"""HNCode grading helpers shared by Flask routes and tests."""

from __future__ import annotations

import csv
import html
import re
import unicodedata
from copy import copy
from pathlib import Path
from typing import Callable


DecodeText = Callable[[bytes], str]


ACCOUNT_COLUMN_ALIASES = {
    "username": {
        "username",
        "user",
        "account",
        "tai khoan",
        "tai_khoan",
        "ten dang nhap",
        "ten_dang_nhap",
        "login",
    },
    "password": {
        "password",
        "pass",
        "mat khau",
        "mat_khau",
        "mk",
    },
    "name": {
        "name",
        "fullname",
        "full name",
        "ho ten",
        "ho_ten",
        "ten",
        "hoc sinh",
        "hoc_sinh",
        "student",
    },
}


def normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def account_fieldnames(fieldnames: list[str] | None) -> dict[str, str]:
    normalized = {normalize_header(name): name for name in fieldnames or []}
    result: dict[str, str] = {}
    for target, aliases in ACCOUNT_COLUMN_ALIASES.items():
        for alias in aliases:
            key = normalize_header(alias)
            if key in normalized:
                result[target] = normalized[key]
                break
    return result


def read_accounts(csv_path: Path, decode_text: DecodeText | None = None) -> list[dict]:
    raw = csv_path.read_bytes()
    text = decode_text(raw) if decode_text else raw.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    fields = account_fieldnames(reader.fieldnames)
    missing = {"username", "password"} - set(fields)
    if missing:
        raise RuntimeError(
            "File tài khoản thiếu cột: "
            + ", ".join(sorted(missing))
            + ". Tool nhận các cột như username/password/name hoặc Tên đăng nhập/Mật khẩu/Họ tên."
        )
    accounts = []
    for index, row in enumerate(reader, 1):
        username = (row.get(fields["username"]) or "").strip()
        password = (row.get(fields["password"]) or "").strip()
        name = (row.get(fields.get("name", "")) or "").strip() if fields.get("name") else ""
        if username and password:
            if not name:
                name = folder_name_from_account(username)
            accounts.append({"index": index, "username": username, "password": password, "name": name})
    if not accounts:
        raise RuntimeError("Không đọc được tài khoản hợp lệ nào trong file CSV. Mỗi dòng cần có ít nhất username và password.")
    return accounts


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def folder_name_from_account(username: str) -> str:
    text = str(username or "").strip()
    for prefix in ("chamthi_", "chamthi-", "chamthi"):
        if text.lower().startswith(prefix):
            return text[len(prefix):].strip("_- ") or text
    return text


def account_folder_key(account: dict) -> str:
    return normalize_key(folder_name_from_account(str(account.get("username") or "")))


def source_root(extract_root: Path) -> Path:
    dirs = [item for item in extract_root.iterdir() if item.is_dir()]
    return dirs[0] if len(dirs) == 1 and any(path.is_file() for path in dirs[0].rglob("*")) else extract_root


def map_problem_code(stem: str, contest_problems: list[dict]) -> str:
    raw = re.sub(r"[^A-Za-z0-9_]+", "", stem).lower()
    codes = [problem["code"] for problem in contest_problems]
    if raw in codes:
        return raw
    for code in codes:
        if code.startswith(raw + "_") or code.split("_", 1)[0] == raw:
            return code
    for code in codes:
        compact = code.replace("_", "")
        if raw in compact or compact in raw:
            return code
    return raw


def collect_submission_files(source_root_path: Path, accounts: list[dict], contest_problems: list[dict]) -> tuple[list[dict], list[str]]:
    account_by_key = {normalize_key(account["name"]): account for account in accounts}
    account_by_folder_key = {account_folder_key(account): account for account in accounts if account_folder_key(account)}
    problem_by_code = {problem["code"]: problem for problem in contest_problems}
    allowed_suffixes = {".cpp", ".cc", ".cxx", ".c", ".py", ".pas"}
    rows: list[dict] = []
    warnings: list[str] = []
    for student_dir in sorted((item for item in source_root_path.iterdir() if item.is_dir()), key=lambda path: path.name.lower()):
        folder_key = normalize_key(student_dir.name)
        account = account_by_key.get(folder_key) or account_by_folder_key.get(folder_key)
        if not account:
            account = {
                "index": 0,
                "username": f"chamthi_{student_dir.name}",
                "password": "",
                "name": student_dir.name,
                "missing": True,
            }
            warnings.append(f"Không tìm thấy tài khoản CSV cho thư mục {student_dir.name}; đã gợi ý username {account['username']}.")
        files = sorted(
            (path for path in student_dir.rglob("*") if path.is_file() and path.suffix.lower() in allowed_suffixes),
            key=lambda path: path.name.lower(),
        )
        if not files:
            warnings.append(f"Thư mục {student_dir.name} không có file code.")
            continue
        for path in files:
            code = map_problem_code(path.stem, contest_problems)
            problem = problem_by_code.get(code)
            rows.append(build_prepare_row(path, source_root_path, account, code, problem))
    if not rows:
        raise RuntimeError("Không tìm thấy file bài làm nào khớp tài khoản trong zip.")
    return rows, warnings


def build_prepare_row(path: Path, source_root_path: Path, account: dict, problem_code: str, problem: dict | None) -> dict:
    relative_path = path.relative_to(source_root_path)
    folder = relative_path.parts[0] if relative_path.parts else ""
    password_missing = bool(account.get("missing"))
    selected = bool(problem) and not password_missing
    status = "Đã chuẩn bị" if problem else "Không khớp bài trong contest"
    if password_missing:
        status = "Thiếu tài khoản trong CSV"
    return {
        "original_key": f"{folder}::{relative_path.as_posix()}",
        "selected": selected,
        "student": account["name"],
        "folder": folder,
        "username": account["username"],
        "password_missing": password_missing,
        "problem": problem_code,
        "problem_title": problem["title"] if problem else "",
        "contest_points": problem["points"] if problem else 0,
        "language": path.suffix.lower().lstrip("."),
        "file": path.name,
        "relative_path": relative_path.as_posix(),
        "local_path": str(path),
        "status": status,
        "submission_url": "",
        "percent": "",
        "score": "",
        "message": "",
    }


def merge_requested_rows(saved_rows: list[dict], requested_rows: list[dict]) -> list[dict]:
    requested = {row.get("original_key"): row for row in requested_rows}
    rows = []
    for base in saved_rows:
        row = dict(base)
        if row["original_key"] in requested:
            update = requested[row["original_key"]]
            row["selected"] = bool(update.get("selected"))
            for field in ("username", "student", "problem"):
                value = str(update.get(field, "")).strip()
                if value:
                    row[field] = value
        rows.append(row)
    return rows


def html_cell_text(fragment: str) -> str:
    text = html.unescape(re.sub(r"<[^>]+>", " ", fragment or ""))
    return re.sub(r"\s+", " ", text).strip()


def number_from_rank_text(value: str) -> float | str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return round(float(normalized), 2)
    except ValueError:
        return text


def parse_ranking_table(page: str) -> tuple[list[dict], list[str]]:
    table_match = re.search(r'<table\b[^>]*id=["\']users-table["\'][^>]*>([\s\S]*?)</table>', page or "", re.I)
    if not table_match:
        return [], []
    table = table_match.group(1)
    thead = re.search(r"<thead\b[^>]*>([\s\S]*?)</thead>", table, re.I)
    problem_codes = []
    if thead:
        for th in re.findall(r"<th\b[^>]*problem-score-col[^>]*>([\s\S]*?)</th>", thead.group(1), re.I):
            code_match = re.search(r'class=["\']problem-code["\'][^>]*>([\s\S]*?)</div>', th, re.I)
            if code_match:
                problem_codes.append(html_cell_text(code_match.group(1)))
    tbody = re.search(r"<tbody\b[^>]*>([\s\S]*?)</tbody>", table, re.I)
    if not tbody:
        return [], problem_codes
    ranking_rows = []
    for tr in re.findall(r"<tr\b[^>]*>([\s\S]*?)</tr>", tbody.group(1), re.I):
        cells = re.findall(r"<td\b([^>]*)>([\s\S]*?)</td>", tr, re.I)
        if len(cells) < 3:
            continue
        rank = html_cell_text(cells[0][1])
        user_cell = cells[1][1]
        username_match = re.search(r'<a\b[^>]*href=["\']/user/[^"\']+["\'][^>]*>([\s\S]*?)</a>', user_cell, re.I)
        fullname_match = re.search(r'class=["\'][^"\']*\bfullname\b[^"\']*["\'][^>]*>([\s\S]*?)</div>', user_cell, re.I)
        participation_match = re.search(r"<sub\b[^>]*>\s*\[([0-9]+)\]\s*</sub>", user_cell, re.I)
        total_cell = cells[2][1]
        total_score = html_cell_text(re.sub(r'<div\b[^>]*class=["\']solving-time["\'][\s\S]*?</div>', "", total_cell, flags=re.I))
        total_time_match = re.search(r'class=["\']solving-time["\'][^>]*>([\s\S]*?)</div>', total_cell, re.I)
        item = {
            "rank": number_from_rank_text(rank),
            "username": html_cell_text(username_match.group(1)) if username_match else html_cell_text(user_cell),
            "fullname": html_cell_text(fullname_match.group(1)) if fullname_match else "",
            "participation": participation_match.group(1) if participation_match else "",
            "total": number_from_rank_text(total_score),
            "time": html_cell_text(total_time_match.group(1)) if total_time_match else "",
            "scores": {},
            "times": {},
        }
        problem_cells = [cell for attrs, cell in cells[3:] if "problem-score-col" in attrs]
        for code, cell in zip(problem_codes, problem_cells):
            score_match = re.search(r"<span\b[^>]*>([\s\S]*?)</span>", cell, re.I)
            time_match = re.search(r'class=["\']solving-time["\'][^>]*>([\s\S]*?)</div>', cell, re.I)
            score_text = html_cell_text(score_match.group(1)) if score_match else html_cell_text(cell)
            item["scores"][code] = number_from_rank_text(score_text)
            item["times"][code] = html_cell_text(time_match.group(1)) if time_match else ""
        ranking_rows.append(item)
    return ranking_rows, problem_codes


def autosize_worksheet(ws) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)


def bold_header(ws) -> None:
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def write_excel(
    rows: list[dict],
    contest_problems: list[dict],
    accounts: list[dict],
    output_path: Path,
    ranking_rows: list[dict] | None = None,
    ranking_problem_codes: list[str] | None = None,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Bang diem"
    problem_codes = list(ranking_problem_codes or []) or [problem["code"] for problem in contest_problems]
    if ranking_rows:
        ws.append(["Rank", "Username", "Họ tên", "Lượt ảo", "Tổng điểm", "Thời gian", *problem_codes])
        for item in ranking_rows:
            ws.append([
                item.get("rank", ""),
                item.get("username", ""),
                item.get("fullname", ""),
                item.get("participation", ""),
                item.get("total", ""),
                item.get("time", ""),
                *[item.get("scores", {}).get(code, "") for code in problem_codes],
            ])
    else:
        ws.append(["STT", "Học sinh", "Username", *problem_codes, "Tổng điểm", "Số bài đã nộp"])
        by_student_problem: dict[tuple[str, str], dict] = {}
        for row in rows:
            key = (row.get("username", ""), row.get("problem", ""))
            current = by_student_problem.get(key)
            if current is None or float(row.get("score") or 0) > float(current.get("score") or 0):
                by_student_problem[key] = row
        for account in accounts:
            total = 0.0
            count = 0
            values = [account["index"], account["name"], account["username"]]
            for code in problem_codes:
                row = by_student_problem.get((account["username"], code))
                score = float(row.get("score") or 0) if row else 0.0
                total += score
                if row and row.get("submission_url"):
                    count += 1
                values.append(round(score, 2))
            values.extend([round(total, 2), count])
            ws.append(values)
    bold_header(ws)
    ws.freeze_panes = "D2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Chi tiet nop bai")
    ws.append(["Học sinh", "Username", "Mã bài", "Tên bài", "Điểm bài", "%", "Điểm quy đổi", "File", "Trạng thái", "Submission", "Thông báo"])
    for row in rows:
        ws.append([
            row.get("student"),
            row.get("username"),
            row.get("problem"),
            row.get("problem_title"),
            row.get("contest_points"),
            row.get("percent"),
            row.get("score"),
            row.get("relative_path"),
            row.get("status"),
            "Mở submission" if row.get("submission_url") else "",
            row.get("message"),
        ])
        if row.get("submission_url"):
            ws.cell(ws.max_row, 10).hyperlink = row["submission_url"]
            ws.cell(ws.max_row, 10).style = "Hyperlink"
    bold_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Danh sach bai")
    ws.append(["Thứ tự", "Mã bài", "Tên bài", "Điểm contest"])
    for problem in contest_problems:
        ws.append([problem["order"], problem["code"], problem["title"], problem["points"]])
    bold_header(ws)
    autosize_worksheet(ws)
    wb.save(output_path)
