#!/usr/bin/env python3
"""Local web UI for preparing, uploading, and transferring OJ problems."""

from __future__ import annotations

import html
import csv
import json
import math
import os
import re
import requests
import shutil
import subprocess
import sys
import time
import unicodedata
import uuid
import zipfile
from collections import Counter, defaultdict
from dataclasses import replace
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import quote, urljoin
from services.quiz import find_image_in_dir, upload_quiz_image
from urllib.request import urlopen
from http.cookies import SimpleCookie

from flask import Flask, Response, jsonify, render_template_string, request, send_file

from services import hncode as hncode_service
from services import jobs as job_service
from services import problem_bundle as bundle_service
from services import problem_upload as upload_service

from transfer_tinhoctre_to_hncode import (
    ProblemInfo,
    checkbox_checked,
    create_hncode_problem,
    destination_problem_exists,
    fetch_source_problem,
    input_value,
    login_hncode,
    selected_option,
    textarea_value,
    upload_hncode_tests,
)
from upload_tinhoctre_batch import (

    GeneratedTests,
    ProblemBundle,
    USER_AGENT,
    clean_statement,
    csrf_token,
    discover_bundles,
    extract_zip,
    form_errors,
    find_named_file,
    generate_tests,
    login as login_tinhoctre_public,
    problem_exists as tinhoctre_problem_exists,
    read_text_smart,
    session as tinhoctre_session,
    statement_body_text,
    submit_solution,
    upload_tests as upload_tinhoctre_tests,
    zip_case_files,
)


ROOT = Path(__file__).resolve().parent
RUNTIME = ROOT / ".runtime"
SAMPLE_TONGHAISO_ZIP = ROOT / "samples" / "bo_mau_1_bai_tonghaiso.zip"
DEFAULT_ZIP = r"E:\Google Drive\Google Drive\1-School\4-KiThi\THT\2026\5Tinh\04-06\tht26_5_bai_files.zip"
QUIZ_TARGETS = {
    "quiz_hncode": {
        "label": "HNCode",
        "base_url": "https://hncode.edu.vn",
        "account_target": "hncode",
        "default_user": "MrTee",
    },
    "quiz_tinhoctre": {
        "label": "TinHocTre",
        "base_url": "https://tinhoctre.vn",
        "account_target": "tinhoctre",
        "default_user": "MrTee",
    },
}

TARGETS = {
    "hnoj": {
        "label": "HNOJ",
        "base_url": "https://hnoj.edu.vn",
        "type_id": "1",
        "group_id": "1",
        "languages": {"C++17": "4", "Pascal": "7", "Python 3": "9", "Scratch": "12"},
        "default_user": "MrTee",
        "test_backend": "dmoj",
    },
    "hncode": {
        "label": "HNCode",
        "base_url": "https://hncode.edu.vn",
        "type_id": "591",
        "group_id": "105",
        "languages": {"C++17": "12", "C++20": "14", "Pascal": "10", "Python 3": "8", "PyPy 3": "16"},
        "default_user": "MrTee",
        "test_backend": "dmoj",
    },
    "tinhoctre": {
        "label": "TinHocTre",
        "base_url": "https://tinhoctre.vn",
        "type_id": "13",
        "group_id": "13",
        "languages": {
            "C++17": "15",
            "C++20": "18",
            "Pascal": "14",
            "Python 3": "8",
            "PyPy 3": "21",
            "Scratch": "23",
        },
        "default_user": "MrTee",
        "test_backend": "dmoj",
    },
}

CONTEST_TARGETS = {
    "contest_hnoj": {
        "label": "HNOJ Contest",
        "base_url": "https://contest.hnoj.edu.vn",
        "default_user": "MrTee",
        "problem_target": "hnoj",
    },
    **TARGETS,
}

PROMPT_GUIDE = """Vá»›i má»—i bÃ i trong danh sÃ¡ch dÆ°á»›i Ä‘Ã¢y, hÃ£y táº¡o Ä‘á»§ 4 file:

1. File sinh test:
   - TÃªn file: gentest_<ma_bai>.py
   - VÃ­ dá»¥: gentest_tht26_tongbi.py

2. File lá»i giáº£i Python:
   - TÃªn file: sol_<ma_bai>.py
   - VÃ­ dá»¥: sol_tht26_tongbi.py

3. File lá»i giáº£i C++:
   - TÃªn file: sol_<ma_bai>.cpp
   - VÃ­ dá»¥: sol_tht26_tongbi.cpp

4. File Ä‘á» bÃ i Markdown:
   - TÃªn file: <ma_bai>.md
   - VÃ­ dá»¥: tht26_tongbi.md
   - DÃ²ng Ä‘áº§u tiÃªn cá»§a file pháº£i cÃ³ Ä‘Ãºng cáº¥u trÃºc:
     TÃªn bÃ i | MÃ£ bÃ i
   - VÃ­ dá»¥:
     Tá»•ng bi | tht26_tongbi
   - Sau dÃ²ng Ä‘áº§u tiÃªn lÃ  toÃ n bá»™ ná»™i dung Ä‘á» bÃ i.

YÃªu cáº§u Ä‘á»‘i vá»›i file sinh test:

- File sinh test lÃ  file Python.
- Trong file sinh test pháº£i nhÃºng lá»i giáº£i chuáº©n báº±ng C++ Ä‘á»ƒ sinh output.
- Khi cháº¡y file sinh test, chÆ°Æ¡ng trÃ¬nh tá»± táº¡o thÆ° má»¥c test cho bÃ i tÆ°Æ¡ng á»©ng.
- TÃªn thÆ° má»¥c test nÃªn lÃ  mÃ£ bÃ i, vÃ­ dá»¥:
  tht26_tongbi/
- CÃ¡c file test trong thÆ° má»¥c cÃ³ dáº¡ng:
  01.inp, 01.out
  02.inp, 02.out
  ...
- Sau khi sinh test, file sinh test tá»± nÃ©n thÆ° má»¥c test thÃ nh:
  tht26_tongbi.zip

YÃªu cáº§u Ä‘á»‘i vá»›i bá»™ test:

- Bá»™ test pháº£i Ä‘á»§ máº¡nh, phá»§ Ä‘á»§ cÃ¡c trÆ°á»ng há»£p Ä‘áº·c biá»‡t vÃ  trÆ°á»ng há»£p biÃªn.
- Dá»¯ liá»‡u pháº£i Ä‘Ãºng giá»›i háº¡n cá»§a Ä‘á» bÃ i.
- Náº¿u Ä‘á» cÃ³ subtask, sá»‘ lÆ°á»£ng test pháº£i phÃ¢n bá»‘ Ä‘Ãºng theo tá»‰ lá»‡ subtask.
- Náº¿u bÃ i Ä‘Æ¡n giáº£n, chá»‰ cáº§n khoáº£ng 10 test.
- Náº¿u bÃ i cáº§n nhiá»u trÆ°á»ng há»£p Ä‘á»ƒ kiá»ƒm tra cháº·t cháº½ hÆ¡n, cÃ³ thá»ƒ sinh khoáº£ng 20 test hoáº·c nhiá»u hÆ¡n.
- Cáº§n cÃ³ 01 test vÃ­ dá»¥, cÃ¡c test nhá», test biÃªn, test ngáº«u nhiÃªn cÃ³ kiá»ƒm soÃ¡t, test Ä‘á»§ cÃ¡c trÆ°á»ng há»£p vÃ  test lá»›n.

Sau khi táº¡o xong, hÃ£y nÃ©n toÃ n bá»™ cÃ¡c file Ä‘Ã£ táº¡o thÃ nh má»™t file zip duy nháº¥t vÃ  gá»­i láº¡i cho tÃ´i.

VÃ­ dá»¥ vá»›i bÃ i:

Tá»•ng bi | tht26_tongbi

Cáº§n táº¡o 4 file:

- gentest_tht26_tongbi.py
- sol_tht26_tongbi.py
- sol_tht26_tongbi.cpp
- tht26_tongbi.md

HÃ£y thá»±c hiá»‡n cho toÃ n bá»™ cÃ¡c bÃ i Ä‘Æ°á»£c cung cáº¥p bÃªn dÆ°á»›i."""

QUIZ_FORMAT_GUIDE = """# Format soáº¡n danh sÃ¡ch quiz

Má»—i cÃ¢u há»i tÃ¡ch nhau báº±ng má»™t dÃ²ng chá»‰ gá»“m `---`.

CÃ¡c loáº¡i há»£p lá»‡:
- `MC` hoáº·c `Tráº¯c nghiá»‡m 1 Ä‘Ã¡p Ã¡n`
- `MA` hoáº·c `Tráº¯c nghiá»‡m nhiá»u Ä‘Ã¡p Ã¡n`
- `SA` hoáº·c `Tráº£ lá»i ngáº¯n`
- `FB` hoáº·c `Äiá»n vÃ o chá»— trá»‘ng`
- `TF` hoáº·c `ÄÃºng / Sai`

Máº«u:

Loáº¡i: MC
TiÃªu Ä‘á»: CÃ¢u há»i vÃ­ dá»¥ 1
Ná»™i dung:
Trong Python, hÃ m nÃ o dÃ¹ng Ä‘á»ƒ in ra mÃ n hÃ¬nh?
Lá»±a chá»n:
- A. input()
- B. print()
- C. len()
- D. range()
ÄÃ¡p Ã¡n: B
Giáº£i thÃ­ch:
`print()` dÃ¹ng Ä‘á»ƒ in dá»¯ liá»‡u ra mÃ n hÃ¬nh.
---
Loáº¡i: MA
TiÃªu Ä‘á»: Sá»‘ nguyÃªn tá»‘
Ná»™i dung:
Nhá»¯ng sá»‘ nÃ o sau Ä‘Ã¢y lÃ  sá»‘ nguyÃªn tá»‘?
Lá»±a chá»n:
- A. 2
- B. 3
- C. 4
- D. 9
ÄÃ¡p Ã¡n: A, B
---
Loáº¡i: SA
TiÃªu Ä‘á»: Káº¿t quáº£ phÃ©p tÃ­nh
Ná»™i dung:
TÃ­nh 6 * 7.
ÄÃ¡p Ã¡n:
- 42
- bá»‘n mÆ°Æ¡i hai
---
Loáº¡i: FB
TiÃªu Ä‘á»: Äiá»n vÃ o chá»— trá»‘ng
Ná»™i dung:
An vÃ  BÃ¬nh cÃ³ $5$ viÃªn bi. An cÃ³ hÆ¡n BÃ¬nh Ä‘Ãºng $1$ viÃªn bi.
Váº­y An cÃ³ \\_\\_\\_(1)\\_\\_\\_ viÃªn bi vÃ  BÃ¬nh cÃ³ \\_\\_\\_(2)\\_\\_\\_ viÃªn bi.
ÄÃ¡p Ã¡n:
- Sá»‘ bi cá»§a An: 3
- Sá»‘ bi cá»§a BÃ¬nh: 2
---
Loáº¡i: FB
TiÃªu Ä‘á»: Nhiá»u cÃ¡ch nháº­p Ä‘Ã¡p Ã¡n
Ná»™i dung:
Äiá»n káº¿t quáº£ Ä‘Ãºng vÃ o hai chá»— trá»‘ng:
$2 + 3 =$ \\_\\_\\_(1)\\_\\_\\_ vÃ  tÃªn ngÃ´n ngá»¯ láº­p trÃ¬nh Python viáº¿t thÆ°á»ng lÃ  \\_\\_\\_(2)\\_\\_\\_.
ÄÃ¡p Ã¡n:
- Ã” 1: 5 | nÄƒm
- Ã” 2: python
Giáº£i thÃ­ch:
Má»—i dÃ²ng Ä‘Ã¡p Ã¡n tÆ°Æ¡ng á»©ng má»™t Ã´ trá»‘ng. CÃ¡c Ä‘Ã¡p Ã¡n Ä‘Ãºng thay tháº¿ cho cÃ¹ng má»™t Ã´ cÃ³ thá»ƒ ngÄƒn báº±ng dáº¥u `|`, `,` hoáº·c `;`.
---
Loáº¡i: TF
TiÃªu Ä‘á»: ÄÃºng sai
Ná»™i dung:
Python lÃ  má»™t ngÃ´n ngá»¯ láº­p trÃ¬nh.
ÄÃ¡p Ã¡n: ÄÃºng

Ghi chÃº:
- NhÃ£n quiz Ä‘á»ƒ trá»‘ng.
- `XÃ¡o trá»™n lá»±a chá»n` vÃ  `CÃ´ng khai` chá»n trÃªn giao diá»‡n tool.
- Vá»›i cÃ¢u `FB`, trong `Ná»™i dung` Ä‘Ã¡nh dáº¥u Ã´ trá»‘ng theo dáº¡ng `\\_\\_\\_(1)\\_\\_\\_`, `\\_\\_\\_(2)\\_\\_\\_`, ... Ä‘á»ƒ há»‡ thá»‘ng nháº­n Ä‘Ãºng vá»‹ trÃ­ cáº§n Ä‘iá»n.
- Vá»›i cÃ¢u `FB`, má»—i dÃ²ng Ä‘Ã¡p Ã¡n cÃ³ dáº¡ng `NhÃ£n: Ä‘Ã¡p Ã¡n`, vÃ­ dá»¥ `Ã” 1: 5 | nÄƒm`. Náº¿u khÃ´ng ghi nhÃ£n, tool tá»± Ä‘áº·t `Ã” 1:`, `Ã” 2:`, ...
- Vá»›i cÃ¢u `FB`, nhiá»u Ä‘Ã¡p Ã¡n Ä‘Ãºng cho cÃ¹ng má»™t Ã´ cÃ³ thá»ƒ ngÄƒn báº±ng `|`, `,` hoáº·c `;`. Tool cháº¥m khÃ´ng phÃ¢n biá»‡t hoa/thÆ°á»ng.
- Vá»›i cÃ¢u `TF`, tool tá»± táº¡o hai lá»±a chá»n `ÄÃºng` vÃ  `Sai`.
"""

app = Flask(__name__)
PROGRESS_DIR = RUNTIME / "progress"
prepared_uploads: dict[str, dict] = {}
prepared_single_uploads: dict[str, dict] = {}
prepared_transfers: dict[str, dict] = {}
prepared_contest_transfers: dict[str, dict] = {}
prepared_quizzes: dict[str, dict] = {}
prepared_lesson_copies: dict[str, dict] = {}
prepared_course_clones: dict[str, dict] = {}
prepared_hncode_grading: dict[str, dict] = {}


class ProblemAlreadyExists(RuntimeError):
    pass


class ContestAlreadyExists(RuntimeError):
    pass


QUESTION_TYPE_ALIASES = {
    "mc": "MC",
    "trac nghiem 1 dap an": "MC",
    "trac nghiem mot dap an": "MC",
    "tráº¯c nghiá»‡m 1 Ä‘Ã¡p Ã¡n": "MC",
    "tráº¯c nghiá»‡m má»™t Ä‘Ã¡p Ã¡n": "MC",
    "ma": "MA",
    "trac nghiem nhieu dap an": "MA",
    "tráº¯c nghiá»‡m nhiá»u Ä‘Ã¡p Ã¡n": "MA",
    "sa": "SA",
    "tra loi ngan": "SA",
    "tráº£ lá»i ngáº¯n": "SA",
    "fb": "FB",
    "dien vao cho trong": "FB",
    "Ä‘iá»n vÃ o chá»— trá»‘ng": "FB",
    "fill blank": "FB",
    "fill in blank": "FB",
    "fill in the blank": "FB",
    "tf": "TF",
    "dung sai": "TF",
    "dung / sai": "TF",
    "Ä‘Ãºng sai": "TF",
    "Ä‘Ãºng / sai": "TF",
}

HNCODE_TYPE_ALIASES = {
    "binary search": "198",
    "binary-search": "198",
    "binary_search": "198",
    "sortings": "188",
    "sorting": "188",
    "sort": "188",
    "dp": "172",
    "dynamic programming": "172",
    "quy hoach dong": "172",
    "quy hoáº¡ch Ä‘á»™ng": "172",
    "two pointers": "196",
    "two-pointers": "196",
    "two_pointers": "196",
    "2 pointers": "196",
    "implementation": "340",
    "cai dat": "340",
    "cÃ i Ä‘áº·t": "340",
    "math": "175",
    "toan": "175",
    "toÃ¡n": "175",
    "greedy": "171",
    "tham lam": "171",
    "strings": "176",
    "string": "176",
    "chuoi": "176",
    "chuá»—i": "176",
}

QUIZ_FIELD_ALIASES = {
    "loáº¡i": "type",
    "loai": "type",
    "type": "type",
    "tiÃªu Ä‘á»": "title",
    "tieu de": "title",
    "title": "title",
    "ná»™i dung": "content",
    "noi dung": "content",
    "content": "content",
    "lá»±a chá»n": "choices",
    "lua chon": "choices",
    "choices": "choices",
    "Ä‘Ã¡p Ã¡n": "answer",
    "dap an": "answer",
    "answer": "answer",
    "answers": "answer",
    "giáº£i thÃ­ch": "explanation",
    "giai thich": "explanation",
    "explanation": "explanation",
}


def normalize_key_text(value: str) -> str:
    value = value.strip().lower()
    replacements = {
        "Ã¡Ã áº£Ã£áº¡Äƒáº¯áº±áº³áºµáº·Ã¢áº¥áº§áº©áº«áº­": "a",
        "Ã©Ã¨áº»áº½áº¹Ãªáº¿á»á»ƒá»…á»‡": "e",
        "Ã­Ã¬á»‰Ä©á»‹": "i",
        "Ã³Ã²á»Ãµá»Ã´á»‘á»“á»•á»—á»™Æ¡á»›á»á»Ÿá»¡á»£": "o",
        "ÃºÃ¹á»§Å©á»¥Æ°á»©á»«á»­á»¯á»±": "u",
        "Ã½á»³á»·á»¹á»µ": "y",
        "Ä‘": "d",
    }
    for chars, repl in replacements.items():
        for ch in chars:
            value = value.replace(ch, repl)
    value = re.sub(r"\s+", " ", value)
    return value


def quiz_field_from_line(line: str) -> tuple[str, str] | None:
    match = re.match(r"^\s*([^:ï¼š]{1,40})\s*[:ï¼š]\s*(.*)$", line)
    if not match:
        return None
    raw_key = match.group(1).strip()
    key = QUIZ_FIELD_ALIASES.get(raw_key.lower()) or QUIZ_FIELD_ALIASES.get(normalize_key_text(raw_key))
    if not key:
        return None
    return key, match.group(2).strip()


def split_quiz_blocks(text: str) -> list[str]:
    blocks: list[list[str]] = [[]]
    for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if line.strip() == "---":
            if blocks[-1]:
                blocks.append([])
            continue
        blocks[-1].append(line)
    return ["\n".join(block).strip() for block in blocks if "\n".join(block).strip()]


def parse_choice_lines(text: str) -> list[dict]:
    choices = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(r"^(?:[-*]\s*)?([A-Za-z0-9]+)\s*[\.\):ï¼š-]\s*(.+)$", line)
        if not match:
            raise RuntimeError(f"Lá»±a chá»n khÃ´ng Ä‘Ãºng dáº¡ng `- A. Ná»™i dung`: {line}")
        choices.append({"id": match.group(1).strip(), "text": match.group(2).strip()})
    return choices


def split_answers(text: str) -> list[str]:
    parts: list[str] = []
    for line in text.splitlines():
        line = re.sub(r"^\s*[-*]\s*", "", line).strip()
        if not line:
            continue
        if "," in line or ";" in line or "|" in line:
            parts.extend(item.strip() for item in re.split(r"[,;|]", line) if item.strip())
        else:
            parts.append(line)
    return parts


def parse_fill_blank_answers(text: str) -> list[dict]:
    blanks: list[dict] = []
    for raw_line in text.splitlines():
        line = re.sub(r"^\s*[-*]\s*", "", raw_line).strip()
        if not line:
            continue
        label = f"Ã” {len(blanks) + 1}:"
        answer_text = line
        match = re.match(r"^(.{1,80}?)\s*[:ï¼š]\s*(.+)$", line)
        if match:
            label = match.group(1).strip()
            if not label.endswith(":"):
                label += ":"
            answer_text = match.group(2).strip()
        answers = [item.strip() for item in re.split(r"[,;|]", answer_text) if item.strip()]
        if not answers:
            raise RuntimeError(f"ÄÃ¡p Ã¡n Ä‘iá»n vÃ o chá»— trá»‘ng chÆ°a há»£p lá»‡: {raw_line}")
        blanks.append({"label": label, "answers": answers})
    return blanks


def parse_quiz_markdown(text: str) -> list[dict]:
    items = []
    for index, block in enumerate(split_quiz_blocks(text), 1):
        fields = {"type": "", "title": "", "content": "", "choices": "", "answer": "", "explanation": ""}
        current: str | None = None
        for line in block.splitlines():
            parsed = quiz_field_from_line(line)
            if parsed:
                current, value = parsed
                fields[current] = value
                continue
            if current:
                fields[current] = (fields[current] + "\n" + line).strip("\n")
        qtype = QUESTION_TYPE_ALIASES.get(fields["type"].strip().lower()) or QUESTION_TYPE_ALIASES.get(normalize_key_text(fields["type"]))
        if not qtype:
            raise RuntimeError(f"CÃ¢u {index}: Loáº¡i cÃ¢u há»i khÃ´ng há»£p lá»‡: {fields['type']!r}")
        content = fields["content"].strip()
        if not content:
            raise RuntimeError(f"CÃ¢u {index}: thiáº¿u Ná»™i dung.")
        title = fields["title"].strip() or re.sub(r"\s+", " ", content)[:80] or f"CÃ¢u há»i {index}"
        choices = parse_choice_lines(fields["choices"]) if fields["choices"].strip() else []
        answers = split_answers(fields["answer"])
        if qtype in {"MC", "MA"}:
            if not choices:
                raise RuntimeError(f"CÃ¢u {index}: cÃ¢u tráº¯c nghiá»‡m cáº§n cÃ³ Lá»±a chá»n.")
            if not answers:
                raise RuntimeError(f"CÃ¢u {index}: cÃ¢u tráº¯c nghiá»‡m cáº§n cÃ³ ÄÃ¡p Ã¡n.")
            valid_ids = {choice["id"] for choice in choices}
            missing = [answer for answer in answers if answer not in valid_ids]
            if missing:
                raise RuntimeError(f"CÃ¢u {index}: Ä‘Ã¡p Ã¡n {', '.join(missing)} khÃ´ng cÃ³ trong lá»±a chá»n.")
            correct = {"answers": answers if qtype == "MA" else (answers[0] if answers else "")}
        elif qtype == "SA":
            if not answers:
                raise RuntimeError(f"CÃ¢u {index}: cÃ¢u tráº£ lá»i ngáº¯n cáº§n cÃ³ Ã­t nháº¥t má»™t ÄÃ¡p Ã¡n.")
            choices = []
            correct = {"type": "exact", "answers": answers, "case_sensitive": False}
            grading_strategy = "all_or_nothing"
        elif qtype == "FB":
            blanks = parse_fill_blank_answers(fields["answer"])
            if not blanks:
                raise RuntimeError(f"CÃ¢u {index}: cÃ¢u Ä‘iá»n vÃ o chá»— trá»‘ng cáº§n cÃ³ Ã­t nháº¥t má»™t dÃ²ng ÄÃ¡p Ã¡n.")
            choices = None
            correct = {"type": "exact", "case_sensitive": False, "blanks": blanks}
            grading_strategy = "correct_only"
        else:
            if not choices:
                choices = [{"id": "T", "text": "ÄÃºng"}, {"id": "F", "text": "Sai"}]
            if not answers:
                raise RuntimeError(f"CÃ¢u {index}: cÃ¢u ÄÃºng/Sai cáº§n cÃ³ ÄÃ¡p Ã¡n.")
            raw = normalize_key_text(answers[0] if answers else "")
            correct_id = "T" if raw in {"dung", "true", "t", "1", "yes"} else "F" if raw in {"sai", "false", "f", "0", "no"} else answers[0] if answers else ""
            if correct_id not in {choice["id"] for choice in choices}:
                raise RuntimeError(f"CÃ¢u {index}: Ä‘Ã¡p Ã¡n ÄÃºng/Sai pháº£i lÃ  ÄÃºng hoáº·c Sai.")
            correct = {"answers": correct_id}
            grading_strategy = "all_or_nothing"
        if qtype in {"MC", "MA"}:
            grading_strategy = "all_or_nothing"
        items.append(
            {
                "index": index,
                "type": qtype,
                "title": title,
                "content": content,
                "choices": choices,
                "correct_answers": correct,
                "grading_strategy": grading_strategy,
                "explanation": fields["explanation"].strip(),
            }
        )
    if not items:
        raise RuntimeError("ChÆ°a cÃ³ cÃ¢u há»i nÃ o trong ná»™i dung quiz.")
    return items


def prepare_quiz_items(text: str) -> tuple[list[dict], list[dict]]:
    rows = []
    valid_questions = []
    blocks = split_quiz_blocks(text)
    if not blocks:
        raise RuntimeError("ChÆ°a cÃ³ cÃ¢u há»i nÃ o trong ná»™i dung quiz.")
    for index, block in enumerate(blocks, 1):
        try:
            question = parse_quiz_markdown(block)[0]
            question["index"] = index
            valid_questions.append(question)
            rows.append(
                {
                    "index": index,
                    "title": question["title"],
                    "type": question["type"],
                    "status": "âœ“ Há»£p lá»‡",
                    "error": "",
                    "can_upload": True,
                }
            )
        except Exception as exc:
            rows.append(
                {
                    "index": index,
                    "title": f"CÃ¢u {index}",
                    "type": "",
                    "status": "âœ— Lá»—i",
                    "error": str(exc),
                    "can_upload": False,
                }
            )
    return valid_questions, rows


def quiz_target_info(target: str) -> dict:
    if target not in QUIZ_TARGETS:
        raise RuntimeError("Web up quiz khÃ´ng há»£p lá»‡.")
    return QUIZ_TARGETS[target]


def login_quiz_target(target: str, account: dict) -> requests.Session:
    info = quiz_target_info(target)
    username = account.get("username", "")
    password = account.get("password", "")
    try:
        session = login_tinhoctre_public(info["base_url"], username, password, "/quiz/questions/create/")
    except Exception as exc:
        raise RuntimeError(f"{info['label']} quiz login failed: {exc}") from exc
    create_url = urljoin(info["base_url"], "/quiz/questions/create/")
    page = session.get(create_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"{info['label']} quiz form failed: HTTP {page.status_code}")
    if "/accounts/login" in page.url or "/admin/login" in page.url:
        raise RuntimeError(f"{info['label']} quiz login did not open create form.")
    return session


def create_quiz_question(session, base_url: str, question: dict, *, shuffle_choices: bool, is_public: bool) -> str:
    create_url = urljoin(base_url, "/quiz/questions/create/")
    page = session.get(create_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form táº¡o quiz: HTTP {page.status_code}")
    if "/accounts/login" in page.url or "/admin/login" in page.url:
        raise RuntimeError("Session quiz Ä‘Ã£ háº¿t háº¡n hoáº·c tÃ i khoáº£n khÃ´ng cÃ³ quyá»n má»Ÿ form táº¡o quiz.")
    data = {
        "csrfmiddlewaretoken": csrf_token(page.text),
        "title": question["title"],
        "question_type": question["type"],
        "content": question["content"],
        "choices": json.dumps(question["choices"], ensure_ascii=False),
        "correct_answers": json.dumps(question["correct_answers"], ensure_ascii=False),
        "grading_strategy": question.get("grading_strategy") or "all_or_nothing",
        "tags": "",
        "explanation": question.get("explanation", ""),
    }
    if shuffle_choices and question.get("type") in {"MC", "MA", "TF"}:
        data["shuffle_choices"] = "on"
    if is_public:
        data["is_public"] = "on"
    result = session.post(create_url, data=data, headers={"Referer": create_url}, allow_redirects=True, timeout=30)
    if not result.ok:
        raise RuntimeError(f"Táº¡o quiz lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text)
    if errors:
        raise RuntimeError("Form táº¡o quiz bÃ¡o lá»—i: " + "; ".join(errors))
    match = re.search(r"/quiz/questions/(\d+)/", result.url)
    if not match:
        raise RuntimeError(f"Táº¡o quiz chÆ°a tráº£ vá» trang cÃ¢u há»i: {result.url}")
    return urljoin(base_url, f"/quiz/questions/{match.group(1)}/")


def valid_progress_id(progress_id: str | None) -> str | None:
    return job_service.valid_job_id(progress_id)


def progress_path(progress_id: str) -> Path:
    return job_service.job_path(PROGRESS_DIR, progress_id)


def progress_update(progress_id: str | None, **payload) -> None:
    job_service.update_job(PROGRESS_DIR, progress_id, **payload)


def progress_finish(progress_id: str | None, ok: bool, message: str = "") -> None:
    job_service.finish_job(PROGRESS_DIR, progress_id, ok, message)


def safe_extract_zip(zip_path: Path, dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    dest_resolved = dest.resolve()
    with zipfile.ZipFile(zip_path) as zf:
        for member in zf.infolist():
            target = (dest / member.filename).resolve()
            if dest_resolved not in target.parents and target != dest_resolved:
                raise RuntimeError(f"File zip cÃ³ Ä‘Æ°á»ng dáº«n khÃ´ng an toÃ n: {member.filename}")
        zf.extractall(dest)


def scratch_submission_score(student_dir: Path) -> int:
    history = student_dir / "$History"
    score = 0
    if history.is_dir():
        score += sum(1 for item in history.iterdir() if item.is_file() and item.suffix.lower() == ".sb3") * 2
    score += sum(1 for item in student_dir.iterdir() if item.is_file() and item.suffix.lower() == ".sb3")
    return score


def find_scratch_data_root(extract_root: Path) -> Path:
    candidates = [extract_root]
    candidates.extend(item for item in extract_root.iterdir() if item.is_dir())
    best = extract_root
    best_score = -1
    for candidate in candidates:
        score = sum(1 for child in candidate.iterdir() if child.is_dir() and scratch_submission_score(child) > 0)
        if score > best_score:
            best = candidate
            best_score = score
    return best


def history_version(path: Path) -> int:
    match = re.search(r"_(\d+)\.sb3$", path.name, re.IGNORECASE)
    return int(match.group(1)) if match else -1


def get_last_scratch_submission(student_dir: Path) -> Path | None:
    history = student_dir / "$History"
    if history.is_dir():
        history_files = [item for item in history.iterdir() if item.is_file() and item.suffix.lower() == ".sb3"]
        if history_files:
            return max(history_files, key=history_version)
    root_files = [item for item in student_dir.iterdir() if item.is_file() and item.suffix.lower() == ".sb3"]
    if root_files:
        return sorted(root_files, key=lambda item: item.name.lower())[0]
    return None


def collect_last_scratch_submissions(data_root: Path, output_dir: Path) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for student_dir in sorted((item for item in data_root.iterdir() if item.is_dir()), key=lambda item: item.name.lower()):
        last_file = get_last_scratch_submission(student_dir)
        row = {"student_id": student_dir.name, "source": "", "output": "", "status": "missing"}
        if last_file:
            output_name = f"{student_dir.name}.sb3"
            output_path = output_dir / output_name
            shutil.copy2(last_file, output_path)
            row.update({"source": last_file.relative_to(data_root).as_posix(), "output": output_name, "status": "ok"})
        rows.append(row)
    report_lines = [
        "student_id\tstatus\toutput_file\tsource_file",
        *[f"{row['student_id']}\t{row['status']}\t{row['output']}\t{row['source']}" for row in rows],
    ]
    (output_dir / "report.txt").write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    found = sum(1 for row in rows if row["status"] == "ok")
    return {"rows": rows, "total": len(rows), "found": found, "missing": len(rows) - found}


CODE_EXTENSIONS = {".py", ".cpp", ".cc", ".cxx", ".c", ".pas", ".java"}
AI_SOURCE_DEFAULT = r"E:\Google Drive\Google Drive\1-School\4-KiThi\THT\2026\TW\KV\Data"


def normalize_contest_name(zip_path: Path) -> str:
    name = zip_path.stem
    return re.sub(r"[-_]?data$", "", name, flags=re.IGNORECASE)


def code_problem_from_name(name: str) -> str:
    stem = Path(name).stem
    return re.sub(r"_\d+$", "", stem)


def history_version_from_name(name: str) -> int:
    match = re.search(r"_(\d+)(?:\.[^.]+)$", name)
    return int(match.group(1)) if match else 10**18


def read_zip_text(zf: zipfile.ZipFile, member: str) -> str:
    raw = zf.read(member)
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def strip_code_comments(text: str, ext: str) -> str:
    if ext == ".py":
        text = re.sub(r'("""[\s\S]*?"""|\'\'\'[\s\S]*?\'\'\')', " ", text)
        return re.sub(r"#.*", " ", text)
    if ext in {".cpp", ".cc", ".cxx", ".c", ".java"}:
        text = re.sub(r"/\*[\s\S]*?\*/", " ", text)
        return re.sub(r"//.*", " ", text)
    if ext == ".pas":
        text = re.sub(r"\{[\s\S]*?\}", " ", text)
        text = re.sub(r"\(\*[\s\S]*?\*\)", " ", text)
        return re.sub(r"//.*", " ", text)
    return text


def comment_line_count(text: str, ext: str) -> int:
    lines = text.splitlines()
    count = sum(1 for line in lines if line.strip().startswith(("#", "//", "{", "(*")))
    if ext in {".cpp", ".cc", ".cxx", ".c", ".java"}:
        count += len(re.findall(r"/\*[\s\S]*?\*/", text))
    if ext == ".py":
        count += len(re.findall(r'("""[\s\S]*?"""|\'\'\'[\s\S]*?\'\'\')', text))
    return count


def code_identifiers(cleaned: str, ext: str) -> list[str]:
    words = re.findall(r"[A-Za-z_][A-Za-z0-9_]*", cleaned)
    keywords = {
        "include", "using", "namespace", "std", "int", "long", "double", "float", "char", "bool", "return", "for",
        "while", "if", "else", "elif", "def", "class", "import", "from", "as", "in", "and", "or", "not", "True",
        "False", "None", "void", "const", "auto", "vector", "map", "set", "dict", "list", "str", "range", "cin",
        "cout", "begin", "end", "var", "procedure", "function", "then", "do",
    }
    return [word for word in words if word not in keywords and not word.isupper()]


def compact_style_bucket(features: dict) -> str:
    if features["ext"] == ".py":
        if features["line_count"] <= 8:
            size = "py-compact"
        elif features["function_count"] >= 2 or features["import_count"] >= 3:
            size = "py-structured"
        else:
            size = "py-simple"
        io = "fastio" if features["fast_io"] else "plainio"
    elif features["ext"] in {".cpp", ".cc", ".cxx", ".c"}:
        if features["macro_count"] >= 4 or features["using_alias_count"] >= 3:
            size = "cpp-template"
        elif features["line_count"] <= 35:
            size = "cpp-short"
        else:
            size = "cpp-plain"
        io = "bits" if features["include_bits"] else "nobits"
    else:
        size = features["ext"].lstrip(".")
        io = "plain"
    return f"{size}/{io}/c{min(features['comment_ratio_bucket'], 4)}/id{min(features['identifier_bucket'], 4)}"


def analyze_code_text(text: str, ext: str) -> dict:
    lines = text.splitlines()
    nonempty = [line for line in lines if line.strip()]
    cleaned = strip_code_comments(text, ext)
    identifiers = code_identifiers(cleaned, ext)
    long_ids = [item for item in identifiers if len(item) >= 10]
    comment_count = comment_line_count(text, ext)
    line_count = max(len(lines), 1)
    comment_ratio = comment_count / line_count
    avg_line_len = sum(len(line) for line in nonempty) / max(len(nonempty), 1)
    ai_phrases = [
        "complexity", "approach", "edge case", "edge cases", "initialize", "iterate", "we need", "we can",
        "let's", "this function", "read input", "print result", "return answer", "time complexity",
        "space complexity", "base case", "recursive", "dynamic programming", "greedy approach",
    ]
    lower = text.lower()
    phrase_hits = [phrase for phrase in ai_phrases if phrase in lower]
    function_count = len(re.findall(r"\bdef\s+\w+\s*\(", text)) + len(re.findall(r"\b[a-zA-Z_][\w:<>,\s*&]*\s+\w+\s*\([^;{}]*\)\s*\{", text))
    features = {
        "ext": ext,
        "line_count": len(lines),
        "char_count": len(text),
        "avg_line_len": round(avg_line_len, 1),
        "comment_ratio": round(comment_ratio, 3),
        "comment_ratio_bucket": int(min(comment_ratio * 12, 9)),
        "blank_ratio": round((line_count - len(nonempty)) / line_count, 3),
        "macro_count": len(re.findall(r"^\s*#\s*define\b", text, re.MULTILINE)),
        "include_count": len(re.findall(r"^\s*#\s*include\b", text, re.MULTILINE)),
        "include_bits": bool(re.search(r"#\s*include\s*<bits/stdc\+\+\.h>", text)),
        "using_alias_count": len(re.findall(r"\b(using|typedef)\b", text)),
        "import_count": len(re.findall(r"^\s*(import|from)\s+", text, re.MULTILINE)),
        "function_count": function_count,
        "class_count": len(re.findall(r"\bclass\s+\w+", text)),
        "fast_io": bool(re.search(r"ios::sync_with_stdio|cin\.tie|sys\.stdin|stdin\.read|readline", text)),
        "long_identifier_ratio": round(len(long_ids) / max(len(identifiers), 1), 3),
        "identifier_bucket": int(min((len(long_ids) / max(len(identifiers), 1)) * 10, 9)),
        "avg_identifier_len": round(sum(len(item) for item in identifiers) / max(len(identifiers), 1), 2),
        "ai_phrase_hits": phrase_hits,
    }
    reasons = []
    score = 0
    if phrase_hits:
        score += min(24, 8 + 4 * len(phrase_hits))
        reasons.append("CÃ³ chÃº thÃ­ch/cá»¥m tá»« giáº£i thÃ­ch kiá»ƒu AI: " + ", ".join(phrase_hits[:4]))
    if comment_ratio >= 0.18 and len(lines) >= 25:
        score += 12
        reasons.append("Tá»‰ lá»‡ chÃº thÃ­ch cao báº¥t thÆ°á»ng")
    if features["long_identifier_ratio"] >= 0.22 and len(identifiers) >= 20:
        score += 10
        reasons.append("Nhiá»u tÃªn biáº¿n/hÃ m dÃ i, mÃ´ táº£ ráº¥t chuáº©n")
    if ext == ".py" and features["function_count"] >= 2 and features["import_count"] >= 3 and len(lines) >= 35:
        score += 10
        reasons.append("Python cÃ³ cáº¥u trÃºc/import khÃ¡ cÃ´ng nghiá»‡p")
    if ext in {".cpp", ".cc", ".cxx", ".c"} and features["macro_count"] >= 6 and features["using_alias_count"] >= 4:
        score += 8
        reasons.append("C++ dÃ¹ng template/macro dÃ y")
    if features["class_count"] >= 1 and len(lines) >= 45:
        score += 6
        reasons.append("CÃ³ class/cáº¥u trÃºc lá»›n so vá»›i bÃ i thi láº­p trÃ¬nh phá»• thÃ´ng")
    features["code_ai_score"] = min(score, 45)
    features["code_reasons"] = reasons
    features["style_bucket"] = compact_style_bucket(features)
    return features


def vector_distance(a: dict, b: dict) -> float:
    if not a or not b:
        return 0.0
    dist = 0.0
    if a["ext"] != b["ext"]:
        dist += 35
    if a["style_bucket"] != b["style_bucket"]:
        dist += 18
    numeric = [
        ("line_count", 80, 10),
        ("comment_ratio", 0.25, 10),
        ("macro_count", 8, 7),
        ("import_count", 6, 7),
        ("function_count", 6, 7),
        ("avg_identifier_len", 8, 6),
        ("long_identifier_ratio", 0.35, 8),
    ]
    for key, scale, weight in numeric:
        dist += min(abs(float(a.get(key, 0)) - float(b.get(key, 0))) / scale, 1.0) * weight
    return round(min(dist, 100), 1)


def normalized_code_tokens(text: str, ext: str) -> list[str]:
    cleaned = strip_code_comments(text, ext).lower()
    cleaned = re.sub(r'"(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\'', " STR ", cleaned)
    raw_tokens = re.findall(r"[a-z_][a-z0-9_]*|\d+(?:\.\d+)?|==|!=|<=|>=|\+\+|--|&&|\|\||[+\-*/%<>=(){}\[\],.;:]", cleaned)
    keywords = {
        "if", "else", "elif", "for", "while", "do", "return", "break", "continue", "switch", "case", "default",
        "def", "class", "import", "from", "as", "in", "and", "or", "not", "true", "false", "none",
        "int", "long", "double", "float", "char", "bool", "void", "const", "auto", "string", "vector", "map", "set",
        "cin", "cout", "scanf", "printf", "readln", "writeln", "begin", "end", "var", "procedure", "function",
    }
    normalized = []
    for token in raw_tokens:
        if re.fullmatch(r"\d+(?:\.\d+)?", token):
            normalized.append("NUM")
        elif token == "str":
            normalized.append("STR")
        elif re.fullmatch(r"[a-z_][a-z0-9_]*", token) and token not in keywords:
            normalized.append("ID")
        else:
            normalized.append(token)
    return normalized


def token_fingerprints(tokens: list[str], size: int = 7) -> set[tuple[str, ...]]:
    if len(tokens) < size:
        return {tuple(tokens)} if tokens else set()
    return {tuple(tokens[i : i + size]) for i in range(len(tokens) - size + 1)}


def code_similarity_percent(a: dict, b: dict) -> float:
    if a["ext"] != b["ext"]:
        return 0.0
    fa = a.get("fingerprints") or set()
    fb = b.get("fingerprints") or set()
    if not fa or not fb:
        return 0.0
    jaccard = len(fa & fb) / max(len(fa | fb), 1)
    containment = max(len(fa & fb) / max(len(fa), 1), len(fa & fb) / max(len(fb), 1))
    return round(max(jaccard * 100, containment * 92), 1)


def classify_copy_similarity(percent: float) -> str:
    if percent >= 88:
        return "Ráº¥t giá»‘ng"
    if percent >= 75:
        return "Giá»‘ng nhiá»u"
    if percent >= 62:
        return "Cáº§n xem láº¡i"
    return "Tháº¥p"


def detect_code_copy_pairs(finals: list[dict]) -> tuple[list[dict], list[dict]]:
    by_problem = defaultdict(list)
    for item in finals:
        tokens = normalized_code_tokens(item.get("text", ""), item["ext"])
        item["norm_token_count"] = len(tokens)
        item["fingerprints"] = token_fingerprints(tokens)
        if len(tokens) >= 25 and len(item["fingerprints"]) >= 5:
            by_problem[(item["contest"], item["problem"], item["ext"])].append(item)

    detail_pairs = []
    for (contest, problem, ext), items in by_problem.items():
        items = sorted(items, key=lambda row: row["student_id"])
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                a, b = items[i], items[j]
                if a["student_id"] == b["student_id"]:
                    continue
                percent = code_similarity_percent(a, b)
                if percent < 62:
                    continue
                shared = len(a["fingerprints"] & b["fingerprints"])
                detail_pairs.append(
                    {
                        "contest": contest,
                        "problem": problem,
                        "language": ext.lstrip("."),
                        "student_a": a["student_id"],
                        "student_b": b["student_id"],
                        "percent": percent,
                        "level": classify_copy_similarity(percent),
                        "shared_fingerprints": shared,
                        "tokens_a": a["norm_token_count"],
                        "tokens_b": b["norm_token_count"],
                        "file_a": a["path"],
                        "file_b": b["path"],
                        "local_a": a.get("local_path", ""),
                        "local_b": b.get("local_path", ""),
                    }
                )

    pair_summary = {}
    for row in detail_pairs:
        key = tuple(sorted([row["student_a"], row["student_b"]]))
        current = pair_summary.setdefault(
            key,
            {
                "student_a": key[0],
                "student_b": key[1],
                "pair_count": 0,
                "max_percent": 0.0,
                "avg_percent": 0.0,
                "contests": set(),
                "problems": [],
                "levels": Counter(),
            },
        )
        current["pair_count"] += 1
        current["max_percent"] = max(current["max_percent"], row["percent"])
        current["avg_percent"] += row["percent"]
        current["contests"].add(row["contest"])
        current["problems"].append(f"{row['contest']}/{row['problem']}:{row['percent']}%")
        current["levels"][row["level"]] += 1

    summaries = []
    for row in pair_summary.values():
        row["avg_percent"] = round(row["avg_percent"] / max(row["pair_count"], 1), 1)
        row["contests"] = ", ".join(sorted(row["contests"]))
        row["problems"] = "; ".join(row["problems"][:12])
        if row["levels"].get("Ráº¥t giá»‘ng"):
            row["level"] = "Ráº¥t giá»‘ng"
        elif row["levels"].get("Giá»‘ng nhiá»u"):
            row["level"] = "Giá»‘ng nhiá»u"
        else:
            row["level"] = "Cáº§n xem láº¡i"
        summaries.append(row)
    return summaries, detail_pairs


def classify_ai_score(score: float) -> str:
    if score >= 60:
        return "Kháº£ nÄƒng cao"
    if score >= 45:
        return "Kháº£ nÄƒng trung bÃ¬nh"
    return "Kháº£ nÄƒng tháº¥p"


def safe_output_part(part: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", part).strip()
    return cleaned or "_"


def extract_code_member(zf: zipfile.ZipFile, member: str, output_root: Path, contest: str) -> Path:
    parts = [safe_output_part(part) for part in Path(member).parts if part not in ("", ".", "..")]
    target = output_root / safe_output_part(contest)
    for part in parts:
        target = target / part
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(zf.read(member))
    return target


def collect_code_records_from_zip(zip_path: Path, extract_root: Path | None = None) -> list[dict]:
    contest = normalize_contest_name(zip_path)
    records = []
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if name.endswith("/"):
                continue
            ext = Path(name).suffix.lower()
            if ext not in CODE_EXTENSIONS:
                continue
            parts = Path(name).parts
            if len(parts) < 2:
                continue
            student_id = parts[0]
            is_history = "$History" in parts
            problem = code_problem_from_name(parts[-1])
            try:
                text = read_zip_text(zf, name)
            except Exception:
                continue
            if not text.strip():
                continue
            local_path = ""
            if extract_root is not None:
                local_path = str(extract_code_member(zf, name, extract_root, contest))
            features = analyze_code_text(text, ext)
            records.append(
                {
                    "contest": contest,
                    "student_id": student_id,
                    "problem": problem,
                    "path": name,
                    "filename": parts[-1],
                    "is_history": is_history,
                    "version": history_version_from_name(parts[-1]) if is_history else 10**18,
                    "ext": ext,
                    "text": text,
                    "local_path": local_path,
                    "features": features,
                }
            )
    return records


def final_code_records(records: list[dict]) -> list[dict]:
    grouped = defaultdict(list)
    for record in records:
        grouped[(record["contest"], record["student_id"], record["problem"])].append(record)
    finals = []
    for items in grouped.values():
        root_items = [item for item in items if not item["is_history"]]
        if root_items:
            finals.append(sorted(root_items, key=lambda item: item["path"])[0])
        else:
            finals.append(max(items, key=lambda item: item["version"]))
    return finals


def analyze_ai_code_records(records: list[dict]) -> dict:
    finals = final_code_records(records)
    copy_summaries, copy_details = detect_code_copy_pairs(finals)
    finals_by_student = defaultdict(list)
    all_by_student_problem = defaultdict(list)
    for record in records:
        all_by_student_problem[(record["contest"], record["student_id"], record["problem"])].append(record)
    for record in finals:
        finals_by_student[record["student_id"]].append(record)

    shifts = []
    shift_by_student = defaultdict(list)
    for key, items in all_by_student_problem.items():
        if len(items) < 2:
            continue
        ordered = sorted(items, key=lambda item: (item["version"], item["path"]))
        first, last = ordered[0], ordered[-1]
        dist = vector_distance(first["features"], last["features"])
        if dist >= 35:
            reason = []
            if first["ext"] != last["ext"]:
                reason.append(f"Äá»•i ngÃ´n ngá»¯ {first['ext']} -> {last['ext']}")
            if first["features"]["style_bucket"] != last["features"]["style_bucket"]:
                reason.append(f"Äá»•i style {first['features']['style_bucket']} -> {last['features']['style_bucket']}")
            row = {
                "contest": key[0],
                "student_id": key[1],
                "problem": key[2],
                "versions": len(items),
                "distance": dist,
                "first_file": first["path"],
                "last_file": last["path"],
                "first_local": first.get("local_path", ""),
                "last_local": last.get("local_path", ""),
                "reason": "; ".join(reason) or "Äá»™ lá»‡ch Ä‘áº·c trÆ°ng code lá»›n",
            }
            shifts.append(row)
            shift_by_student[key[1]].append(row)

    students = []
    for student_id, items in sorted(finals_by_student.items()):
        languages = sorted({item["ext"].lstrip(".") for item in items})
        buckets = Counter(item["features"]["style_bucket"] for item in items)
        code_scores = [item["features"]["code_ai_score"] for item in items]
        pair_distances = []
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                pair_distances.append(vector_distance(items[i]["features"], items[j]["features"]))
        max_pair = max(pair_distances) if pair_distances else 0
        avg_pair = sum(pair_distances) / len(pair_distances) if pair_distances else 0
        inconsistency = 0
        reasons = []
        if len(languages) >= 2 and len(items) >= 3:
            inconsistency += min(18, 7 * (len(languages) - 1))
            reasons.append("DÃ¹ng nhiá»u ngÃ´n ngá»¯ trong cÃ¡c bÃ i: " + ", ".join(languages))
        if len(buckets) >= 3 and len(items) >= 3:
            inconsistency += min(20, 6 * (len(buckets) - 2))
            reasons.append("Template/phong cÃ¡ch giá»¯a cÃ¡c bÃ i khÃ¡c nhau")
        if max_pair >= 60:
            inconsistency += 16
            reasons.append("CÃ³ cáº·p bÃ i cÃ¹ng thÃ­ sinh lá»‡ch phong cÃ¡ch ráº¥t máº¡nh")
        elif avg_pair >= 42:
            inconsistency += 10
            reasons.append("Äá»™ lá»‡ch phong cÃ¡ch trung bÃ¬nh cao")
        if shift_by_student.get(student_id):
            inconsistency += min(24, 10 + 4 * len(shift_by_student[student_id]))
            reasons.append("CÃ³ láº§n ná»™p cÃ¹ng bÃ i Ä‘á»•i phong cÃ¡ch/template rÃµ")
        top_code = max(code_scores) if code_scores else 0
        avg_code = sum(code_scores) / len(code_scores) if code_scores else 0
        score = min(100, round(top_code * 0.8 + avg_code * 0.35 + inconsistency, 1))
        code_reason_hits = []
        for item in sorted(items, key=lambda row: row["features"]["code_ai_score"], reverse=True)[:3]:
            code_reason_hits.extend(item["features"]["code_reasons"][:2])
        all_reasons = reasons + code_reason_hits
        students.append(
            {
                "student_id": student_id,
                "level": classify_ai_score(score),
                "score": score,
                "final_count": len(items),
                "history_shift_count": len(shift_by_student.get(student_id, [])),
                "languages": ", ".join(languages),
                "style_count": len(buckets),
                "max_pair_distance": round(max_pair, 1),
                "avg_pair_distance": round(avg_pair, 1),
                "reasons": "; ".join(dict.fromkeys(all_reasons)) or "Ãt dáº¥u hiá»‡u báº¥t thÆ°á»ng",
                "sample_files": "; ".join(item["path"] for item in sorted(items, key=lambda row: row["features"]["code_ai_score"], reverse=True)[:3]),
            }
        )
    details = []
    for item in sorted(records, key=lambda row: (row["contest"], row["student_id"], row["problem"], row["is_history"], row["path"])):
        f = item["features"]
        details.append(
            {
                "contest": item["contest"],
                "student_id": item["student_id"],
                "problem": item["problem"],
                "kind": "history" if item["is_history"] else "final/root",
                "file": item["path"],
                "local_path": item.get("local_path", ""),
                "ext": item["ext"].lstrip("."),
                "code_ai_score": f["code_ai_score"],
                "style_bucket": f["style_bucket"],
                "line_count": f["line_count"],
                "comment_ratio": f["comment_ratio"],
                "macro_count": f["macro_count"],
                "import_count": f["import_count"],
                "function_count": f["function_count"],
                "avg_identifier_len": f["avg_identifier_len"],
                "long_identifier_ratio": f["long_identifier_ratio"],
                "reasons": "; ".join(f["code_reasons"]),
            }
        )
    return {
        "students": students,
        "details": details,
        "shifts": shifts,
        "finals": finals,
        "copy_summaries": copy_summaries,
        "copy_details": copy_details,
    }


def autosize_worksheet(ws) -> None:
    for column in ws.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        ws.column_dimensions[letter].width = max(10, min(max_length + 2, 55))


def write_ai_warning_excel(analysis: dict, output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    def style_header(sheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    def set_hyperlink(cell, target: str, label: str | None = None) -> None:
        if label is not None:
            cell.value = label
        if not target:
            return
        try:
            if target.startswith("#"):
                cell.hyperlink = target
            else:
                cell.hyperlink = Path(target).resolve().as_uri()
            cell.style = "Hyperlink"
        except Exception:
            pass

    def add_sheet_link(sheet, row: int, label: str, sheet_name: str, note: str) -> None:
        sheet.cell(row=row, column=1, value=label)
        set_hyperlink(sheet.cell(row=row, column=2), f"#'{sheet_name}'!A1", "Má»Ÿ sheet")
        sheet.cell(row=row, column=3, value=note)

    ws = wb.active
    ws.title = "Tong quan"
    high = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng cao")
    medium = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng trung bÃ¬nh")
    low = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng tháº¥p")
    copy_very = sum(1 for row in analysis["copy_summaries"] if row["level"] == "Ráº¥t giá»‘ng")
    copy_many = sum(1 for row in analysis["copy_summaries"] if row["level"] == "Giá»‘ng nhiá»u")
    ws["A1"] = "Tá»•ng quan bÃ¡o cÃ¡o cáº£nh bÃ¡o AI code vÃ  chÃ©p code"
    ws["A1"].font = Font(bold=True, size=14)
    overview_rows = [
        ("Sá»‘ thÃ­ sinh", len(analysis["students"])),
        ("Kháº£ nÄƒng AI cao", high),
        ("Kháº£ nÄƒng AI trung bÃ¬nh", medium),
        ("Kháº£ nÄƒng AI tháº¥p", low),
        ("Sá»‘ file code phÃ¢n tÃ­ch", len(analysis["details"])),
        ("Sá»‘ trÆ°á»ng há»£p Ä‘á»•i style cÃ¹ng bÃ i", len(analysis["shifts"])),
        ("Sá»‘ cáº·p nghi chÃ©p code", len(analysis["copy_summaries"])),
        ("Cáº·p ráº¥t giá»‘ng", copy_very),
        ("Cáº·p giá»‘ng nhiá»u", copy_many),
    ]
    row_idx = 3
    for label, value in overview_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="CÃ¡c sheet chi tiáº¿t").font = Font(bold=True)
    row_idx += 1
    add_sheet_link(ws, row_idx, "Cáº£nh bÃ¡o AI theo thÃ­ sinh", "Canh bao AI", "Má»©c cao/trung bÃ¬nh/tháº¥p vÃ  lÃ½ do")
    row_idx += 1
    add_sheet_link(ws, row_idx, "ChÃ©p code tá»•ng há»£p", "Chep code tong hop", "Má»—i cáº·p thÃ­ sinh chá»‰ liá»‡t kÃª má»™t láº§n")
    row_idx += 1
    add_sheet_link(ws, row_idx, "ChÃ©p code chi tiáº¿t", "Chep code chi tiet", "Chi tiáº¿t theo contest/bÃ i, cÃ³ % giá»‘ng nhau")
    row_idx += 1
    add_sheet_link(ws, row_idx, "Chi tiáº¿t file code", "Chi tiet file code", "CÃ³ link má»Ÿ file code Ä‘Ã£ giáº£i nÃ©n")
    row_idx += 1
    add_sheet_link(ws, row_idx, "Äá»•i style cÃ¹ng bÃ i", "Doi style cung bai", "CÃ¡c láº§n ná»™p cÃ¹ng bÃ i Ä‘á»•i template/phong cÃ¡ch")
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Top cáº£nh bÃ¡o AI").font = Font(bold=True)
    row_idx += 1
    ws.append(["MÃ£ thÃ­ sinh", "Má»©c cáº£nh bÃ¡o", "Äiá»ƒm", "LÃ½ do"])
    for item in sorted(analysis["students"], key=lambda r: (-r["score"], r["student_id"]))[:15]:
        ws.append([item["student_id"], item["level"], item["score"], item["reasons"]])
    row_idx = ws.max_row + 2
    ws.cell(row=row_idx, column=1, value="Top cáº·p nghi chÃ©p code").font = Font(bold=True)
    row_idx += 1
    ws.append(["ThÃ­ sinh A", "ThÃ­ sinh B", "Má»©c", "% cao nháº¥t", "Sá»‘ bÃ i/cáº·p", "BÃ i liÃªn quan"])
    for item in sorted(analysis["copy_summaries"], key=lambda r: (-r["max_percent"], -r["pair_count"], r["student_a"], r["student_b"]))[:15]:
        ws.append([item["student_a"], item["student_b"], item["level"], item["max_percent"], item["pair_count"], item["problems"]])
    autosize_worksheet(ws)

    ws = wb.create_sheet("Canh bao AI")
    headers = [
        "MÃ£ thÃ­ sinh", "Má»©c cáº£nh bÃ¡o", "Äiá»ƒm nghi váº¥n", "Sá»‘ bÃ i final", "Sá»‘ Ä‘á»•i style trong history",
        "NgÃ´n ngá»¯", "Sá»‘ nhÃ³m style", "Lá»‡ch lá»›n nháº¥t", "Lá»‡ch trung bÃ¬nh", "LÃ½ do", "File máº«u cáº§n xem",
    ]
    ws.append(headers)
    fills = {
        "Kháº£ nÄƒng cao": PatternFill("solid", fgColor="FCA5A5"),
        "Kháº£ nÄƒng trung bÃ¬nh": PatternFill("solid", fgColor="FDE68A"),
        "Kháº£ nÄƒng tháº¥p": PatternFill("solid", fgColor="BBF7D0"),
    }
    for row in sorted(analysis["students"], key=lambda item: (-item["score"], item["student_id"])):
        ws.append([
            row["student_id"], row["level"], row["score"], row["final_count"], row["history_shift_count"],
            row["languages"], row["style_count"], row["max_pair_distance"], row["avg_pair_distance"],
            row["reasons"], row["sample_files"],
        ])
        ws.cell(ws.max_row, 2).fill = fills.get(row["level"], PatternFill())
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Chep code tong hop")
    ws.append(["ThÃ­ sinh A", "ThÃ­ sinh B", "Má»©c giá»‘ng", "% cao nháº¥t", "% trung bÃ¬nh", "Sá»‘ bÃ i/cáº·p giá»‘ng", "Contest", "BÃ i liÃªn quan"])
    for row in sorted(analysis["copy_summaries"], key=lambda item: (-item["max_percent"], -item["pair_count"], item["student_a"], item["student_b"])):
        ws.append([
            row["student_a"], row["student_b"], row["level"], row["max_percent"], row["avg_percent"],
            row["pair_count"], row["contests"], row["problems"],
        ])
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Chep code chi tiet")
    ws.append([
        "Contest", "MÃ£ bÃ i", "NgÃ´n ngá»¯", "ThÃ­ sinh A", "ThÃ­ sinh B", "% giá»‘ng", "Má»©c",
        "Fingerprint chung", "Token A", "Token B", "File A", "Má»Ÿ file A", "File B", "Má»Ÿ file B",
    ])
    for row in sorted(analysis["copy_details"], key=lambda item: (-item["percent"], item["contest"], item["problem"], item["student_a"], item["student_b"])):
        ws.append([
            row["contest"], row["problem"], row["language"], row["student_a"], row["student_b"],
            row["percent"], row["level"], row["shared_fingerprints"], row["tokens_a"], row["tokens_b"],
            row["file_a"], "Má»Ÿ file", row["file_b"], "Má»Ÿ file",
        ])
        set_hyperlink(ws.cell(ws.max_row, 12), row.get("local_a", ""), "Má»Ÿ file")
        set_hyperlink(ws.cell(ws.max_row, 14), row.get("local_b", ""), "Má»Ÿ file")
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Chi tiet file code")
    detail_headers = [
        "Contest", "MÃ£ thÃ­ sinh", "MÃ£ bÃ i", "Loáº¡i", "File", "Má»Ÿ file", "NgÃ´n ngá»¯", "Äiá»ƒm dáº¥u hiá»‡u AI",
        "NhÃ³m style", "Sá»‘ dÃ²ng", "Tá»‰ lá»‡ comment", "Macro", "Import", "HÃ m", "Äá»™ dÃ i tÃªn TB",
        "Tá»‰ lá»‡ tÃªn dÃ i", "LÃ½ do",
    ]
    ws.append(detail_headers)
    for row in analysis["details"]:
        ws.append([
            row["contest"], row["student_id"], row["problem"], row["kind"], row["file"], "Má»Ÿ file", row["ext"],
            row["code_ai_score"], row["style_bucket"], row["line_count"], row["comment_ratio"],
            row["macro_count"], row["import_count"], row["function_count"], row["avg_identifier_len"],
            row["long_identifier_ratio"], row["reasons"],
        ])
        set_hyperlink(ws.cell(ws.max_row, 6), row.get("local_path", ""), "Má»Ÿ file")
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Doi style cung bai")
    shift_headers = ["Contest", "MÃ£ thÃ­ sinh", "MÃ£ bÃ i", "Sá»‘ phiÃªn báº£n", "Äá»™ lá»‡ch", "File Ä‘áº§u", "Má»Ÿ file Ä‘áº§u", "File cuá»‘i", "Má»Ÿ file cuá»‘i", "LÃ½ do"]
    ws.append(shift_headers)
    for row in sorted(analysis["shifts"], key=lambda item: (-item["distance"], item["student_id"])):
        ws.append([
            row["contest"], row["student_id"], row["problem"], row["versions"], row["distance"],
            row["first_file"], "Má»Ÿ file", row["last_file"], "Má»Ÿ file", row["reason"],
        ])
        set_hyperlink(ws.cell(ws.max_row, 7), row.get("first_local", ""), "Má»Ÿ file")
        set_hyperlink(ws.cell(ws.max_row, 9), row.get("last_local", ""), "Má»Ÿ file")
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws = wb.create_sheet("Giai thich")
    notes = [
        ["LÆ°u Ã½", "ÄÃ¢y lÃ  bÃ¡o cÃ¡o cáº£nh bÃ¡o/nghi váº¥n, khÃ´ng pháº£i káº¿t luáº­n cháº¯c cháº¯n thÃ­ sinh dÃ¹ng AI."],
        ["Nguá»“n Ä‘iá»ƒm", "Äiá»ƒm káº¿t há»£p dáº¥u hiá»‡u trong tá»«ng file code, Ä‘á»™ lá»‡ch phong cÃ¡ch giá»¯a cÃ¡c bÃ i, vÃ  Ä‘á»•i phong cÃ¡ch trong history cÃ¹ng bÃ i."],
        ["Kháº£ nÄƒng cao", "Äiá»ƒm nghi váº¥n tá»« 60 trá»Ÿ lÃªn."],
        ["Kháº£ nÄƒng trung bÃ¬nh", "Äiá»ƒm nghi váº¥n tá»« 45 Ä‘áº¿n dÆ°á»›i 60."],
        ["Kháº£ nÄƒng tháº¥p", "Äiá»ƒm nghi váº¥n dÆ°á»›i 45."],
        ["NÃªn xem láº¡i", "Æ¯u tiÃªn má»Ÿ cÃ¡c file trong cá»™t File máº«u cáº§n xem vÃ  sheet Äá»•i style cÃ¹ng bÃ i."],
        ["ChÃ©p code", "So khá»›p cÃ¡c cáº·p final/root cÃ¹ng contest vÃ  cÃ¹ng bÃ i. File quÃ¡ ngáº¯n khÃ´ng Ä‘Æ°á»£c cháº¥m Ä‘á»ƒ trÃ¡nh nhiá»…u."],
        ["% giá»‘ng", "Dá»±a trÃªn token code Ä‘Ã£ bá» comment, chuáº©n hÃ³a tÃªn biáº¿n/háº±ng sá»‘ vÃ  so fingerprint k-gram."],
        ["Link file", "BÃ¡o cÃ¡o cÃ³ link tá»›i thÆ° má»¥c code Ä‘Ã£ giáº£i nÃ©n trong .runtime/misc cá»§a tool local."],
        ["KhÃ´ng phÃ¢n tÃ­ch", "File Scratch .sb3 lÃ  nhá»‹ phÃ¢n nÃªn khÃ´ng Ä‘Æ°á»£c cháº¥m báº±ng heuristic code vÄƒn báº£n."],
    ]
    for row in notes:
        ws.append(row)
    style_header(ws)
    autosize_worksheet(ws)
    wb.save(output_path)


def build_ai_warning_report(source_zips: list[Path], output_path: Path) -> dict:
    records = []
    extract_root = output_path.parent / "extracted_code"
    if extract_root.exists():
        shutil.rmtree(extract_root)
    extract_root.mkdir(parents=True, exist_ok=True)
    for zip_path in source_zips:
        records.extend(collect_code_records_from_zip(zip_path, extract_root))
    if not records:
        raise RuntimeError("KhÃ´ng tÃ¬m tháº¥y file code vÄƒn báº£n (.py/.cpp/.pas/.c/.java) trong dá»¯ liá»‡u.")
    analysis = analyze_ai_code_records(records)
    write_ai_warning_excel(analysis, output_path)
    high = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng cao")
    medium = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng trung bÃ¬nh")
    low = sum(1 for row in analysis["students"] if row["level"] == "Kháº£ nÄƒng tháº¥p")
    return {
        "zip_count": len(source_zips),
        "code_file_count": len(records),
        "student_count": len(analysis["students"]),
        "high": high,
        "medium": medium,
        "low": low,
        "shift_count": len(analysis["shifts"]),
        "copy_pair_count": len(analysis["copy_summaries"]),
        "copy_detail_count": len(analysis["copy_details"]),
        "copy_very_similar": sum(1 for row in analysis["copy_summaries"] if row["level"] == "Ráº¥t giá»‘ng"),
        "copy_many": sum(1 for row in analysis["copy_summaries"] if row["level"] == "Giá»‘ng nhiá»u"),
        "extracted_folder": str(extract_root),
        "filename": output_path.name,
    }


@app.before_request
def require_basic_auth():
    auth_user = os.getenv("TOOL_OJ_AUTH_USER")
    auth_pass = os.getenv("TOOL_OJ_AUTH_PASS")
    if not auth_user and not auth_pass:
        return None
    auth = request.authorization
    if auth and auth.username == auth_user and auth.password == auth_pass:
        return None
    return Response(
        "Authentication required",
        401,
        {"WWW-Authenticate": 'Basic realm="Tool HNCode"'},
    )


PAGE = r"""
<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Tool HNCode</title>
  <link rel="icon" type="image/svg+xml" href="/static/favicon-HNCode.svg">
  <style>
    :root { --bg:#f5f7fb; --panel:#fff; --ink:#172033; --muted:#667085; --line:#d8dee9; --soft:#eef2f6; --accent:#0f766e; --ok:#087443; --bad:#b42318; --warn:#b54708; --code:#101828; }
    * { box-sizing:border-box; }
    body { margin:0; background:var(--bg); color:var(--ink); font-family:"Segoe UI", Arial, sans-serif; font-size:14px; }
    header { background:var(--panel); border-bottom:1px solid var(--line); padding:16px 22px; display:flex; align-items:center; justify-content:space-between; gap:16px; flex-wrap:wrap; }
    h1 { margin:0; font-size:22px; letter-spacing:0; }
    h2 { margin:0 0 8px; font-size:18px; }
    h3 { margin:16px 0 8px; font-size:15px; }
    p { color:var(--muted); line-height:1.45; margin:0 0 12px; }
    .nav { display:flex; gap:10px; flex-wrap:wrap; align-items:stretch; }
    .nav-group { display:flex; gap:6px; flex-wrap:wrap; align-items:center; border:1px solid var(--line); background:#f8fafc; border-radius:8px; padding:6px; }
    .nav-label { color:#475467; font-size:12px; font-weight:800; padding:0 4px; text-transform:uppercase; letter-spacing:.03em; }
    .nav button, button.action { border:1px solid #b8c2d3; border-radius:6px; padding:10px 14px; background:#fff; color:var(--ink); font:inherit; font-weight:700; cursor:pointer; box-shadow:0 1px 2px rgba(16,24,40,.08); }
    .nav button:hover, button.action:hover { border-color:#8fa1b8; background:#f8fafc; }
    .nav button.active, button.primary { background:var(--accent); border-color:var(--accent); color:#fff; box-shadow:0 2px 6px rgba(15,118,110,.24); }
    button.primary:hover { background:#0b665f; border-color:#0b665f; }
    button:disabled { opacity:.5; cursor:not-allowed; }
    main { max-width:1480px; margin:0 auto; padding:20px; display:flex; align-items:flex-start; gap:18px; }
    section { background:var(--panel); border:1px solid var(--line); border-radius:8px; }
    main > section:first-child { flex:1 1 auto; min-width:0; }
    .panel { display:none; padding:18px; }
    .panel.active { display:block; }
    label { display:block; margin:12px 0 6px; color:#344054; font-weight:650; }
    input[type=text], input[type=password], select, textarea { width:100%; border:1px solid var(--line); border-radius:6px; padding:9px 10px; font:inherit; background:#fff; color:var(--ink); }
    textarea { min-height:78px; resize:vertical; line-height:1.45; }
    .grid-2 { display:grid; grid-template-columns:1fr 1fr; gap:10px; }
    .grid-3 { display:grid; grid-template-columns:1fr 1fr 1fr; gap:10px; }
    .row { display:flex; gap:10px; align-items:end; flex-wrap:wrap; }
    .row > .grow { flex:1 1 340px; }
    .actions { display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }
    #toggleGuide { margin-top:6px; }
    .table-tools { display:flex; gap:8px; margin-top:14px; flex-wrap:wrap; }
    .note, .guide { border:1px solid #b8d8d3; background:#f0fdfa; color:#134e48; border-radius:8px; padding:12px; line-height:1.48; margin:12px 0; }
    .guide { border-color:var(--line); background:#fafbfc; color:var(--ink); }
    .tool-card { border:1px solid #cbd5e1; background:#fff; border-radius:8px; padding:16px; margin-top:16px; box-shadow:0 1px 3px rgba(16,24,40,.08); }
    .tool-card + .tool-card { margin-top:18px; }
    .tool-title { display:flex; align-items:center; gap:10px; margin:0 0 8px; font-size:18px; color:#0f172a; }
    .tool-title::before { content:""; width:6px; height:24px; border-radius:999px; background:var(--accent); display:inline-block; }
    .tool-subtitle { margin-bottom:14px; }
    .sample, pre#log { background:var(--code); color:#f2f4f7; border-radius:6px; padding:12px; white-space:pre-wrap; overflow:auto; font-family:Consolas, "Cascadia Mono", monospace; font-size:12px; line-height:1.45; }
    .log-panel { flex:0 0 auto; width:min(380px, 34vw); min-width:280px; max-width:680px; display:grid; grid-template-rows:auto minmax(360px, 1fr); min-height:520px; resize:horizontal; overflow:auto; }
    .log-head { padding:14px 16px; border-bottom:1px solid var(--line); display:flex; justify-content:space-between; gap:12px; align-items:center; }
    pre#log { margin:0; border-radius:0 0 8px 8px; }
    .status { border-radius:999px; padding:4px 10px; background:var(--soft); color:var(--muted); font-weight:650; font-size:12px; }
    .status.ok { background:#dcfae6; color:var(--ok); }
    .status.err { background:#fee4e2; color:var(--bad); }
    .status.warn { background:#fef0c7; color:var(--warn); }
    .row-status.ok { color:var(--ok); font-weight:700; }
    .row-status.err { color:var(--bad); font-weight:700; }
    .row-status.warn { color:var(--warn); font-weight:700; }
    .log-ok { color:#86efac; font-weight:700; }
    .log-err { color:#fca5a5; font-weight:700; }
    .log-warn { color:#fde68a; font-weight:700; }
    .log-progress { color:#bfdbfe; font-weight:700; }
    .login-badge { display:inline-flex; align-items:center; min-height:24px; border-radius:999px; padding:3px 9px; background:var(--soft); color:var(--muted); font-size:12px; font-weight:700; margin-top:6px; }
    .login-badge.ok { background:#dcfae6; color:var(--ok); }
    .login-badge.err { background:#fee4e2; color:var(--bad); }
    table { width:100%; border-collapse:collapse; margin-top:14px; font-size:13px; }
    th, td { border-bottom:1px solid var(--line); padding:8px; vertical-align:top; text-align:left; }
    th { background:#f8fafc; font-weight:700; }
    .inner-table { margin-top:0; font-size:12px; }
    .inner-table th, .inner-table td { padding:5px 6px; }
    td input[type=text] { padding:6px 7px; }
    a.problem-link { color:var(--accent); font-weight:700; text-decoration:none; }
    .test-meta { color:var(--muted); font-size:12px; line-height:1.4; }
    .lang-list { display:grid; grid-template-columns:repeat(3, minmax(0,1fr)); gap:8px; margin-top:8px; }
    .check { display:flex; align-items:center; gap:7px; }
    .hidden { display:none; }
    @media (max-width:980px) { main { display:block; padding:14px; } .log-panel { width:100%; max-width:none; margin-top:14px; resize:vertical; } .grid-2,.grid-3,.lang-list { grid-template-columns:1fr; } }
  </style>
</head>
<body>
  <header>
    <h1>Tool HNCode</h1>
    <div class="nav">
      <div class="nav-group">
        <span class="nav-label">Chung</span>
        <button type="button" class="active" data-panel="accounts">TÃ i khoáº£n & HÆ°á»›ng dáº«n</button>
      </div>
      <div class="nav-group">
        <span class="nav-label">BÃ i táº­p</span>
        <button type="button" data-panel="upload">Up nhiá»u bÃ i</button>
        <button type="button" data-panel="single-upload">Up 1 bÃ i</button>
        <button type="button" data-panel="transfer">Chuyá»ƒn bÃ i</button>
      </div>
      <div class="nav-group">
        <span class="nav-label">Contest / Course</span>
        <button type="button" data-panel="contest-transfer">Chuyá»ƒn contest</button>
        <button type="button" data-panel="contest-create">Táº¡o contest</button>
        <button type="button" data-panel="contest-lesson-copy">Contest â†’ Lesson</button>
        <button type="button" data-panel="course-clone">Clone Course</button>
      </div>
      <div class="nav-group">
        <span class="nav-label">KhÃ¡c</span>
        <button type="button" data-panel="quiz-upload">Up Quiz</button>
        <button type="button" data-panel="misc-tools">Tool láº»</button>
      </div>
    </div>
  </header>

  <main>
    <section>
      <div class="panel active" id="panel-accounts">
        <h2>TÃ i khoáº£n & HÆ°á»›ng dáº«n</h2>
        <p>LÆ°u táº¡m tÃ i khoáº£n trÃªn trÃ¬nh duyá»‡t mÃ¡y nÃ y. Khi cháº¡y tÃ¡c vá»¥, form sáº½ tá»± Ä‘iá»n cÃ¡c thÃ´ng tin Ä‘Ã£ lÆ°u.</p>
        <div class="grid-3">
          <div><label>HNOJ user</label><input id="acct_hnoj_user" type="text" value="MrTee"><span id="login_hnoj" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>HNCode user</label><input id="acct_hncode_user" type="text" value="MrTee"><span id="login_hncode" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>TinHocTre user</label><input id="acct_tinhoctre_user" type="text" value="MrTee"><span id="login_tinhoctre" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
        </div>
        <div class="grid-3">
          <div><label>HNOJ password</label><input id="acct_hnoj_pass" type="password"></div>
          <div><label>HNCode password</label><input id="acct_hncode_pass" type="password"></div>
          <div><label>TinHocTre password</label><input id="acct_tinhoctre_pass" type="password"></div>
        </div>
        <textarea id="acct_tinhoctre_cookie" class="hidden"></textarea>
        <div class="actions">
          <button class="action primary" type="button" id="saveAccounts">LÆ°u táº¡m</button>
          <button class="action" type="button" id="checkAccounts">Kiá»ƒm tra Ä‘Äƒng nháº­p</button>
          <button class="action" type="button" id="clearAccounts">XÃ³a thÃ´ng tin Ä‘Ã£ lÆ°u</button>
        </div>
        <button class="action" type="button" id="toggleGuide">áº¨n / Hiá»‡n hÆ°á»›ng dáº«n prompt</button>
        <div class="guide hidden" id="promptGuide"><div class="sample">{{ prompt_guide }}</div></div>
      </div>

      <div class="panel" id="panel-upload">
        <h2>Up nhiá»u bÃ i</h2>
        <p>Chá»n web Ä‘Ã­ch, chá»n zip bá»™ bÃ i, báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u Ä‘á»ƒ xem báº£ng bÃ i trÆ°á»›c khi up tháº­t.</p>
        <div class="grid-2">
          <div>
            <label>Web Ä‘Ã­ch</label>
            <select id="uploadTarget">
              <option value="hnoj">HNOJ</option>
              <option value="hncode">HNCode</option>
              <option value="tinhoctre">TinHocTre</option>
            </select><span id="uploadTargetLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span>
          </div>
          <div>
            <label>File zip bá»™ bÃ i hoáº·c file Markdown tá»•ng há»£p</label>
            <div class="row">
              <div class="grow"><input id="uploadZip" type="text" value="{{ default_zip }}"></div>
              <button class="action" type="button" id="chooseZip">Chá»n file</button>
              <button class="action" type="button" id="useBatchSample">DÃ¹ng máº«u Tá»•ng hai sá»‘</button>
              <input id="zipFileInput" class="hidden" type="file" accept=".zip,.md,application/zip,text/markdown,text/plain">
            </div>
          </div>
        </div>
        <div class="note" style="margin-top:12px">
          <b>Cáº¥u trÃºc file zip bá»™ bÃ i:</b><br>
          Má»—i bÃ i nÃªn cÃ³ Ä‘á»§ file <code>&lt;ma_bai&gt;.md</code>, <code>gentest_&lt;ma_bai&gt;.py</code> hoáº·c <code>&lt;ma_bai&gt;.zip</code>, <code>sol_&lt;ma_bai&gt;.md</code> vÃ  náº¿u cáº§n ná»™p thá»­ thÃ¬ cÃ³ <code>sol_&lt;ma_bai&gt;.cpp</code>, <code>sol_&lt;ma_bai&gt;.py</code>.<br>
          File Markdown nÃªn cÃ³ dÃ²ng Ä‘áº§u <code>TÃªn bÃ i | MÃ£ bÃ i | Äiá»ƒm | CÃ¡c Tags</code>. File sinh test sáº½ táº¡o thÆ° má»¥c test vÃ  nÃ©n thÃ nh <code>&lt;ma_bai&gt;.zip</code>; náº¿u zip test cÃ³ sáºµn thÃ¬ tool dÃ¹ng trá»±c tiáº¿p. ThÃ´ng tin nÃ o thiáº¿u sáº½ Ä‘á»ƒ trá»‘ng hoáº·c dÃ¹ng máº·c Ä‘á»‹nh.
          <br><b>RÃ ng buá»™c gentest:</b> nÃªn lÃ  Python, tÃªn <code>gentest_&lt;ma_bai&gt;.py</code>, tá»± táº¡o zip <code>&lt;ma_bai&gt;.zip</code> hoáº·c má»™t zip duy nháº¥t cÃ³ Ä‘á»§ cáº·p <code>.inp/.out</code>; khÃ´ng cáº§n input tÆ°Æ¡ng tÃ¡c; cháº¡y trong 120 giÃ¢y; náº¿u dÃ¹ng C++ trong gentest thÃ¬ mÃ¡y/VPS cáº§n cÃ³ <code>g++</code>.
          <br><a class="problem-link" href="/samples/bo_mau_1_bai_tonghaiso.zip" target="_blank" rel="noopener">Táº£i máº«u bo_mau_1_bai_tonghaiso.zip</a>
        </div>
        <div class="grid-2">
          <div><label>Giá»›i háº¡n thá»i gian</label><input id="timeLimit" type="text" value="1.0"></div>
          <div><label>Giá»›i háº¡n bá»™ nhá»›</label><input id="memoryLimit" type="text" value="1048576"></div>
        </div>
        <h3>NgÃ´n ngá»¯ cho phÃ©p</h3>
        <div id="languages" class="lang-list"></div>

        <div class="actions">
          <button class="action" type="button" id="toggleAdvanced">Má»Ÿ rá»™ng thÃ´ng tin khÃ¡c</button>
        </div>
        <div id="advancedUpload" class="hidden">
          <div class="grid-3">
            <div><label>NgÆ°á»i táº¡o (Creators)</label><input id="creator" type="text" value="mrtee"></div>
            <div><label>Äiá»ƒm máº·c Ä‘á»‹nh</label><input id="uploadPoints" type="text" value="100"></div>
            <div><label>Dáº¡ng bÃ i táº­p / Tags máº·c Ä‘á»‹nh</label><input id="uploadTags" type="text" placeholder="ChÆ°a phÃ¢n loáº¡i, implementation, math, hoáº·c Type ID"></div>
          </div>
          <div class="grid-3">
            <label class="check"><input type="checkbox" id="uploadPartial" checked> Cho phÃ©p Ä‘iá»ƒm thÃ nh pháº§n</label>
            <label class="check"><input type="checkbox" id="overwriteExisting"> Ghi Ä‘Ã¨ bÃ i Ä‘Ã£ cÃ³</label>
            <label class="check"><input type="checkbox" id="overwriteStatement"> Ghi Ä‘Ã¨ Ä‘á» bÃ i náº¿u mÃ£ bÃ i Ä‘Ã£ cÃ³</label>
            <label class="check"><input type="checkbox" id="overwriteTests"> Ghi Ä‘Ã¨ test náº¿u mÃ£ bÃ i Ä‘Ã£ cÃ³</label>
          </div>
          <div class="grid-3">
            <div><label>Dáº¡ng Ä‘á» (Problem types)</label><input id="typeLabel" type="text" value="ChÆ°a phÃ¢n loáº¡i" disabled></div>
            <div><label>NhÃ³m bÃ i (Problem group)</label><input id="groupLabel" type="text" value="ChÆ°a phÃ¢n loáº¡i" disabled></div>
          </div>
        </div>

        <div class="grid-3" style="margin-top:12px">
          <label class="check"><input type="checkbox" id="submitCpp"> Ná»™p bÃ i cháº¥m thá»­ C++</label>
          <label class="check"><input type="checkbox" id="submitPython" checked> Ná»™p bÃ i cháº¥m thá»­ Python</label>
          <label class="check"><input type="checkbox" id="noSubmit"> KhÃ´ng ná»™p bÃ i cháº¥m thá»­</label>
        </div>
        <label class="check" style="margin-top:12px"><input type="checkbox" id="skipStatementTitle" checked> Bá» dÃ²ng Ä‘áº§u tiÃªn trong file Ä‘á» bÃ i</label>
        <div class="actions">
          <button class="action primary" type="button" id="prepareUpload">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmUpload" disabled>XÃ¡c nháº­n Up nhiá»u bÃ i</button>
        </div>
        <div id="uploadTable"></div>
      </div>

      <div class="panel" id="panel-single-upload">
        <h2>Up 1 bÃ i</h2>
        <p>Nháº­p trá»±c tiáº¿p má»™t bÃ i, kiá»ƒm tra dá»¯ liá»‡u trÆ°á»›c, rá»“i xÃ¡c nháº­n up. Thiáº¿u pháº§n nÃ o thÃ¬ tool bá» qua pháº§n Ä‘Ã³.</p>
        <div class="grid-2">
          <div>
            <label>Web Ä‘Ã­ch</label>
            <select id="singleUploadTarget">
              <option value="hnoj">HNOJ</option>
              <option value="hncode">HNCode</option>
              <option value="tinhoctre">TinHocTre</option>
            </select><span id="singleUploadLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span>
          </div>
          <div>
            <label>MÃ£ bÃ i</label>
            <input id="singleCode" type="text" placeholder="tht26_tongbi">
          </div>
        </div>
        <div><label>TÃªn bÃ i toÃ¡n</label><input id="singleName" type="text" placeholder="Tá»•ng bi"></div>
        <div class="grid-2">
          <div><label>Giá»›i háº¡n thá»i gian</label><input id="singleTimeLimit" type="text" value="1.0"></div>
          <div><label>Giá»›i háº¡n bá»™ nhá»›</label><input id="singleMemoryLimit" type="text" value="1024M"></div>
        </div>
        <div class="grid-3">
          <div><label>Äiá»ƒm</label><input id="singlePoints" type="text" value="100"></div>
          <div><label>Dáº¡ng bÃ i táº­p / Tags</label><input id="singleTags" type="text" placeholder="CÃ³ thá»ƒ Ä‘á»ƒ trá»‘ng; náº¿u nháº­p sá»‘ ID thÃ¬ dÃ¹ng lÃ m Type ID"></div>
          <label class="check"><input type="checkbox" id="singlePartial" checked> Cho phÃ©p Ä‘iá»ƒm thÃ nh pháº§n</label>
        </div>
        <div class="grid-2" style="margin-top:12px">
          <label class="check"><input type="checkbox" id="singleOverwrite"> Ghi Ä‘Ã¨ náº¿u mÃ£ bÃ i Ä‘Ã£ cÃ³</label>
        </div>
        <h3>NgÃ´n ngá»¯ cho phÃ©p</h3>
        <div id="singleLanguages" class="lang-list"></div>

        <div class="tool-card">
          <h3 class="tool-title">Äá» bÃ i</h3>
          <div class="actions">
            <button class="action" type="button" id="toggleSingleStatement">Thu gá»n Ä‘á» bÃ i</button>
            <button class="action" type="button" id="chooseSingleStatement">Chá»n file .md</button>
            <input id="singleStatementFile" class="hidden" type="file" accept=".md,text/markdown,text/plain">
          </div>
          <div id="singleStatementBox">
            <textarea id="singleStatement" placeholder="DÃ²ng Ä‘áº§u cÃ³ thá»ƒ lÃ : TÃªn bÃ i | ma_bai | Äiá»ƒm | Tags&#10;&#10;Sau Ä‘Ã³ lÃ  ná»™i dung Ä‘á» bÃ i."></textarea>
            <label class="check" style="margin-top:8px"><input type="checkbox" id="singleSkipStatementTitle" checked> Bá» dÃ²ng Ä‘áº§u tiÃªn trong file Ä‘á» bÃ i</label>
          </div>
        </div>

        <div class="tool-card">
          <h3 class="tool-title">Code sinh test</h3>
          <div class="actions">
            <button class="action" type="button" id="toggleSingleGenerator">Thu gá»n sinh test</button>
            <button class="action" type="button" id="chooseSingleGenerator">Chá»n code Python / C++</button>
            <button class="action" type="button" id="chooseSingleTestZip">Chá»n zip test cÃ³ sáºµn</button>
            <button class="action" type="button" id="useSingleSample">DÃ¹ng máº«u Tá»•ng hai sá»‘</button>
            <input id="singleGeneratorFile" class="hidden" type="file" accept=".py,.cpp,text/plain">
            <input id="singleTestZipFile" class="hidden" type="file" accept=".zip,application/zip">
          </div>
          <div id="singleGeneratorBox">
            <input id="singleGeneratorName" type="text" placeholder="ChÆ°a chá»n file sinh test" readonly>
            <input id="singleTestZipName" type="text" placeholder="ChÆ°a chá»n zip test cÃ³ sáºµn" readonly>
            <textarea id="singleGenerator" placeholder="DÃ¡n code gentest Python vÃ o Ä‘Ã¢y. Gentest cáº§n tá»± sinh zip test, khÃ´ng chá» nháº­p bÃ n phÃ­m, cháº¡y trong 120 giÃ¢y; náº¿u gá»i g++ thÃ¬ mÃ¡y/VPS pháº£i cÃ³ g++."></textarea>
            <div class="note"><b>RÃ ng buá»™c gentest:</b> nÃªn Ä‘áº·t tÃªn <code>gentest_&lt;ma_bai&gt;.py</code>; táº¡o zip <code>&lt;ma_bai&gt;.zip</code> hoáº·c má»™t zip duy nháº¥t; trong zip cÃ³ Ä‘á»§ cáº·p <code>.inp/.out</code>. Náº¿u chá»n zip test cÃ³ sáºµn thÃ¬ tool Æ°u tiÃªn zip Ä‘Ã³.</div>
          </div>
        </div>

        <div class="tool-card">
          <h3 class="tool-title">Lá»i giáº£i / hÆ°á»›ng dáº«n</h3>
          <div class="actions">
            <button class="action" type="button" id="toggleSingleSolution">Thu gá»n lá»i giáº£i</button>
            <button class="action" type="button" id="chooseSingleSolution">Chá»n file .md</button>
            <input id="singleSolutionFile" class="hidden" type="file" accept=".md,text/markdown,text/plain">
          </div>
          <div id="singleSolutionBox">
            <textarea id="singleSolution" placeholder="DÃ¡n lá»i giáº£i/hÆ°á»›ng dáº«n Markdown náº¿u muá»‘n up kÃ¨m."></textarea>
          </div>
        </div>

        <div class="actions">
          <button class="action primary" type="button" id="prepareSingleUpload">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmSingleUpload" disabled>XÃ¡c nháº­n Up 1 bÃ i</button>
        </div>
        <div id="singleUploadTable"></div>
      </div>

      <div class="panel" id="panel-transfer">
        <h2>Chuyá»ƒn bÃ i</h2>
        <p>Chá»n nguá»“n, Ä‘Ã­ch vÃ  danh sÃ¡ch mÃ£ bÃ i. Tool sáº½ láº¥y Ä‘á»/test tá»« nguá»“n rá»“i táº¡o bÃ i vÃ  upload test á»Ÿ Ä‘Ã­ch.</p>
        <div class="grid-2">
          <div><label>Nguá»“n</label><select id="transferSource"><option value="tinhoctre">TinHocTre</option><option value="hnoj">HNOJ</option><option value="hncode">HNCode</option></select><span id="transferSourceLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>ÄÃ­ch</label><select id="transferDest"><option value="hncode">HNCode</option><option value="hnoj">HNOJ</option><option value="tinhoctre">TinHocTre</option></select><span id="transferDestLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
        </div>
        <div class="grid-2">
          <div><label>Giá»›i háº¡n thá»i gian máº·c Ä‘á»‹nh</label><input id="transferTimeLimit" type="text" value="1.0"></div>
          <div><label>Giá»›i háº¡n bá»™ nhá»› máº·c Ä‘á»‹nh</label><input id="transferMemoryLimit" type="text" value="1048576"></div>
        </div>
        <div class="actions">
          <button class="action" type="button" id="applyTransferLimits">Ãp dá»¥ng cho táº¥t cáº£ cÃ¡c bÃ i</button>
          <button class="action" type="button" id="resetTransferLimits">Máº·c Ä‘á»‹nh</button>
        </div>
        <h3>NgÃ´n ngá»¯ cho phÃ©p á»Ÿ Ä‘Ã­ch</h3>
        <div id="transferLanguages" class="lang-list"></div>
        <div class="actions">
          <button class="action" type="button" id="toggleTransferAdvanced">Má»Ÿ rá»™ng thÃ´ng tin khÃ¡c</button>
        </div>
        <div id="advancedTransfer" class="hidden">
          <div class="grid-3">
            <div><label>NgÆ°á»i táº¡o (Creators)</label><input id="transferCreator" type="text" value="mrtee"></div>
            <div><label>Dáº¡ng Ä‘á» (Problem types)</label><input id="transferTypeLabel" type="text" value="ChÆ°a phÃ¢n loáº¡i" disabled></div>
            <div><label>NhÃ³m bÃ i (Problem group)</label><input id="transferGroupLabel" type="text" value="ChÆ°a phÃ¢n loáº¡i" disabled></div>
          </div>
        </div>
        <label>Danh sÃ¡ch mÃ£ bÃ i cáº§n chuyá»ƒn</label>
        <textarea id="transferCodes" placeholder="tht26_tongbi&#10;tht26_quatang"></textarea>
        <div class="actions">
          <button class="action primary" type="button" id="prepareTransfer">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmTransfer" disabled>XÃ¡c nháº­n chuyá»ƒn bÃ i</button>
        </div>
        <div id="transferTable"></div>
      </div>

      <div class="panel" id="panel-contest-transfer">
        <h2>Chuyá»ƒn contest</h2>
        <p>Chuyá»ƒn contest gá»“m thÃ´ng tin cÆ¡ báº£n, danh sÃ¡ch bÃ i, Ä‘iá»ƒm vÃ  bá»™ test cá»§a tá»«ng bÃ i. KhÃ´ng chuyá»ƒn bÃ i ná»™p cá»§a há»c sinh.</p>
        <div class="grid-2">
          <div>
            <label>Nguá»“n</label>
            <select id="contestSource">
              <option value="hnoj">HNOJ</option>
              <option value="hncode">HNCode</option>
              <option value="tinhoctre">TinHocTre</option>
            </select><span id="contestSourceLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span>
          </div>
          <div>
            <label>ÄÃ­ch</label>
            <select id="contestDest">
              <option value="hnoj">HNOJ</option>
              <option value="hncode">HNCode</option>
              <option value="tinhoctre">TinHocTre</option>
            </select><span id="contestDestLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span>
          </div>
        </div>
        <label>Danh sÃ¡ch mÃ£ contest cáº§n chuyá»ƒn</label>
        <textarea id="contestCodes" placeholder="tht2026_hn_ck_a&#10;tht2026_hn_ck_b&#10;tht2026_hn_ck_c"></textarea>
        <div class="grid-2">
          <div><label>Time máº·c Ä‘á»‹nh cho bÃ i thiáº¿u thÃ´ng tin</label><input id="contestProblemTime" type="text" value="1.0"></div>
          <div><label>Memory máº·c Ä‘á»‹nh cho bÃ i thiáº¿u thÃ´ng tin</label><input id="contestProblemMemory" type="text" value="1048576"></div>
        </div>
        <label class="check" style="margin-top:12px"><input type="checkbox" id="contestReuseExistingProblems" checked> Náº¿u bÃ i Ä‘Ã£ cÃ³ á»Ÿ Ä‘Ã­ch thÃ¬ dÃ¹ng láº¡i bÃ i Ä‘Ã³</label>
        <label class="check" style="margin-top:8px"><input type="checkbox" id="contestCreateMissingProblems" checked> Tá»± chuyá»ƒn bÃ i/test cÃ²n thiáº¿u trÆ°á»›c khi táº¡o contest</label>
        <div class="actions">
          <button class="action primary" type="button" id="prepareContestTransfer">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmContestTransfer" disabled>XÃ¡c nháº­n chuyá»ƒn contest</button>
        </div>
        <div id="contestTransferTable"></div>
      </div>

      <div class="panel" id="panel-contest-create">
        <h2>Táº¡o contest tá»« mÃ£ bÃ i</h2>
        <p>Táº¡o contest cÆ¡ báº£n vÃ  gáº¯n cÃ¡c mÃ£ bÃ i Ä‘Ã£ cÃ³ trÃªn web Ä‘Ã­ch. CÃ¡c thiáº¿t láº­p chi tiáº¿t cÃ³ thá»ƒ chá»‰nh láº¡i trong admin sau.</p>
        <div class="grid-2">
          <div><label>Web Ä‘Ã­ch</label><select id="createContestTarget"><option value="hnoj">HNOJ</option><option value="hncode">HNCode</option><option value="tinhoctre">TinHocTre</option></select><span id="createContestLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>MÃ£ contest</label><input id="createContestKey" type="text" placeholder="tht2026_hn_ck_a"></div>
        </div>
        <label>TÃªn contest</label><input id="createContestName" type="text" placeholder="TIN Há»ŒC TRáºº 2026 - HÃ€ Ná»˜I - CHUNG Káº¾T - Báº¢NG A">
        <div class="grid-2">
          <div><label>Báº¯t Ä‘áº§u</label><input id="createContestStart" type="datetime-local"></div>
          <div><label>Káº¿t thÃºc</label><input id="createContestEnd" type="datetime-local"></div>
        </div>
        <div class="actions">
          <button class="action" type="button" id="contestTimeToday">HÃ´m nay 8:00-11:00</button>
          <button class="action" type="button" id="contestTimeTomorrow">NgÃ y mai 8:00-11:00</button>
          <button class="action" type="button" id="contestTime90">Káº¿t thÃºc sau 90 phÃºt</button>
        </div>
        <label>Danh sÃ¡ch mÃ£ bÃ i</label>
        <textarea id="createContestProblems" placeholder="tht26hn_cka_thieunhi&#10;tht26hn_cka_tongdayso"></textarea>
        <div class="actions">
          <button class="action primary" type="button" id="createContestButton">Táº¡o contest</button>
        </div>
      </div>

      <div class="panel" id="panel-contest-lesson-copy">
        <h2>Sao chÃ©p bÃ i tá»« Contest â†’ Lesson HNCode</h2>
        <p>Láº¥y danh sÃ¡ch bÃ i theo Ä‘Ãºng thá»© tá»± trong contest HNCode/HNOJ, chuyá»ƒn bÃ i thiáº¿u sang HNCode náº¿u cáº§n, rá»“i thÃªm vÃ o lesson HNCode. BÃ i Ä‘Ã£ cÃ³ trong lesson sáº½ Ä‘Æ°á»£c bÃ¡o rÃµ vÃ  bá» qua.</p>
        <div class="grid-2">
          <div><label>Nguá»“n contest</label><select id="lessonCopySource"><option value="hncode">HNCode</option><option value="hnoj">HNOJ</option></select><span id="lessonCopySourceLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>Lesson Ä‘Ã­ch</label><input id="lessonCopyLessonUrl" type="text" value="https://hncode.edu.vn/course/26nc202/lesson/3073"></div>
        </div>
        <div><label>HNCode Ä‘Ã­ch</label><input id="lessonCopyUserMirror" type="text" value="MrTee" readonly><span id="lessonCopyLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
        <label>Contest nguá»“n</label>
        <input id="lessonCopyContestUrl" type="text" value="https://hnoj.edu.vn/contest/ctp_4">
        <div class="actions">
          <button class="action primary" type="button" id="prepareContestLessonCopy">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmContestLessonCopy" disabled>Sao chÃ©p bÃ i</button>
        </div>
        <div class="row" style="margin-top:12px">
          <div style="max-width:180px"><label>Äiá»ƒm chung</label><input id="lessonCopyBulkScore" type="text" value="100"></div>
          <button class="action" type="button" id="fillLessonCopyScores">Ãp dá»¥ng Ä‘iá»ƒm cho táº¥t cáº£ bÃ i</button>
        </div>
        <div id="contestLessonCopyTable"></div>
      </div>

      <div class="panel" id="panel-course-clone">
        <h2>Clone Course HNCode</h2>
        <p>Clone cÃ¡c lesson vÃ  contest tá»« course nguá»“n sang course Ä‘Ã­ch. Lesson dÃ¹ng nÃºt NhÃ¢n báº£n native cá»§a HNCode; contest sáº½ táº¡o báº£n clone vá»›i mÃ£ má»›i Ä‘á»ƒ khÃ´ng Ä‘á»¥ng contest gá»‘c.</p>
        <div class="grid-2">
          <div><label>HNCode user</label><input id="courseCloneUserMirror" type="text" value="MrTee" readonly><span id="courseCloneLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>Háº­u tá»‘ mÃ£ contest Ä‘Ã­ch</label><input id="courseCloneContestSuffix" type="text" placeholder="Äá»ƒ trá»‘ng thÃ¬ tá»± dÃ¹ng _<course Ä‘Ã­ch>"></div>
        </div>
        <label>Course nguá»“n</label>
        <input id="courseCloneSourceUrl" type="text" value="https://hncode.edu.vn/course/sach_cppcoban_share">
        <label>Course Ä‘Ã­ch</label>
        <input id="courseCloneDestUrl" type="text" value="https://hncode.edu.vn/course/ngs_cpp_cb_01">
        <div class="grid-2" style="margin-top:12px">
          <label class="check"><input type="checkbox" id="courseCloneLessons" checked> Clone lesson</label>
          <label class="check"><input type="checkbox" id="courseCloneContests" checked> Clone contest</label>
        </div>
        <div class="actions">
          <button class="action primary" type="button" id="prepareCourseClone">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="confirmCourseClone" disabled>XÃ¡c nháº­n Clone Course</button>
        </div>
        <div id="courseCloneTable"></div>
      </div>

      <div class="panel" id="panel-quiz-upload">
        <h2>Up Quiz</h2>
        <p>Up danh sÃ¡ch cÃ¢u há»i lÃªn Quiz. Há»— trá»£ HNCode vÃ  TinHocTre; nhÃ£n Ä‘á»ƒ trá»‘ng.</p>
        <div class="grid-2">
          <div><label>Web Ä‘Ã­ch</label><select id="quizTarget"><option value="quiz_hncode">HNCode</option><option value="quiz_tinhoctre">TinHocTre</option></select><span id="quizLogin" class="login-badge">ChÆ°a kiá»ƒm tra</span></div>
          <div><label>TÃ i khoáº£n Ä‘ang dÃ¹ng</label><input id="quizUserMirror" type="text" value="MrTee" readonly></div>
        </div>
        <div class="grid-2" style="margin-top:12px">
          <div><label>Form táº¡o quiz</label><input id="quizCreateUrl" type="text" value="https://hncode.edu.vn/quiz/questions/create/" readonly></div>
          <div><label>File quiz Markdown/TXT</label><div class="row"><div class="grow"><input id="quizFileName" type="text" placeholder="CÃ³ thá»ƒ bá» trá»‘ng vÃ  dÃ¡n trá»±c tiáº¿p ná»™i dung bÃªn dÆ°á»›i" readonly></div><button class="action" type="button" id="chooseQuizFile">Chá»n file</button><input id="quizFileInput" class="hidden" type="file" accept=".md,.txt,text/markdown,text/plain,.zip"></div></div>
        </div>
        <div class="grid-2" style="margin-top:12px">
          <label class="check"><input type="checkbox" id="quizShuffleChoices" checked> XÃ¡o trá»™n lá»±a chá»n</label>
          <label class="check"><input type="checkbox" id="quizPublic"> CÃ´ng khai</label>
        </div>
        <label>Ná»™i dung danh sÃ¡ch quiz</label>
        <textarea id="quizMarkdown" style="min-height:340px" placeholder="DÃ¡n danh sÃ¡ch quiz theo format, hoáº·c báº¥m ChÃ¨n máº«u format."></textarea>
        <div class="actions">
          <button class="action" type="button" id="fillQuizSample">ChÃ¨n máº«u format</button>
          <button class="action primary" type="button" id="prepareQuizButton">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
          <button class="action primary" type="button" id="uploadQuizButton">Up list quiz</button>
        </div>
        <div id="quizUploadSummary"></div>
      </div>

      <div class="panel" id="panel-misc-tools">
        <h2>Tool láº»</h2>
        <p>CÃ¡c chá»©c nÄƒng phá»¥ cháº¡y á»•n Ä‘á»‹nh trÃªn local. Má»™t sá»‘ tool cÃ³ dÃ¹ng tÃ i khoáº£n OJ Ä‘Ã£ lÆ°u á»Ÿ tab TÃ i khoáº£n.</p>
        <div class="tool-card">
          <h3 class="tool-title">Láº¥y list mÃ£ bÃ i tá»« Contest / Lesson</h3>
          <p class="tool-subtitle">Láº¥y danh sÃ¡ch mÃ£ bÃ i theo Ä‘Ãºng thá»© tá»± tá»« HNCode Contest, HNCode Lesson hoáº·c HNOJ Contest. Káº¿t quáº£ cÃ³ báº£ng chi tiáº¿t vÃ  Ã´ mÃ£ bÃ i Ä‘á»ƒ copy nhanh.</p>
          <div class="grid-2">
            <div><label>Web nguá»“n</label><select id="codeListSite"><option value="hncode">HNCode</option><option value="hnoj">HNOJ</option></select></div>
            <div><label>Loáº¡i nguá»“n</label><select id="codeListType"><option value="contest">Contest</option><option value="lesson">Lesson (chá»‰ HNCode)</option></select></div>
          </div>
          <label>URL Contest / Lesson</label>
          <input id="codeListUrl" type="text" value="https://hncode.edu.vn/contest/nt26exam01">
          <div class="actions">
            <button class="action primary" type="button" id="runCodeList">Láº¥y danh sÃ¡ch mÃ£ bÃ i</button>
          </div>
          <label>Danh sÃ¡ch mÃ£ bÃ i</label>
          <textarea id="codeListOutput" rows="6" readonly placeholder="MÃ£ bÃ i sáº½ hiá»‡n á»Ÿ Ä‘Ã¢y, má»—i dÃ²ng má»™t mÃ£."></textarea>
          <div id="codeListSummary"></div>
        </div>

        <div class="tool-card">
          <h3 class="tool-title">Láº¥y last submissions Scratch</h3>
          <p class="tool-subtitle">Upload file zip data. Tool sáº½ láº¥y má»—i thÃ­ sinh 1 file `.sb3`: Æ°u tiÃªn file trong thÆ° má»¥c `$History` cÃ³ sá»‘ cuá»‘i lá»›n nháº¥t, náº¿u khÃ´ng cÃ³ thÃ¬ láº¥y file `.sb3` á»Ÿ thÆ° má»¥c gá»‘c cá»§a thÃ­ sinh.</p>
          <label>File zip data</label>
          <div class="row">
            <div class="grow"><input id="lastSubZipName" type="text" placeholder="ChÆ°a chá»n file zip" readonly></div>
            <button class="action" type="button" id="chooseLastSubZip">Chá»n file</button>
            <input id="lastSubZipFile" class="hidden" type="file" accept=".zip,application/zip">
          </div>
          <div class="actions">
            <button class="action primary" type="button" id="runLastSubmissions">Táº¡o zip last submissions</button>
          </div>
          <div id="lastSubmissionsSummary"></div>
        </div>

        <div class="tool-card">
          <h3 class="tool-title">Cháº¥m bÃ i HNCode</h3>
          <p class="tool-subtitle">Äá»c file zip bÃ i lÃ m theo dáº¡ng <code>BaiLam/TenHocSinh/MABAI.cpp</code>, Ä‘á»c file CSV tÃ i khoáº£n cÃ³ cá»™t <code>username,password,name</code>, Ä‘Äƒng nháº­p tá»«ng tÃ i khoáº£n, tham gia contest vÃ  ná»™p cÃ¡c bÃ i tÆ°Æ¡ng á»©ng. Káº¿t quáº£ xuáº¥t ra Excel.</p>
          <label>File zip bÃ i lÃ m cá»§a há»c sinh</label>
          <div class="row">
            <div class="grow"><input id="gradingZipName" type="text" placeholder="ChÆ°a chá»n file BaiLam.zip" readonly></div>
            <button class="action" type="button" id="chooseGradingZip">Chá»n file</button>
            <input id="gradingZipFile" class="hidden" type="file" accept=".zip,application/zip">
          </div>
          <label>File CSV tÃ i khoáº£n ná»™p bÃ i</label>
          <div class="row">
            <div class="grow"><input id="gradingCsvName" type="text" placeholder="ChÆ°a chá»n file TaiKhoan.csv" readonly></div>
            <button class="action" type="button" id="chooseGradingCsv">Chá»n file</button>
            <input id="gradingCsvFile" class="hidden" type="file" accept=".csv,text/csv">
          </div>
          <div class="grid-2">
            <div><label>URL contest HNCode</label><input id="gradingContestUrl" type="text" value="https://hncode.edu.vn/contest/_nt26tst"></div>
            <div><label>Máº­t kháº©u contest náº¿u cÃ³</label><input id="gradingContestPassword" type="password" value="amsvodich*8*^^"></div>
          </div>
          <div class="grid-2">
            <div><label>Thá»i gian chá» má»—i submission</label><input id="gradingPollSeconds" type="text" value="Äáº¿n khi cháº¥m xong" readonly></div>
            <div><label>Quy Ä‘á»•i Ä‘iá»ƒm</label><input type="text" value="% cháº¥m x Ä‘iá»ƒm contest" readonly></div>
          </div>
          <div class="actions">
            <button class="action primary" type="button" id="prepareGrading">Chuáº©n bá»‹ dá»¯ liá»‡u</button>
            <button class="action primary" type="button" id="confirmGrading" disabled>XÃ¡c nháº­n ná»™p vÃ  cháº¥m</button>
            <a class="action primary hidden" id="downloadGradingResult" href="#" download="bang_diem_hncode.xlsx">Táº£i báº£ng Ä‘iá»ƒm Excel</a>
          </div>
          <div id="gradingSummary"></div>
        </div>

        <div class="tool-card">
          <h3 class="tool-title">Cáº£nh bÃ¡o sá»­ dá»¥ng AI Ä‘á»ƒ code</h3>
          <p class="tool-subtitle">Nháº­n vÃ o má»™t folder chá»©a nhiá»u file zip contest hoáº·c chá»n má»™t file zip data contest. Tool phÃ¢n tÃ­ch dáº¥u hiá»‡u AI code, Ä‘á»•i phong cÃ¡ch code vÃ  nghi váº¥n chÃ©p code nhau, rá»“i xuáº¥t Excel cÃ³ link má»Ÿ file code.</p>
          <label>Folder chá»©a cÃ¡c zip contest</label>
          <input id="aiWarningFolder" type="text" value="{{ ai_source_default }}">
          <label>Hoáº·c chá»n 1 file zip data contest</label>
          <div class="row">
            <div class="grow"><input id="aiWarningZipName" type="text" placeholder="KhÃ´ng chá»n thÃ¬ dÃ¹ng folder á»Ÿ trÃªn" readonly></div>
            <button class="action" type="button" id="chooseAiWarningZip">Chá»n file zip</button>
            <input id="aiWarningZipFile" class="hidden" type="file" accept=".zip,application/zip">
          </div>
          <p>ÄÃ¢y lÃ  bÃ¡o cÃ¡o cáº£nh bÃ¡o/nghi váº¥n, khÃ´ng pháº£i káº¿t luáº­n cháº¯c cháº¯n. NÃªn má»Ÿ cÃ¡c file máº«u trong Excel Ä‘á»ƒ kiá»ƒm tra láº¡i.</p>
          <div class="actions">
            <button class="action primary" type="button" id="runAiWarning">Táº¡o bÃ¡o cÃ¡o Excel</button>
          </div>
          <div id="aiWarningSummary"></div>
        </div>
      </div>
    </section>

    <section class="log-panel">
      <div class="log-head"><h2>ThÃ´ng tin tráº£ vá»</h2><span id="jobStatus" class="status">idle</span></div>
      <pre id="log">Sáºµn sÃ ng.</pre>
    </section>
  </main>

<script>
const TARGETS = {{ targets_json | safe }};
const QUIZ_TARGETS = {{ quiz_targets_json | safe }};
let preparedUpload = null;
let preparedSingleUpload = null;
let preparedTransfer = null;
let preparedContestTransfer = null;
let preparedQuiz = null;
let preparedContestLessonCopy = null;
let preparedCourseClone = null;
let preparedGrading = null;
let selectedZipFile = null;
let selectedSingleTestZipFile = null;
let selectedGradingZipFile = null;
let selectedGradingCsvFile = null;
const QUIZ_FORMAT_GUIDE = {{ quiz_format_guide_json | safe }};

const logEl = document.getElementById("log");
const statusEl = document.getElementById("jobStatus");
let logText = "Sáºµn sÃ ng.";
const progressTimers = new Map();
function colorizeLog(text) {
  return String(text).split("\n").map(line => {
    const trimmed = line.trim();
    let cls = "";
    if (trimmed.startsWith("âœ“") || trimmed.includes("ThÃ nh cÃ´ng") || trimmed.includes("ÄÃ£ táº¡o") || trimmed.includes("ÄÃ£ upload")) cls = "log-ok";
    else if (trimmed.startsWith("âœ—") || trimmed.startsWith("Error:") || trimmed.includes("Lá»—i")) cls = "log-err";
    else if (trimmed.includes("Ä‘Ã£ tá»“n táº¡i") || trimmed.includes("ÄÃ£ tá»“n táº¡i") || trimmed.includes("BÃ i Ä‘Ã£ tá»“n táº¡i") || trimmed.includes("Contest Ä‘Ã£ tá»“n táº¡i")) cls = "log-warn";
    else if (trimmed.startsWith("Tiáº¿n Ä‘á»™:") || trimmed.startsWith("Äang ")) cls = "log-progress";
    const safe = escapeHtml(line);
    return cls ? `<span class="${cls}">${safe}</span>` : safe;
  }).join("\n");
}
function renderLog() { logEl.innerHTML = colorizeLog(logText); logEl.scrollTop = logEl.scrollHeight; }
function log(text) { logText = String(text); renderLog(); }
function append(text) { logText += "\n" + String(text); renderLog(); }
function status(text, cls="") { statusEl.textContent = text; statusEl.className = "status " + cls; }

for (const button of document.querySelectorAll(".nav button")) {
  button.addEventListener("click", () => {
    document.querySelectorAll(".nav button").forEach(item => item.classList.remove("active"));
    document.querySelectorAll(".panel").forEach(item => item.classList.remove("active"));
    button.classList.add("active");
    document.getElementById("panel-" + button.dataset.panel).classList.add("active");
  });
}

const accountFields = {
  hnoj_user: document.getElementById("acct_hnoj_user"),
  hnoj_pass: document.getElementById("acct_hnoj_pass"),
  hncode_user: document.getElementById("acct_hncode_user"),
  hncode_pass: document.getElementById("acct_hncode_pass"),
  tinhoctre_user: document.getElementById("acct_tinhoctre_user"),
  tinhoctre_pass: document.getElementById("acct_tinhoctre_pass"),
  tinhoctre_cookie: document.getElementById("acct_tinhoctre_cookie"),
};
function loadAccounts() {
  for (const [key, input] of Object.entries(accountFields)) {
    const value = localStorage.getItem("chuyenbai." + key);
    if (value !== null) input.value = value;
  }
}
function saveAccounts() {
  for (const [key, input] of Object.entries(accountFields)) localStorage.setItem("chuyenbai." + key, input.value);
}
loadAccounts();
document.getElementById("saveAccounts").onclick = () => { saveAccounts(); append("ÄÃ£ lÆ°u táº¡m tÃ i khoáº£n."); };
document.getElementById("checkAccounts").onclick = () => { log("Äang kiá»ƒm tra Ä‘Äƒng nháº­p cÃ¡c trang..."); checkAllAccounts(); };
const openTinHocTreBrowserButton = document.getElementById("openTinHocTreBrowser");
if (openTinHocTreBrowserButton) openTinHocTreBrowserButton.onclick = async () => {
  try {
    status("running");
    const data = await postJson("/api/tinhoctre-browser/start", {});
    append(data.message || "ÄÃ£ má»Ÿ Edge Ä‘Äƒng nháº­p TinHocTre.");
    status("ready", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};
const pullTinHocTreCookieButton = document.getElementById("pullTinHocTreCookie");
if (pullTinHocTreCookieButton) pullTinHocTreCookieButton.onclick = async () => {
  try {
    status("running");
    const data = await postJson("/api/tinhoctre-browser/cookie", {});
    accountFields.tinhoctre_cookie.value = data.cookie || "";
    saveAccounts();
    append(data.message || "ÄÃ£ láº¥y vÃ  lÆ°u Cookie TinHocTre.");
    await checkLogin("tinhoctre", "login_tinhoctre", firstToken(document.getElementById("transferCodes").value));
    status("ready", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};
const quickTinHocTreCookieButton = document.getElementById("quickTinHocTreCookie");
if (quickTinHocTreCookieButton) quickTinHocTreCookieButton.onclick = async () => {
  try {
    status("running");
    append("Äang Ä‘Ã³ng/má»Ÿ láº¡i Edge Ä‘á»ƒ láº¥y cookie TinHocTre...");
    const data = await postJson("/api/tinhoctre-browser/quick-cookie", {});
    accountFields.tinhoctre_cookie.value = data.cookie || "";
    saveAccounts();
    append(data.message || "ÄÃ£ láº¥y vÃ  lÆ°u Cookie TinHocTre tá»« Edge.");
    await checkLogin("tinhoctre", "login_tinhoctre", firstToken(document.getElementById("transferCodes").value));
    status("ready", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};
document.getElementById("clearAccounts").onclick = () => {
  for (const key of Object.keys(accountFields)) localStorage.removeItem("chuyenbai." + key);
  for (const [key, input] of Object.entries(accountFields)) if (key.endsWith("_pass") || key.endsWith("_cookie")) input.value = "";
  append("ÄÃ£ xÃ³a thÃ´ng tin Ä‘Ã£ lÆ°u.");
};
document.getElementById("toggleGuide").onclick = () => document.getElementById("promptGuide").classList.toggle("hidden");
document.getElementById("toggleAdvanced").onclick = () => {
  const box = document.getElementById("advancedUpload");
  box.classList.toggle("hidden");
  document.getElementById("toggleAdvanced").textContent = box.classList.contains("hidden") ? "Má»Ÿ rá»™ng thÃ´ng tin khÃ¡c" : "Thu gá»n thÃ´ng tin khÃ¡c";
};
document.getElementById("toggleTransferAdvanced").onclick = () => {
  const box = document.getElementById("advancedTransfer");
  box.classList.toggle("hidden");
  document.getElementById("toggleTransferAdvanced").textContent = box.classList.contains("hidden") ? "Má»Ÿ rá»™ng thÃ´ng tin khÃ¡c" : "Thu gá»n thÃ´ng tin khÃ¡c";
};
document.getElementById("applyTransferLimits").onclick = () => {
  const timeLimit = document.getElementById("transferTimeLimit").value;
  const memoryLimit = document.getElementById("transferMemoryLimit").value;
  for (const tr of document.querySelectorAll("#transferTable tbody tr")) {
    const timeInput = tr.querySelector(".row-time");
    const memoryInput = tr.querySelector(".row-memory");
    if (timeInput) timeInput.value = timeLimit;
    if (memoryInput) memoryInput.value = memoryLimit;
  }
  append("ÄÃ£ Ã¡p dá»¥ng time/memory máº·c Ä‘á»‹nh cho táº¥t cáº£ bÃ i trong báº£ng chuyá»ƒn.");
};
document.getElementById("resetTransferLimits").onclick = () => {
  for (const tr of document.querySelectorAll("#transferTable tbody tr")) {
    const timeInput = tr.querySelector(".row-time");
    const memoryInput = tr.querySelector(".row-memory");
    if (timeInput) timeInput.value = tr.dataset.sourceTime || "1.0";
    if (memoryInput) memoryInput.value = tr.dataset.sourceMemory || "1048576";
  }
  append("ÄÃ£ tráº£ time/memory vá» thÃ´ng sá»‘ láº¥y tá»« nguá»“n.");
};
function localDateTimeValue(date) {
  const pad = value => String(value).padStart(2, "0");
  return `${date.getFullYear()}-${pad(date.getMonth() + 1)}-${pad(date.getDate())}T${pad(date.getHours())}:${pad(date.getMinutes())}`;
}
function backendDateTimeValue(value) {
  if (!value) return "";
  return value.length === 16 ? value.replace("T", " ") + ":00" : value.replace("T", " ");
}
function setContestTime(dayOffset=0, startHour=8, startMinute=0, durationMinutes=180) {
  const start = new Date();
  start.setDate(start.getDate() + dayOffset);
  start.setHours(startHour, startMinute, 0, 0);
  const end = new Date(start.getTime() + durationMinutes * 60000);
  document.getElementById("createContestStart").value = localDateTimeValue(start);
  document.getElementById("createContestEnd").value = localDateTimeValue(end);
}
document.getElementById("contestTimeToday").onclick = () => setContestTime(0, 8, 0, 180);
document.getElementById("contestTimeTomorrow").onclick = () => setContestTime(1, 8, 0, 180);
document.getElementById("contestTime90").onclick = () => {
  const startInput = document.getElementById("createContestStart");
  const start = startInput.value ? new Date(startInput.value) : new Date();
  document.getElementById("createContestStart").value = localDateTimeValue(start);
  document.getElementById("createContestEnd").value = localDateTimeValue(new Date(start.getTime() + 90 * 60000));
};
document.getElementById("chooseZip").onclick = () => document.getElementById("zipFileInput").click();
document.getElementById("zipFileInput").onchange = event => {
  selectedZipFile = event.target.files[0] || null;
  if (selectedZipFile) document.getElementById("uploadZip").value = selectedZipFile.name;
};
document.getElementById("useBatchSample").onclick = async () => {
  selectedZipFile = null;
  document.getElementById("zipFileInput").value = "";
  const data = await postJson("/api/sample/tonghaiso", {});
  document.getElementById("uploadZip").value = data.zip_path;
  append("ÄÃ£ Ä‘iá»n file máº«u Tá»•ng hai sá»‘ cho Up nhiá»u bÃ i.");
};
function toggleBox(buttonId, boxId, openText, closedText) {
  const box = document.getElementById(boxId);
  box.classList.toggle("hidden");
  document.getElementById(buttonId).textContent = box.classList.contains("hidden") ? openText : closedText;
}
document.getElementById("toggleSingleStatement").onclick = () => toggleBox("toggleSingleStatement", "singleStatementBox", "Má»Ÿ Ä‘á» bÃ i", "Thu gá»n Ä‘á» bÃ i");
document.getElementById("toggleSingleGenerator").onclick = () => toggleBox("toggleSingleGenerator", "singleGeneratorBox", "Má»Ÿ sinh test", "Thu gá»n sinh test");
document.getElementById("toggleSingleSolution").onclick = () => toggleBox("toggleSingleSolution", "singleSolutionBox", "Má»Ÿ lá»i giáº£i", "Thu gá»n lá»i giáº£i");
document.getElementById("chooseSingleStatement").onclick = () => document.getElementById("singleStatementFile").click();
document.getElementById("chooseSingleGenerator").onclick = () => document.getElementById("singleGeneratorFile").click();
document.getElementById("chooseSingleTestZip").onclick = () => document.getElementById("singleTestZipFile").click();
document.getElementById("chooseSingleSolution").onclick = () => document.getElementById("singleSolutionFile").click();
document.getElementById("useSingleSample").onclick = async () => {
  const data = await postJson("/api/sample/tonghaiso", {});
  document.getElementById("singleCode").value = data.code;
  document.getElementById("singleName").value = data.name;
  document.getElementById("singlePoints").value = data.points || "100";
  document.getElementById("singleTags").value = data.tags || "";
  document.getElementById("singleTimeLimit").value = "1.0";
  document.getElementById("singleMemoryLimit").value = "1024M";
  document.getElementById("singlePartial").checked = true;
  document.getElementById("singleStatement").value = data.statement || "";
  document.getElementById("singleGenerator").value = data.generator || "";
  document.getElementById("singleGeneratorName").value = "gentest_" + data.code + ".py";
  document.getElementById("singleSolution").value = data.solution_md || "";
  selectedSingleTestZipFile = null;
  document.getElementById("singleTestZipFile").value = "";
  document.getElementById("singleTestZipName").value = "CÃ³ zip test trong máº«u; Up 1 bÃ i sáº½ sinh tá»« gentest";
  append("ÄÃ£ náº¡p máº«u Tá»•ng hai sá»‘ vÃ o Up 1 bÃ i. Báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u Ä‘á»ƒ kiá»ƒm tra.");
};
document.getElementById("singleStatementFile").addEventListener("change", async event => {
  const file = event.target.files && event.target.files[0];
  if (file) document.getElementById("singleStatement").value = await file.text();
});
document.getElementById("singleGeneratorFile").addEventListener("change", async event => {
  const file = event.target.files && event.target.files[0];
  document.getElementById("singleGeneratorName").value = file ? file.name : "";
  if (file) document.getElementById("singleGenerator").value = await file.text();
});
document.getElementById("singleTestZipFile").addEventListener("change", event => {
  selectedSingleTestZipFile = event.target.files && event.target.files[0] || null;
  document.getElementById("singleTestZipName").value = selectedSingleTestZipFile ? selectedSingleTestZipFile.name : "";
});
document.getElementById("singleSolutionFile").addEventListener("change", async event => {
  const file = event.target.files && event.target.files[0];
  if (file) document.getElementById("singleSolution").value = await file.text();
});

function renderLanguages() {
  const target = document.getElementById("uploadTarget").value;
  const langs = TARGETS[target].languages;
  document.getElementById("languages").innerHTML = Object.keys(langs).map(name =>
    `<label class="check"><input type="checkbox" value="${name}" checked> ${name}</label>`
  ).join("");
}
function renderTransferLanguages() {
  const target = document.getElementById("transferDest").value;
  const langs = TARGETS[target].languages;
  document.getElementById("transferLanguages").innerHTML = Object.keys(langs).map(name =>
    `<label class="check"><input type="checkbox" value="${name}" checked> ${name}</label>`
  ).join("");
}
function renderSingleLanguages() {
  const target = document.getElementById("singleUploadTarget").value;
  const langs = TARGETS[target].languages;
  document.getElementById("singleLanguages").innerHTML = Object.keys(langs).map(name =>
    `<label class="check"><input type="checkbox" value="${name}" checked> ${name}</label>`
  ).join("");
}
document.getElementById("uploadTarget").addEventListener("change", renderLanguages);
document.getElementById("singleUploadTarget").addEventListener("change", renderSingleLanguages);
document.getElementById("transferDest").addEventListener("change", renderTransferLanguages);
document.getElementById("uploadTarget").addEventListener("change", checkUploadLogin);
document.getElementById("singleUploadTarget").addEventListener("change", checkSingleUploadLogin);
document.getElementById("transferSource").addEventListener("change", checkTransferLogins);
document.getElementById("transferDest").addEventListener("change", checkTransferLogins);
document.getElementById("transferCodes").addEventListener("blur", checkTransferLogins);
document.getElementById("contestSource").addEventListener("change", checkContestLogins);
document.getElementById("contestDest").addEventListener("change", checkContestLogins);
document.getElementById("contestCodes").addEventListener("blur", checkContestLogins);
document.getElementById("createContestTarget").addEventListener("change", checkCreateContestLogin);
document.getElementById("lessonCopySource").addEventListener("change", checkLessonCopyLogin);
document.getElementById("lessonCopyContestUrl").addEventListener("blur", () => {
  const value = document.getElementById("lessonCopyContestUrl").value.toLowerCase();
  if (value.includes("hnoj.edu.vn")) document.getElementById("lessonCopySource").value = "hnoj";
  if (value.includes("hncode.edu.vn") || value.includes("oj.hncode.edu.vn")) document.getElementById("lessonCopySource").value = "hncode";
  checkLessonCopyLogin();
});
renderLanguages();
renderSingleLanguages();
renderTransferLanguages();
setTimeout(() => { checkUploadLogin(); checkSingleUploadLogin(); checkTransferLogins(); checkContestLogins(); checkCreateContestLogin(); checkQuizLogin(); checkLessonCopyLogin(); checkCourseCloneLogin(); }, 300);

function selectedLanguages() {
  return [...document.querySelectorAll("#languages input:checked")].map(item => item.value);
}
function selectedSingleLanguages() {
  return [...document.querySelectorAll("#singleLanguages input:checked")].map(item => item.value);
}
function selectedTransferLanguages() {
  return [...document.querySelectorAll("#transferLanguages input:checked")].map(item => item.value);
}
function accountPayload(target) {
  const payload = {
    username: accountFields[target + "_user"].value,
    password: accountFields[target + "_pass"].value,
  };
  if (target === "tinhoctre") payload.cookie = accountFields.tinhoctre_cookie.value;
  return payload;
}
function firstToken(value) {
  return (value || "").split(/[\s,]+/).filter(Boolean)[0] || "";
}
function setLoginBadge(id, state, text) {
  const el = document.getElementById(id);
  if (!el) return;
  el.textContent = text;
  el.className = "login-badge " + (state || "");
}
async function checkLogin(target, badgeId, probeCode="") {
  setLoginBadge(badgeId, "", "Äang kiá»ƒm tra...");
  try {
    const accountTarget = QUIZ_TARGETS[target] ? QUIZ_TARGETS[target].account_target : target;
    const data = await postJson("/api/check-login", {target, account: accountPayload(accountTarget), probe_code: probeCode});
    setLoginBadge(badgeId, data.ok ? "ok" : "err", data.ok ? "âœ“ ÄÄƒng nháº­p OK" : "âœ— " + (data.message || "Lá»—i"));
    return data.ok;
  } catch (err) {
    setLoginBadge(badgeId, "err", "âœ— " + String(err).replace(/^Error:\s*/, ""));
    return false;
  }
}
function quizAccountTarget() {
  const target = document.getElementById("quizTarget").value;
  return (QUIZ_TARGETS[target] && QUIZ_TARGETS[target].account_target) || "hncode";
}
function quizAccountPayload() {
  return accountPayload(quizAccountTarget());
}
function updateQuizTargetUi() {
  const target = document.getElementById("quizTarget").value;
  const info = QUIZ_TARGETS[target] || QUIZ_TARGETS.quiz_hncode;
  const accountTarget = info.account_target || "hncode";
  document.getElementById("quizUserMirror").value = accountFields[accountTarget + "_user"].value || info.default_user || TARGETS[accountTarget]?.default_user || "";
  document.getElementById("quizCreateUrl").value = (info.base_url || "") + "/quiz/questions/create/";
}
async function checkAllAccounts() {
  saveAccounts();
  await Promise.all([
    checkLogin("hnoj", "login_hnoj"),
    checkLogin("hncode", "login_hncode"),
    checkLogin("tinhoctre", "login_tinhoctre", firstToken(document.getElementById("transferCodes").value)),
  ]);
}
function checkUploadLogin() {
  checkLogin(document.getElementById("uploadTarget").value, "uploadTargetLogin");
}
function checkSingleUploadLogin() {
  checkLogin(document.getElementById("singleUploadTarget").value, "singleUploadLogin");
}
function checkTransferLogins() {
  const probe = firstToken(document.getElementById("transferCodes").value);
  checkLogin(document.getElementById("transferSource").value, "transferSourceLogin", probe);
  checkLogin(document.getElementById("transferDest").value, "transferDestLogin");
}
function checkContestLogins() {
  checkLogin(document.getElementById("contestSource").value, "contestSourceLogin");
  checkLogin(document.getElementById("contestDest").value, "contestDestLogin");
}
function checkCreateContestLogin() {
  checkLogin(document.getElementById("createContestTarget").value, "createContestLogin");
}
function checkQuizLogin() {
  updateQuizTargetUi();
  checkLogin(document.getElementById("quizTarget").value, "quizLogin");
}
function checkLessonCopyLogin() {
  document.getElementById("lessonCopyUserMirror").value = accountFields.hncode_user.value || "MrTee";
  const source = document.getElementById("lessonCopySource").value;
  checkLogin(source, "lessonCopySourceLogin");
  checkLogin("hncode", "lessonCopyLogin");
}
function checkCourseCloneLogin() {
  document.getElementById("courseCloneUserMirror").value = accountFields.hncode_user.value || "MrTee";
  checkLogin("hncode", "courseCloneLogin");
}
function uploadSettings() {
  const target = document.getElementById("uploadTarget").value;
  return {
    target,
    zip_path: selectedZipFile ? "" : document.getElementById("uploadZip").value,
    creator: document.getElementById("creator").value,
    points: document.getElementById("uploadPoints").value.trim() || "100",
    tags: document.getElementById("uploadTags").value.trim(),
    partial: document.getElementById("uploadPartial").checked,
    time_limit: document.getElementById("timeLimit").value,
    memory_limit: document.getElementById("memoryLimit").value,
    languages: selectedLanguages(),
    no_submit: document.getElementById("noSubmit").checked,
    submit_cpp: document.getElementById("submitCpp").checked,
    submit_python: document.getElementById("submitPython").checked,
    skip_statement_title: document.getElementById("skipStatementTitle").checked,
    overwrite_existing: document.getElementById("overwriteExisting").checked,
    overwrite_statement: document.getElementById("overwriteStatement").checked,
    overwrite_tests: document.getElementById("overwriteTests").checked,
    ...accountPayload(target),
  };
}
function singleUploadSettings() {
  const target = document.getElementById("singleUploadTarget").value;
  return {
    target,
    code: document.getElementById("singleCode").value.trim(),
    name: document.getElementById("singleName").value.trim(),
    points: document.getElementById("singlePoints").value.trim() || "100",
    tags: document.getElementById("singleTags").value.trim(),
    time_limit: document.getElementById("singleTimeLimit").value.trim() || "1.0",
    memory_limit: document.getElementById("singleMemoryLimit").value.trim() || "1024M",
    partial: document.getElementById("singlePartial").checked,
    overwrite_statement: document.getElementById("singleOverwrite").checked,
    overwrite_tests: document.getElementById("singleOverwrite").checked,
    languages: selectedSingleLanguages(),
    skip_statement_title: document.getElementById("singleSkipStatementTitle").checked,
    statement_text: document.getElementById("singleStatement").value,
    generator_text: document.getElementById("singleGenerator").value,
    generator_filename: document.getElementById("singleGeneratorName").value,
    solution_text: document.getElementById("singleSolution").value,
    upload_solution: Boolean(document.getElementById("singleSolution").value.trim()),
    no_submit: true,
    ...accountPayload(target),
  };
}
function transferSettings() {
  const dest = document.getElementById("transferDest").value;
  return {
    creator: document.getElementById("transferCreator").value,
    time_limit: document.getElementById("transferTimeLimit").value,
    memory_limit: document.getElementById("transferMemoryLimit").value,
    languages: selectedTransferLanguages(),
    ...accountPayload(dest),
  };
}
async function postJson(url, payload) {
  const res = await fetch(url, {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify(payload)});
  const data = await parseJsonResponse(res);
  if (!res.ok) throw new Error(data.error || "Request failed");
  return data;
}
async function parseJsonResponse(res) {
  const text = await res.text();
  try {
    return text ? JSON.parse(text) : {};
  } catch (err) {
    const preview = text.replace(/<[^>]*>/g, " ").replace(/\s+/g, " ").trim().slice(0, 300);
    throw new Error(`Server tráº£ vá» HTML/text thay vÃ¬ JSON (HTTP ${res.status}). ${preview || "KhÃ´ng cÃ³ ná»™i dung lá»—i."}`);
  }
}
async function prepareUploadRequest(settings) {
  if (!selectedZipFile) return postJson("/api/prepare-upload", settings);
  const form = new FormData();
  form.append("zip_file", selectedZipFile);
  form.append("payload", JSON.stringify(settings));
  const res = await fetch("/api/prepare-upload", {method:"POST", body:form});
  const data = await parseJsonResponse(res);
  if (!res.ok) throw new Error(data.error || "Request failed");
  return data;
}
async function prepareSingleUploadRequest(settings) {
  const form = new FormData();
  if (selectedSingleTestZipFile) form.append("test_zip", selectedSingleTestZipFile);
  form.append("payload", JSON.stringify(settings));
  const res = await fetch("/api/prepare-single-upload", {method:"POST", body:form});
  const data = await parseJsonResponse(res);
  if (!res.ok) throw new Error(data.error || "Request failed");
  return data;
}
function newProgressId() {
  if (window.crypto && crypto.randomUUID) return crypto.randomUUID().replaceAll("-", "");
  return Array.from({length: 32}, () => Math.floor(Math.random() * 16).toString(16)).join("");
}
function statusClass(text) {
  const value = String(text || "");
  if (value.startsWith("âœ“") || value.includes("ThÃ nh cÃ´ng") || value.includes("ÄÃ£ Ä‘á»c")) return "ok";
  if (value.includes("Ä‘Ã£ tá»“n táº¡i") || value.includes("ÄÃ£ tá»“n táº¡i") || value.includes("Ä‘Ã£ cÃ³") || value.includes("ÄÃ£ cÃ³")) return "warn";
  if (value.startsWith("âœ—") || value.includes("Lá»—i")) return "err";
  return "";
}
function setStatusCell(cell, text, link="") {
  cell.className = "row-status " + statusClass(text);
  const linkHtml = link ? ` <a class="problem-link" href="${escapeHtml(link)}" target="_blank" rel="noopener">Link</a>` : "";
  cell.innerHTML = `${escapeHtml(text || "")}${linkHtml}`;
}
function progressMessage(data) {
  const total = data.total || 0;
  const done = data.done || 0;
  const prefix = total ? `Tiáº¿n Ä‘á»™: ${done}/${total}` : "Tiáº¿n Ä‘á»™:";
  return data.message ? `${prefix} - ${data.message}` : prefix;
}
function startProgressPolling(progressId, tableSelector, mode="problem") {
  stopProgressPolling(progressId);
  const timer = setInterval(async () => {
    try {
      const res = await fetch(`/api/progress/${progressId}`, {cache: "no-store"});
      if (!res.ok) return;
      const data = await res.json();
      if (data.rows) {
        if (mode === "contest") applyContestStatuses(data.rows);
        else if (mode === "grading") applyGradingStatuses(data.rows);
        else if (tableSelector) applyStatuses(data.rows, tableSelector);
      }
      if (data.message || data.total) append(progressMessage(data));
      if (data.finished) stopProgressPolling(progressId);
    } catch (err) {
      stopProgressPolling(progressId);
    }
  }, 1000);
  progressTimers.set(progressId, timer);
  return progressId;
}
function stopProgressPolling(progressId) {
  const timer = progressTimers.get(progressId);
  if (timer) clearInterval(timer);
  progressTimers.delete(progressId);
}

document.getElementById("prepareUpload").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang chuáº©n bá»‹ dá»¯ liá»‡u...");
    startProgressPolling(progressId, "#uploadTable");
    const settings = uploadSettings();
    settings.progress_id = progressId;
    const data = await prepareUploadRequest(settings);
    stopProgressPolling(progressId);
    preparedUpload = data.prepare_id;
    renderUploadTable(data.rows);
    document.getElementById("confirmUpload").disabled = false;
    log(data.log);
    status("ready", "ok");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

function renderUploadTable(rows) {
  const overwriteDefault = document.getElementById("overwriteExisting").checked;
  document.getElementById("uploadTable").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#uploadTable', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#uploadTable', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i toÃ¡n</th><th>Äiá»ƒm</th><th>Dáº¡ng bÃ i táº­p / Tags</th><th>Time</th><th>Memory</th><th>Äiá»ƒm thÃ nh pháº§n</th><th>Ghi Ä‘Ã¨</th><th>Up Ä‘á»</th><th>Up test</th><th>Up lá»i giáº£i</th><th>File test</th><th>Sá»‘ test</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-original="${escapeHtml(row.original_code)}" data-source-time="${escapeHtml(row.source_time_limit || row.time_limit || "1.0")}" data-source-memory="${escapeHtml(row.source_memory_limit || row.memory_limit || "1048576")}">
      <td><input type="checkbox" class="row-selected" checked></td>
      <td><input type="text" class="row-code" value="${escapeHtml(row.code)}"></td>
      <td><input type="text" class="row-name" value="${escapeHtml(row.name)}"></td>
      <td><input type="text" class="row-points" value="${escapeHtml(row.points || "100")}"></td>
      <td><input type="text" class="row-tags" value="${escapeHtml(row.tags || "")}"></td>
      <td><input type="text" class="row-time" value="${escapeHtml(row.time_limit || "1.0")}"></td>
      <td><input type="text" class="row-memory" value="${escapeHtml(row.memory_limit || "1048576")}"></td>
      <td><input type="checkbox" class="row-partial" ${row.partial === false ? "" : "checked"}></td>
      <td><input type="checkbox" class="row-overwrite" ${row.overwrite_default === true || overwriteDefault ? "checked" : ""}></td>
      <td><input type="checkbox" class="row-statement" checked></td>
      <td><input type="checkbox" class="row-tests" ${row.upload_tests_default === false ? "" : "checked"}></td>
      <td><input type="checkbox" class="row-solution" ${row.upload_solution_default ? "checked" : ""}></td>
      <td><div class="test-meta">${escapeHtml(row.test_file)}</div></td>
      <td>${row.test_count}</td>
      <td class="row-status">ChÆ°a up</td>
    </tr>`).join("")}</tbody></table>`;
}
function collectUploadRows() {
  return [...document.querySelectorAll("#uploadTable tbody tr")].map(tr => ({
    original_code: tr.dataset.original,
    selected: tr.querySelector(".row-selected").checked,
    code: tr.querySelector(".row-code").value.trim(),
    name: tr.querySelector(".row-name").value.trim(),
    points: tr.querySelector(".row-points").value.trim(),
    tags: tr.querySelector(".row-tags").value.trim(),
    time_limit: tr.querySelector(".row-time").value.trim(),
    memory_limit: tr.querySelector(".row-memory").value.trim(),
    partial: tr.querySelector(".row-partial").checked,
    overwrite: tr.querySelector(".row-overwrite").checked,
    upload_statement: tr.querySelector(".row-statement").checked,
    upload_tests: tr.querySelector(".row-tests").checked,
    upload_solution: tr.querySelector(".row-solution").checked,
  }));
}
document.getElementById("confirmUpload").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang up bÃ i...");
    markRowsProcessing("#uploadTable", "Äang up...");
    startProgressPolling(progressId, "#uploadTable");
    const settings = uploadSettings();
    settings.progress_id = progressId;
    const data = await postJson("/api/confirm-upload", {prepare_id: preparedUpload, settings, rows: collectUploadRows(), progress_id: progressId});
    stopProgressPolling(progressId);
    applyStatuses(data.rows, "#uploadTable");
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("prepareSingleUpload").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang chuáº©n bá»‹ dá»¯ liá»‡u 1 bÃ i...");
    const settings = singleUploadSettings();
    settings.progress_id = progressId;
    const data = await prepareSingleUploadRequest(settings);
    preparedSingleUpload = data.prepare_id;
    renderSingleUploadTable(data.rows || []);
    document.getElementById("confirmSingleUpload").disabled = false;
    log(data.log);
    status("ready", "ok");
  } catch (err) {
    preparedSingleUpload = null;
    document.getElementById("confirmSingleUpload").disabled = true;
    log(String(err));
    status("failed", "err");
  }
};

function renderSingleUploadTable(rows) {
  document.getElementById("singleUploadTable").innerHTML = `<table>
    <thead><tr><th>Chá»n</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i toÃ¡n</th><th>Äiá»ƒm</th><th>Dáº¡ng bÃ i táº­p / Tags</th><th>Time</th><th>Memory</th><th>Äiá»ƒm thÃ nh pháº§n</th><th>Up Ä‘á»</th><th>Up test</th><th>Up lá»i giáº£i</th><th>Test</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-original="${escapeHtml(row.original_code)}">
      <td><input type="checkbox" class="row-selected" checked></td>
      <td><input type="text" class="row-code" value="${escapeHtml(row.code)}"></td>
      <td><input type="text" class="row-name" value="${escapeHtml(row.name)}"></td>
      <td><input type="text" class="row-points" value="${escapeHtml(row.points || "100")}"></td>
      <td><input type="text" class="row-tags" value="${escapeHtml(row.tags || "")}"></td>
      <td><input type="text" class="row-time" value="${escapeHtml(row.time_limit || "1.0")}"></td>
      <td><input type="text" class="row-memory" value="${escapeHtml(row.memory_limit || "1024M")}"></td>
      <td><input type="checkbox" class="row-partial" ${row.partial === false ? "" : "checked"}></td>
      <td><input type="checkbox" class="row-statement" ${row.upload_statement_default ? "checked" : ""}></td>
      <td><input type="checkbox" class="row-tests" ${row.upload_tests_default ? "checked" : ""}></td>
      <td><input type="checkbox" class="row-solution" ${row.upload_solution_default ? "checked" : ""}></td>
      <td><div class="test-meta">${escapeHtml(row.test_file || "KhÃ´ng cÃ³ test")}<br>${escapeHtml(row.test_count || 0)} test</div></td>
      <td class="row-status ${statusClass(row.status)}">${escapeHtml(row.status || "ÄÃ£ chuáº©n bá»‹")}</td>
    </tr>`).join("")}</tbody></table>`;
}

function collectSingleUploadRows() {
  return [...document.querySelectorAll("#singleUploadTable tbody tr")].map(tr => ({
    original_code: tr.dataset.original,
    selected: tr.querySelector(".row-selected").checked,
    code: tr.querySelector(".row-code").value.trim(),
    name: tr.querySelector(".row-name").value.trim(),
    points: tr.querySelector(".row-points").value.trim(),
    tags: tr.querySelector(".row-tags").value.trim(),
    time_limit: tr.querySelector(".row-time").value.trim(),
    memory_limit: tr.querySelector(".row-memory").value.trim(),
    partial: tr.querySelector(".row-partial").checked,
    upload_statement: tr.querySelector(".row-statement").checked,
    upload_tests: tr.querySelector(".row-tests").checked,
    upload_solution: tr.querySelector(".row-solution").checked,
  }));
}

document.getElementById("confirmSingleUpload").onclick = async () => {
  const progressId = newProgressId();
  try {
    if (!preparedSingleUpload) throw new Error("HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u trÆ°á»›c khi xÃ¡c nháº­n up.");
    status("running");
    log("Äang up 1 bÃ i...");
    markRowsProcessing("#singleUploadTable", "Äang up...");
    startProgressPolling(progressId, "#singleUploadTable");
    const settings = singleUploadSettings();
    settings.progress_id = progressId;
    const data = await postJson("/api/confirm-single-upload", {prepare_id: preparedSingleUpload, settings, rows: collectSingleUploadRows(), progress_id: progressId});
    stopProgressPolling(progressId);
    applyStatuses(data.rows || [], "#singleUploadTable");
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("prepareTransfer").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang Ä‘á»c dá»¯ liá»‡u bÃ i nguá»“n...");
    const source = document.getElementById("transferSource").value;
    const dest = document.getElementById("transferDest").value;
    const codes = document.getElementById("transferCodes").value.split(/[\s,]+/).filter(Boolean);
    startProgressPolling(progressId, "#transferTable");
    const data = await postJson("/api/prepare-transfer", {
      source, dest, codes,
      source_account: accountPayload(source),
      settings: transferSettings(),
      progress_id: progressId,
    });
    stopProgressPolling(progressId);
    preparedTransfer = data.prepare_id;
    renderTransferTable(data.rows);
    document.getElementById("confirmTransfer").disabled = false;
    log(data.log);
    status("ready", "ok");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};
function renderTransferTable(rows) {
  document.getElementById("transferTable").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#transferTable', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#transferTable', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i toÃ¡n</th><th>Time</th><th>Memory</th><th>Up Ä‘á»</th><th>Up test</th><th>Bá»™ test</th><th>Sá»‘ test</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-original="${escapeHtml(row.original_code)}">
      <td><input type="checkbox" class="row-selected" checked></td>
      <td><input type="text" class="row-code" value="${escapeHtml(row.code)}"></td>
      <td><input type="text" class="row-name" value="${escapeHtml(row.name || "")}"></td>
      <td><input type="text" class="row-time" value="${escapeHtml(row.time_limit || "1.0")}"></td>
      <td><input type="text" class="row-memory" value="${escapeHtml(row.memory_limit || "1048576")}"></td>
      <td><input type="checkbox" class="row-statement" checked></td>
      <td><input type="checkbox" class="row-tests" checked></td>
      <td>${row.test_link ? `<a class="problem-link" href="${escapeHtml(row.test_link)}" target="_blank" rel="noopener">Bá»™ test</a>` : escapeHtml(row.test_file)}</td><td>${row.test_count}</td><td class="row-status">${escapeHtml(row.status)}</td>
    </tr>`).join("")}</tbody></table>`;
}
document.getElementById("confirmTransfer").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang chuyá»ƒn bÃ i...");
    const source = document.getElementById("transferSource").value;
    const dest = document.getElementById("transferDest").value;
    markRowsProcessing("#transferTable", "Äang chuyá»ƒn...");
    startProgressPolling(progressId, "#transferTable");
    const data = await postJson("/api/confirm-transfer", {
      prepare_id: preparedTransfer,
      source, dest, rows: collectRows("#transferTable"),
      settings: transferSettings(),
      source_account: accountPayload(source),
      dest_account: accountPayload(dest),
      progress_id: progressId,
    });
    stopProgressPolling(progressId);
    applyStatuses(data.rows, "#transferTable");
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("prepareContestTransfer").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang Ä‘á»c dá»¯ liá»‡u contest nguá»“n...");
    const source = document.getElementById("contestSource").value;
    const dest = document.getElementById("contestDest").value;
    const codes = document.getElementById("contestCodes").value.split(/[\s,]+/).filter(Boolean);
    startProgressPolling(progressId, "#contestTransferTable", "contest");
    const data = await postJson("/api/prepare-contest-transfer", {
      source, dest, codes,
      source_account: accountPayload(source),
      dest_account: accountPayload(dest),
      settings: contestTransferSettings(),
      progress_id: progressId,
    });
    stopProgressPolling(progressId);
    preparedContestTransfer = data.prepare_id;
    renderContestTransferTable(data.rows);
    document.getElementById("confirmContestTransfer").disabled = false;
    log(data.log);
    status("ready", "ok");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("confirmContestTransfer").onclick = async () => {
  const progressId = newProgressId();
  try {
    status("running");
    log("Äang chuyá»ƒn contest...");
    const source = document.getElementById("contestSource").value;
    const dest = document.getElementById("contestDest").value;
    markRowsProcessing("#contestTransferTable", "Äang chuyá»ƒn...");
    startProgressPolling(progressId, "#contestTransferTable", "contest");
    const data = await postJson("/api/confirm-contest-transfer", {
      prepare_id: preparedContestTransfer,
      source, dest, rows: collectContestRows(),
      source_account: accountPayload(source),
      dest_account: accountPayload(dest),
      settings: contestTransferSettings(),
      progress_id: progressId,
    });
    stopProgressPolling(progressId);
    applyContestStatuses(data.rows);
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("createContestButton").onclick = async () => {
  try {
    status("running");
    log("Äang táº¡o contest...");
    const target = document.getElementById("createContestTarget").value;
    const data = await postJson("/api/create-contest", {
      target,
      account: accountPayload(target),
      key: document.getElementById("createContestKey").value.trim(),
      name: document.getElementById("createContestName").value.trim(),
      start_time: backendDateTimeValue(document.getElementById("createContestStart").value.trim()),
      end_time: backendDateTimeValue(document.getElementById("createContestEnd").value.trim()),
      problems: document.getElementById("createContestProblems").value.split(/[\s,]+/).filter(Boolean),
    });
    log(data.log);
    status("done", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("prepareContestLessonCopy").onclick = async () => {
  try {
    status("running");
    log("Äang Ä‘á»c danh sÃ¡ch bÃ i trong contest vÃ  lesson Ä‘Ã­ch...");
    saveAccounts();
    const source = document.getElementById("lessonCopySource").value;
    const data = await postJson("/api/prepare-contest-to-lesson", {
      source,
      source_account: accountPayload(source),
      account: accountPayload("hncode"),
      contest_url: document.getElementById("lessonCopyContestUrl").value.trim(),
      lesson_url: document.getElementById("lessonCopyLessonUrl").value.trim(),
    });
    preparedContestLessonCopy = data.prepare_id;
    renderContestLessonCopyTable(data.rows || []);
    document.getElementById("confirmContestLessonCopy").disabled = !data.can_copy;
    log(data.log);
    status(data.can_copy ? "ready" : "done", data.can_copy ? "ok" : "warn");
  } catch (err) {
    preparedContestLessonCopy = null;
    document.getElementById("confirmContestLessonCopy").disabled = true;
    document.getElementById("contestLessonCopyTable").innerHTML = "";
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("confirmContestLessonCopy").onclick = async () => {
  try {
    if (!preparedContestLessonCopy) throw new Error("HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u trÆ°á»›c khi sao chÃ©p bÃ i.");
    status("running");
    log("Äang sao chÃ©p bÃ i vÃ o lesson HNCode...");
    markRowsProcessing("#contestLessonCopyTable", "Äang thÃªm...");
    saveAccounts();
    const source = document.getElementById("lessonCopySource").value;
    const res = await fetch("/api/confirm-contest-to-lesson", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({
      prepare_id: preparedContestLessonCopy,
      source_account: accountPayload(source),
      account: accountPayload("hncode"),
      rows: collectContestLessonCopyRows(),
    })});
    const data = await parseJsonResponse(res);
    applyContestLessonCopyStatuses(data.rows || []);
    if (!res.ok) throw new Error(data.error || "KhÃ´ng sao chÃ©p Ä‘Æ°á»£c bÃ i vÃ o lesson.");
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};
document.getElementById("fillLessonCopyScores").onclick = () => {
  const value = document.getElementById("lessonCopyBulkScore").value.trim();
  if (!value) {
    log("HÃ£y nháº­p Ä‘iá»ƒm chung trÆ°á»›c khi Ã¡p dá»¥ng.");
    status("failed", "err");
    return;
  }
  document.querySelectorAll("#contestLessonCopyTable .row-score").forEach(input => { input.value = value; });
  append(`ÄÃ£ Ã¡p dá»¥ng Ä‘iá»ƒm ${value} cho táº¥t cáº£ bÃ i trong báº£ng Contest â†’ Lesson.`);
};

function renderContestLessonCopyTable(rows) {
  document.getElementById("contestLessonCopyTable").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#contestLessonCopyTable', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#contestLessonCopyTable', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>STT</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i</th><th>Äiá»ƒm lesson</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-code="${escapeHtml(row.code)}">
      <td><input type="checkbox" class="row-selected" ${row.selected ? "checked" : ""} ${row.problem_id ? "" : "disabled"}></td>
      <td>${escapeHtml(row.index || "")}</td>
      <td><a class="problem-link" href="https://hncode.edu.vn/problem/${escapeHtml(row.code)}" target="_blank" rel="noopener">${escapeHtml(row.code)}</a></td>
      <td>${escapeHtml(row.title || "")}</td>
      <td><input type="text" class="row-score" value="${escapeHtml(row.score || "100")}"></td>
      <td class="row-status ${statusClass(row.status)}">${escapeHtml(row.status || "")}</td>
    </tr>`).join("")}</tbody></table>`;
}

function collectContestLessonCopyRows() {
  return [...document.querySelectorAll("#contestLessonCopyTable tbody tr")].map(tr => ({
    code: tr.dataset.code,
    selected: tr.querySelector(".row-selected").checked,
    score: tr.querySelector(".row-score").value.trim(),
  }));
}

function applyContestLessonCopyStatuses(rows) {
  const byCode = new Map(rows.map(row => [row.code, row]));
  for (const tr of document.querySelectorAll("#contestLessonCopyTable tbody tr")) {
    const row = byCode.get(tr.dataset.code);
    if (!row) continue;
    const detail = row.error ? "\n" + row.error : "";
    setStatusCell(tr.querySelector(".row-status"), (row.status || "") + detail, row.link || "");
  }
}

document.getElementById("prepareCourseClone").onclick = async () => {
  try {
    status("running");
    log("Äang Ä‘á»c lesson vÃ  contest cá»§a course nguá»“n...");
    saveAccounts();
    const data = await postJson("/api/prepare-course-clone", {
      account: accountPayload("hncode"),
      source_url: document.getElementById("courseCloneSourceUrl").value.trim(),
      dest_url: document.getElementById("courseCloneDestUrl").value.trim(),
      contest_suffix: document.getElementById("courseCloneContestSuffix").value.trim(),
      include_lessons: document.getElementById("courseCloneLessons").checked,
      include_contests: document.getElementById("courseCloneContests").checked,
    });
    preparedCourseClone = data.prepare_id;
    renderCourseCloneTable(data.rows || []);
    document.getElementById("confirmCourseClone").disabled = !data.can_clone;
    log(data.log);
    status(data.can_clone ? "ready" : "done", data.can_clone ? "ok" : "warn");
  } catch (err) {
    preparedCourseClone = null;
    document.getElementById("confirmCourseClone").disabled = true;
    document.getElementById("courseCloneTable").innerHTML = "";
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("confirmCourseClone").onclick = async () => {
  try {
    if (!preparedCourseClone) throw new Error("HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u trÆ°á»›c khi Clone Course.");
    status("running");
    log("Äang clone course HNCode...");
    markRowsProcessing("#courseCloneTable", "Äang clone...");
    saveAccounts();
    const res = await fetch("/api/confirm-course-clone", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({
      prepare_id: preparedCourseClone,
      account: accountPayload("hncode"),
      rows: collectCourseCloneRows(),
    })});
    const data = await parseJsonResponse(res);
    applyCourseCloneStatuses(data.rows || []);
    if (!res.ok) throw new Error(data.error || "KhÃ´ng clone Ä‘Æ°á»£c course.");
    log(data.log);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};

function renderCourseCloneTable(rows) {
  document.getElementById("courseCloneTable").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#courseCloneTable', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#courseCloneTable', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>Loáº¡i</th><th>Thá»© tá»±</th><th>MÃ£/ID nguá»“n</th><th>TÃªn</th><th>MÃ£ contest Ä‘Ã­ch</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-kind="${escapeHtml(row.kind)}" data-key="${escapeHtml(row.key)}">
      <td><input type="checkbox" class="row-selected" ${row.selected ? "checked" : ""} ${row.can_clone ? "" : "disabled"}></td>
      <td>${row.kind === "contest" ? "Contest" : "Lesson"}</td>
      <td>${escapeHtml(row.order || "")}</td>
      <td>${escapeHtml(row.key || "")}</td>
      <td>${escapeHtml(row.title || "")}</td>
      <td>${row.kind === "contest" ? `<input type="text" class="row-new-key" value="${escapeHtml(row.new_key || "")}">` : ""}</td>
      <td class="row-status ${statusClass(row.status)}">${escapeHtml(row.status || "")}</td>
    </tr>`).join("")}</tbody></table>`;
}

function collectCourseCloneRows() {
  return [...document.querySelectorAll("#courseCloneTable tbody tr")].map(tr => ({
    kind: tr.dataset.kind,
    key: tr.dataset.key,
    selected: tr.querySelector(".row-selected").checked,
    new_key: tr.querySelector(".row-new-key") ? tr.querySelector(".row-new-key").value.trim() : "",
  }));
}

function applyCourseCloneStatuses(rows) {
  const byId = new Map(rows.map(row => [row.kind + ":" + row.key, row]));
  for (const tr of document.querySelectorAll("#courseCloneTable tbody tr")) {
    const row = byId.get(tr.dataset.kind + ":" + tr.dataset.key);
    if (!row) continue;
    const detail = row.error ? "\n" + row.error : "";
    setStatusCell(tr.querySelector(".row-status"), (row.status || "") + detail, row.link || "");
  }
}

document.getElementById("chooseQuizFile").onclick = () => document.getElementById("quizFileInput").click();

let selectedQuizZipFile = null;

document.getElementById("quizFileInput").addEventListener("change", async event => {
  const file = event.target.files && event.target.files[0];
  document.getElementById("quizFileName").value = file ? file.name : "";
  if (file) {
    if (file.name.toLowerCase().endsWith(".zip")) {
      selectedQuizZipFile = file;
      document.getElementById("quizMarkdown").value = "[Sáº½ xá»­ lÃ½ tá»« file ZIP: " + file.name + "]";
    } else {
      selectedQuizZipFile = null;
      document.getElementById("quizMarkdown").value = await file.text();
    }
    preparedQuiz = null;
    document.getElementById("uploadQuizButton").disabled = true;
  }
});

document.getElementById("fillQuizSample").onclick = () => {
  document.getElementById("quizMarkdown").value = QUIZ_FORMAT_GUIDE;
  preparedQuiz = null;
  document.getElementById("uploadQuizButton").disabled = true;
};
document.getElementById("uploadQuizButton").disabled = true;
document.getElementById("quizMarkdown").addEventListener("input", () => {
  preparedQuiz = null;
  document.getElementById("uploadQuizButton").disabled = true;
});
document.getElementById("quizTarget").addEventListener("change", () => {
  updateQuizTargetUi();
  checkQuizLogin();
});
document.getElementById("prepareQuizButton").onclick = async () => {
  try {
    status("running");
    log("Äang kiá»ƒm tra dá»¯ liá»‡u quiz...");
    let data;
    if (selectedQuizZipFile) {
        const form = new FormData();
        form.append("zip_file", selectedQuizZipFile);
        const res = await fetch("/api/prepare-quiz-zip", {method:"POST", body:form});
        if (!res.ok) { const errData = await res.json().catch(()=>({})); throw new Error(errData.error || "Lá»—i gá»i API prepare-quiz-zip"); }
        data = await res.json();
    } else {
        data = await postJson("/api/prepare-quiz", {text: document.getElementById("quizMarkdown").value});
    }
    preparedQuiz = data.prepare_id;
    renderQuizTable(data.rows || []);
    log(data.log);
    document.getElementById("uploadQuizButton").disabled = !data.can_upload;
    status(data.can_upload ? "ready" : "failed", data.can_upload ? "ok" : "err");
  } catch (err) {
    preparedQuiz = null;
    document.getElementById("uploadQuizButton").disabled = true;
    document.getElementById("quizUploadSummary").innerHTML = "";
    log(String(err));
    status("failed", "err");
  }
};
document.getElementById("uploadQuizButton").onclick = async () => {
  try {
    if (!preparedQuiz) throw new Error("HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u trÆ°á»›c khi up quiz.");
    status("running");
    updateQuizTargetUi();
    const quizTarget = document.getElementById("quizTarget").value;
    const quizTargetInfo = QUIZ_TARGETS[quizTarget] || QUIZ_TARGETS.quiz_hncode;
    log("Äang up list quiz lÃªn " + (quizTargetInfo.label || "Quiz") + "...");
    saveAccounts();
    const data = await postJson("/api/upload-quiz", {
      prepare_id: preparedQuiz,
      target: quizTarget,
      account: quizAccountPayload(),
      shuffle_choices: document.getElementById("quizShuffleChoices").checked,
      is_public: document.getElementById("quizPublic").checked,
    });
    const rows = (data.rows || []).map(row => `${row.status} ${row.index}. ${row.title}${row.link ? " - " + row.link : ""}`).join("\n");
    applyQuizStatuses(data.rows || []);
    document.getElementById("quizUploadSummary").innerHTML = `<div class="note">${escapeHtml(rows || data.log || "").replaceAll("\n", "<br>")}</div>`;
    log(data.log || rows);
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};

function renderQuizTable(rows) {
  document.getElementById("quizUploadSummary").innerHTML = `<table>
    <thead><tr><th>STT</th><th>TiÃªu Ä‘á»</th><th>Loáº¡i</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-quiz-index="${row.index}">
      <td>${row.index}</td>
      <td>${escapeHtml(row.title || "")}</td>
      <td>${escapeHtml(row.type || "")}</td>
      <td class="row-status ${statusClass(row.status)}">${escapeHtml(row.status || "")}${row.error ? `<div class="test-meta">${escapeHtml(row.error)}</div>` : ""}</td>
    </tr>`).join("")}</tbody>
  </table>`;
}

function applyQuizStatuses(rows) {
  const byIndex = new Map(rows.map(row => [String(row.index), row]));
  for (const tr of document.querySelectorAll("#quizUploadSummary tr[data-quiz-index]")) {
    const row = byIndex.get(tr.dataset.quizIndex);
    if (!row) continue;
    const cell = tr.querySelector(".row-status");
    cell.className = "row-status " + statusClass(row.status);
    const linkHtml = row.link ? ` <a class="problem-link" href="${escapeHtml(row.link)}" target="_blank" rel="noopener">Link</a>` : "";
    const errorHtml = row.error ? `<div class="test-meta">${escapeHtml(row.error)}</div>` : "";
    cell.innerHTML = `${escapeHtml(row.status || "")}${linkHtml}${errorHtml}`;
  }
}

function syncCodeListType() {
  const site = document.getElementById("codeListSite").value;
  const type = document.getElementById("codeListType");
  if (site === "hnoj") {
    type.value = "contest";
    [...type.options].forEach(option => option.disabled = option.value === "lesson");
    document.getElementById("codeListUrl").value = document.getElementById("codeListUrl").value || "https://hnoj.edu.vn/contest/ctp_4";
  } else {
    [...type.options].forEach(option => option.disabled = false);
  }
}
document.getElementById("codeListSite").addEventListener("change", syncCodeListType);
document.getElementById("runCodeList").onclick = async () => {
  try {
    status("running");
    saveAccounts();
    syncCodeListType();
    const site = document.getElementById("codeListSite").value;
    const sourceType = document.getElementById("codeListType").value;
    log("Äang láº¥y danh sÃ¡ch mÃ£ bÃ i...");
    const data = await postJson("/api/misc/list-problem-codes", {
      site,
      source_type: sourceType,
      url: document.getElementById("codeListUrl").value.trim(),
      account: accountPayload(site),
    });
    document.getElementById("codeListOutput").value = data.codes_text || "";
    const rows = data.rows || [];
    document.getElementById("codeListSummary").innerHTML = `<div class="note">TÃ¬m tháº¥y ${rows.length} bÃ i.</div>
      <table>
        <thead><tr><th>STT</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i</th><th>Äiá»ƒm</th></tr></thead>
        <tbody>${rows.map(row => `<tr>
          <td>${row.index || row.order || ""}</td>
          <td><code>${escapeHtml(row.code || "")}</code></td>
          <td>${escapeHtml(row.title || "")}</td>
          <td>${escapeHtml(row.points || row.score || "")}</td>
        </tr>`).join("")}</tbody>
      </table>`;
    log(data.log || `ÄÃ£ láº¥y ${rows.length} mÃ£ bÃ i.`);
    status("done", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};
syncCodeListType();

document.getElementById("chooseLastSubZip").onclick = () => document.getElementById("lastSubZipFile").click();
document.getElementById("lastSubZipFile").addEventListener("change", event => {
  const file = event.target.files && event.target.files[0];
  document.getElementById("lastSubZipName").value = file ? file.name : "";
});
document.getElementById("runLastSubmissions").onclick = async () => {
  try {
    const input = document.getElementById("lastSubZipFile");
    const file = input.files && input.files[0];
    if (!file) throw new Error("HÃ£y chá»n file zip data trÆ°á»›c.");
    status("running");
    log("Äang xá»­ lÃ½ last submissions...");
    const form = new FormData();
    form.append("zip_file", file);
    const res = await fetch("/api/misc/last-submissions", {method:"POST", body:form});
    if (!res.ok) {
      const data = await parseJsonResponse(res);
      throw new Error(data.error || "KhÃ´ng xá»­ lÃ½ Ä‘Æ°á»£c file zip.");
    }
    const summaryRaw = res.headers.get("X-Last-Submissions-Summary") || "";
    const summary = summaryRaw ? JSON.parse(decodeURIComponent(summaryRaw)) : {};
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = summary.filename || "last_submissions.zip";
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    const text = `âœ“ ÄÃ£ táº¡o file zip last submissions.\nTÃ¬m tháº¥y: ${summary.found || 0}/${summary.total || 0} thÃ­ sinh\nThiáº¿u file: ${summary.missing || 0}\nFile táº£i vá»: ${summary.filename || "last_submissions.zip"}`;
    document.getElementById("lastSubmissionsSummary").innerHTML = `<div class="note">${escapeHtml(text).replaceAll("\n", "<br>")}</div>`;
    log(text);
    status("done", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("chooseGradingZip").onclick = () => document.getElementById("gradingZipFile").click();
document.getElementById("chooseGradingCsv").onclick = () => document.getElementById("gradingCsvFile").click();
document.getElementById("gradingZipFile").addEventListener("change", event => {
  selectedGradingZipFile = event.target.files && event.target.files[0] || null;
  document.getElementById("gradingZipName").value = selectedGradingZipFile ? selectedGradingZipFile.name : "";
});
document.getElementById("gradingCsvFile").addEventListener("change", event => {
  selectedGradingCsvFile = event.target.files && event.target.files[0] || null;
  document.getElementById("gradingCsvName").value = selectedGradingCsvFile ? selectedGradingCsvFile.name : "";
});
function renderGradingTable(rows) {
  document.getElementById("gradingSummary").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#gradingSummary', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#gradingSummary', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>Há»c sinh</th><th>Username</th><th>MÃ£ bÃ i</th><th>TÃªn bÃ i</th><th>Äiá»ƒm bÃ i</th><th>File</th><th>%</th><th>Äiá»ƒm</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-original="${escapeHtml(row.original_key)}">
      <td><input type="checkbox" class="row-selected" ${row.selected ? "checked" : ""}></td>
      <td>${escapeHtml(row.student || "")}</td>
      <td>${escapeHtml(row.username || "")}</td>
      <td>${escapeHtml(row.problem || "")}</td>
      <td>${escapeHtml(row.problem_title || "")}</td>
      <td>${escapeHtml(row.contest_points || "")}</td>
      <td><div class="test-meta">${escapeHtml(row.relative_path || row.file || "")}</div></td>
      <td class="row-percent">${escapeHtml(row.percent || "")}</td>
      <td class="row-score">${escapeHtml(row.score || "")}</td>
      <td class="row-status ${statusClass(row.status)}">${escapeHtml(row.status || "")}${row.submission_url ? ` <a class="problem-link" href="${escapeHtml(row.submission_url)}" target="_blank" rel="noopener">Link</a>` : ""}${row.message ? `<div class="test-meta">${escapeHtml(row.message)}</div>` : ""}</td>
    </tr>`).join("")}</tbody></table>`;
}
function collectGradingRows() {
  return [...document.querySelectorAll("#gradingSummary tbody tr")].map(tr => ({
    original_key: tr.dataset.original,
    selected: tr.querySelector(".row-selected").checked,
  }));
}
function applyGradingStatuses(rows) {
  const byKey = new Map(rows.map(row => [row.original_key, row]));
  for (const tr of document.querySelectorAll("#gradingSummary tbody tr")) {
    const row = byKey.get(tr.dataset.original);
    if (!row) continue;
    tr.querySelector(".row-percent").textContent = row.percent || "";
    tr.querySelector(".row-score").textContent = row.score || "";
    const cell = tr.querySelector(".row-status");
    cell.className = "row-status " + statusClass(row.status);
    const linkHtml = row.submission_url ? ` <a class="problem-link" href="${escapeHtml(row.submission_url)}" target="_blank" rel="noopener">Link</a>` : "";
    const msgHtml = row.message ? `<div class="test-meta">${escapeHtml(row.message)}</div>` : "";
    cell.innerHTML = `${escapeHtml(row.status || "")}${linkHtml}${msgHtml}`;
  }
}
document.getElementById("prepareGrading").onclick = async () => {
  const progressId = newProgressId();
  try {
    if (!selectedGradingZipFile) throw new Error("HÃ£y chá»n file zip bÃ i lÃ m.");
    if (!selectedGradingCsvFile) throw new Error("HÃ£y chá»n file CSV tÃ i khoáº£n.");
    status("running");
    log("Äang chuáº©n bá»‹ dá»¯ liá»‡u cháº¥m HNCode...");
    document.getElementById("downloadGradingResult").classList.add("hidden");
    startProgressPolling(progressId, "#gradingSummary", "grading");
    const form = new FormData();
    form.append("zip_file", selectedGradingZipFile);
    form.append("csv_file", selectedGradingCsvFile);
    form.append("contest_url", document.getElementById("gradingContestUrl").value.trim());
    form.append("progress_id", progressId);
    form.append("admin_username", accountFields.hncode_user.value);
    form.append("admin_password", accountFields.hncode_pass.value);
    const res = await fetch("/api/prepare-hncode-grading", {method:"POST", body:form});
    const data = await parseJsonResponse(res);
    if (!res.ok) throw new Error(data.error || "KhÃ´ng chuáº©n bá»‹ Ä‘Æ°á»£c dá»¯ liá»‡u cháº¥m.");
    stopProgressPolling(progressId);
    preparedGrading = data.prepare_id;
    renderGradingTable(data.rows || []);
    document.getElementById("confirmGrading").disabled = false;
    log(data.log || "ÄÃ£ chuáº©n bá»‹ dá»¯ liá»‡u cháº¥m.");
    status("ready", "ok");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};
document.getElementById("confirmGrading").onclick = async () => {
  const progressId = newProgressId();
  try {
    if (!preparedGrading) throw new Error("HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u trÆ°á»›c.");
    status("running");
    log("Äang Ä‘Äƒng nháº­p há»c sinh, tham gia contest vÃ  ná»™p bÃ i...");
    markRowsProcessing("#gradingSummary", "Äang cháº¥m...");
    startProgressPolling(progressId, "#gradingSummary", "grading");
    const data = await postJson("/api/confirm-hncode-grading", {
      prepare_id: preparedGrading,
      rows: collectGradingRows(),
      contest_password: document.getElementById("gradingContestPassword").value,
      admin_account: accountPayload("hncode"),
      progress_id: progressId,
    });
    stopProgressPolling(progressId);
    applyGradingStatuses(data.rows || []);
    const link = data.download_url ? `\nTáº£i báº£ng Ä‘iá»ƒm: ${location.origin}${data.download_url}` : "";
    log((data.log || "ÄÃ£ cháº¥m xong.") + link);
    if (data.download_url) {
      const a = document.getElementById("downloadGradingResult");
      a.href = data.download_url;
      a.classList.remove("hidden");
    }
    status(data.ok ? "done" : "failed", data.ok ? "ok" : "err");
  } catch (err) {
    stopProgressPolling(progressId);
    log(String(err));
    status("failed", "err");
  }
};

document.getElementById("chooseAiWarningZip").onclick = () => document.getElementById("aiWarningZipFile").click();
document.getElementById("aiWarningZipFile").addEventListener("change", event => {
  const file = event.target.files && event.target.files[0];
  document.getElementById("aiWarningZipName").value = file ? file.name : "";
});
document.getElementById("runAiWarning").onclick = async () => {
  try {
    status("running");
    log("Äang phÃ¢n tÃ­ch dáº¥u hiá»‡u sá»­ dá»¥ng AI Ä‘á»ƒ code...");
    const input = document.getElementById("aiWarningZipFile");
    const file = input.files && input.files[0];
    const folder = document.getElementById("aiWarningFolder").value.trim();
    const form = new FormData();
    if (file) form.append("zip_file", file);
    else form.append("folder_path", folder);
    const res = await fetch("/api/misc/ai-code-warning", {method:"POST", body:form});
    if (!res.ok) {
      const data = await parseJsonResponse(res);
      throw new Error(data.error || "KhÃ´ng táº¡o Ä‘Æ°á»£c bÃ¡o cÃ¡o.");
    }
    const summaryRaw = res.headers.get("X-AI-Warning-Summary") || "";
    const summary = summaryRaw ? JSON.parse(decodeURIComponent(summaryRaw)) : {};
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = summary.filename || "ai_code_warning_report.xlsx";
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    const text = `âœ“ ÄÃ£ táº¡o bÃ¡o cÃ¡o Excel cáº£nh bÃ¡o AI code.\nSá»‘ contest zip: ${summary.zip_count || 0}\nSá»‘ file code: ${summary.code_file_count || 0}\nSá»‘ thÃ­ sinh: ${summary.student_count || 0}\nKháº£ nÄƒng cao: ${summary.high || 0}\nKháº£ nÄƒng trung bÃ¬nh: ${summary.medium || 0}\nKháº£ nÄƒng tháº¥p: ${summary.low || 0}\nÄá»•i style cÃ¹ng bÃ i: ${summary.shift_count || 0}\nCáº·p nghi chÃ©p code: ${summary.copy_pair_count || 0}\nCáº·p ráº¥t giá»‘ng: ${summary.copy_very_similar || 0}\nChi tiáº¿t cáº·p theo bÃ i: ${summary.copy_detail_count || 0}\nThÆ° má»¥c code Ä‘Ã£ giáº£i nÃ©n: ${summary.extracted_folder || ""}\nFile táº£i vá»: ${summary.filename || "ai_code_warning_report.xlsx"}`;
    document.getElementById("aiWarningSummary").innerHTML = `<div class="note">${escapeHtml(text).replaceAll("\n", "<br>")}</div>`;
    log(text);
    status("done", "ok");
  } catch (err) {
    log(String(err));
    status("failed", "err");
  }
};

function contestTransferSettings() {
  return {
    reuse_existing_problems: document.getElementById("contestReuseExistingProblems").checked,
    create_missing_problems: document.getElementById("contestCreateMissingProblems").checked,
    time_limit: document.getElementById("contestProblemTime").value,
    memory_limit: document.getElementById("contestProblemMemory").value,
  };
}

function renderContestTransferTable(rows) {
  document.getElementById("contestTransferTable").innerHTML = `<div class="table-tools">
    <button class="action" type="button" onclick="setRowSelection('#contestTransferTable', true)">Chá»n táº¥t cáº£</button>
    <button class="action" type="button" onclick="setRowSelection('#contestTransferTable', false)">Bá» chá»n táº¥t cáº£</button>
  </div><table>
    <thead><tr><th>Chá»n</th><th>MÃ£ contest</th><th>TÃªn contest</th><th>Thá»i gian</th><th>BÃ i trong contest</th><th>Tráº¡ng thÃ¡i</th></tr></thead>
    <tbody>${rows.map(row => `<tr data-original="${escapeHtml(row.original_key)}">
      <td><input type="checkbox" class="row-selected" ${row.can_transfer ? "checked" : ""}></td>
      <td><input type="text" class="row-key" value="${escapeHtml(row.key)}"></td>
      <td><input type="text" class="row-name" value="${escapeHtml(row.name || "")}"></td>
      <td><div class="test-meta">${escapeHtml(row.start_time || "")}<br>${escapeHtml(row.end_time || "")}</div></td>
      <td>${renderContestProblemList(row.problems || [])}</td>
      <td class="row-status">${escapeHtml(row.status)}</td>
    </tr>`).join("")}</tbody></table>`;
}

function renderContestProblemList(problems) {
  if (!problems.length) return `<div class="test-meta">KhÃ´ng cÃ³ bÃ i.</div>`;
  return `<table class="inner-table"><thead><tr><th>Chá»n</th><th>MÃ£ bÃ i</th><th>Äiá»ƒm</th><th>Thá»© tá»±</th><th>Tráº¡ng thÃ¡i</th></tr></thead><tbody>
    ${problems.map(p => `<tr data-problem-code="${escapeHtml(p.code)}">
      <td><input type="checkbox" class="problem-selected" checked></td>
      <td>${escapeHtml(p.code)}</td>
      <td>${escapeHtml(p.points || "100")}</td>
      <td>${escapeHtml(p.order || "")}</td>
      <td>${escapeHtml(p.status || "")}</td>
    </tr>`).join("")}
  </tbody></table>`;
}

function collectContestRows() {
  return [...document.querySelectorAll("#contestTransferTable > table > tbody > tr")].map(tr => ({
    original_key: tr.dataset.original,
    selected: tr.querySelector(".row-selected").checked,
    key: tr.querySelector(".row-key").value.trim(),
    name: tr.querySelector(".row-name").value.trim(),
    problems: [...tr.querySelectorAll(".inner-table tbody tr")].map(pr => ({
      code: pr.dataset.problemCode,
      selected: pr.querySelector(".problem-selected").checked,
    })),
  }));
}

function applyContestStatuses(rows) {
  const byOriginal = new Map(rows.map(row => [row.original_key, row]));
  for (const tr of document.querySelectorAll("#contestTransferTable > table > tbody > tr")) {
    const row = byOriginal.get(tr.dataset.original);
    if (!row) continue;
    setStatusCell(tr.querySelector(".row-status"), row.status, row.link || "");
  }
}

function collectRows(selector) {
  return [...document.querySelectorAll(selector + " tbody tr")].map(tr => ({
    original_code: tr.dataset.original,
    selected: tr.querySelector(".row-selected").checked,
    code: tr.querySelector(".row-code").value.trim(),
    name: tr.querySelector(".row-name").value.trim(),
    time_limit: tr.querySelector(".row-time") ? tr.querySelector(".row-time").value.trim() : "",
    memory_limit: tr.querySelector(".row-memory") ? tr.querySelector(".row-memory").value.trim() : "",
    upload_statement: tr.querySelector(".row-statement").checked,
    upload_tests: tr.querySelector(".row-tests").checked,
  }));
}
function setRowSelection(selector, checked) {
  document.querySelectorAll(selector + " .row-selected").forEach(item => { item.checked = checked; });
}
function markRowsProcessing(selector, text="Äang xá»­ lÃ½...") {
  for (const tr of document.querySelectorAll(selector + " tbody tr")) {
    const selected = tr.querySelector(".row-selected");
    const statusCell = tr.querySelector(".row-status");
    if (selected && selected.checked && statusCell) {
      statusCell.className = "row-status";
      statusCell.textContent = text;
    }
  }
}
function applyStatuses(rows, selector) {
  const byOriginal = new Map(rows.map(row => [row.original_code, row]));
  for (const tr of document.querySelectorAll(selector + " tbody tr")) {
    const row = byOriginal.get(tr.dataset.original);
    if (!row) continue;
    setStatusCell(tr.querySelector(".row-status"), row.status, row.link || "");
  }
}
function escapeHtml(text) {
  return String(text).replace(/[&<>"']/g, ch => ({"&":"&amp;","<":"&lt;",">":"&gt;","\"":"&quot;","'":"&#039;"}[ch]));
}
</script>
</body>
</html>
"""


@app.get("/")
def index():
    return render_template_string(
        PAGE,
        default_zip=DEFAULT_ZIP,
        ai_source_default=AI_SOURCE_DEFAULT,
        prompt_guide=PROMPT_GUIDE,
        quiz_format_guide_json=json.dumps(QUIZ_FORMAT_GUIDE, ensure_ascii=False),
        quiz_targets_json=json.dumps(QUIZ_TARGETS, ensure_ascii=False),
        targets_json=json.dumps(TARGETS, ensure_ascii=False),
    )


@app.get("/samples/bo_mau_1_bai_tonghaiso.zip")
def sample_tonghaiso_zip():
    if not SAMPLE_TONGHAISO_ZIP.exists():
        return jsonify({"error": "KhÃ´ng tÃ¬m tháº¥y file máº«u."}), 404
    return send_file(SAMPLE_TONGHAISO_ZIP, as_attachment=True, download_name=SAMPLE_TONGHAISO_ZIP.name)


@app.post("/api/sample/tonghaiso")
def api_sample_tonghaiso():
    try:
        if not SAMPLE_TONGHAISO_ZIP.exists():
            raise FileNotFoundError(f"KhÃ´ng tÃ¬m tháº¥y file máº«u: {SAMPLE_TONGHAISO_ZIP}")
        with zipfile.ZipFile(SAMPLE_TONGHAISO_ZIP) as archive:
            statement = read_zip_member_text(archive, "tonghaiso.md")
            generator = read_zip_member_text(archive, "gentest_tonghaiso.py")
            solution_md = read_zip_member_text(archive, "sol_tonghaiso.md")
        parts = first_markdown_header_parts(statement)
        return jsonify(
            {
                "zip_path": str(SAMPLE_TONGHAISO_ZIP),
                "code": parts[1] if len(parts) > 1 else "tonghaiso",
                "name": parts[0] if parts else "Tá»•ng hai sá»‘",
                "points": parts[2] if len(parts) > 2 else "800",
                "tags": parts[3] if len(parts) > 3 else "implementation, math",
                "statement": statement,
                "generator": generator,
                "solution_md": solution_md,
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


def read_zip_member_text(archive: zipfile.ZipFile, name: str) -> str:
    raw = archive.read(name)
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def first_markdown_header_parts(text: str) -> list[str]:
    for line in text.splitlines():
        stripped = line.strip().strip("#* ")
        if stripped:
            return [part.strip() for part in stripped.split("|")]
    return []


@app.post("/api/check-login")
def api_check_login():
    payload = request.get_json(force=True)
    target = payload.get("target", "")
    account = payload.get("account", {})
    probe_code = (payload.get("probe_code") or "").strip()
    try:
        if target == "tinhoctre":
            login_hncode(TARGETS[target]["base_url"], account.get("username", ""), account.get("password", ""))
            return jsonify({"ok": True, "message": "ÄÄƒng nháº­p OK"})
        if target in QUIZ_TARGETS:
            login_quiz_target(target, account)
            return jsonify({"ok": True, "message": "ÄÄƒng nháº­p OK"})
        if target == "contest_hnoj":
            info = CONTEST_TARGETS[target]
        else:
            info = TARGETS[target]
        login_hncode(info["base_url"], account.get("username", ""), account.get("password", ""))
        return jsonify({"ok": True, "message": "ÄÄƒng nháº­p OK"})
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)[:180]})


@app.post("/api/misc/list-problem-codes")
def api_misc_list_problem_codes():
    payload = request.get_json(force=True)
    site = payload.get("site", "hncode")
    source_type = payload.get("source_type", "contest")
    source_url = (payload.get("url") or "").strip()
    account = payload.get("account", {})
    try:
        if site == "hncode":
            session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
            if source_type == "lesson":
                course_slug, lesson_id = extract_hncode_lesson_ref(source_url)
                rows = hncode_lesson_problem_code_rows(session, course_slug, lesson_id)
                source_label = f"HNCode Lesson: {course_slug}/lesson/{lesson_id}"
            else:
                contest_key = extract_hncode_contest_key(source_url)
                rows = hncode_contest_problem_rows(session, contest_key)
                source_label = f"HNCode Contest: {contest_key}"
        elif site == "hnoj":
            if source_type != "contest":
                raise RuntimeError("HNOJ hiá»‡n chá»‰ há»— trá»£ láº¥y mÃ£ bÃ i tá»« Contest.")
            session = login_hncode(TARGETS["hnoj"]["base_url"], account.get("username", ""), account.get("password", ""))
            contest_key = extract_hncode_contest_key(source_url)
            rows = hnoj_contest_problem_rows(session, contest_key)
            source_label = f"HNOJ Contest: {contest_key}"
        else:
            raise RuntimeError("Nguá»“n khÃ´ng há»£p lá»‡. HÃ£y chá»n HNCode hoáº·c HNOJ.")
        for index, row in enumerate(rows, 1):
            row["index"] = index
        codes_text = "\n".join(row["code"] for row in rows)
        compact_text = " ".join(row["code"] for row in rows)
        log_lines = [
            f"Nguá»“n: {source_label}",
            f"Sá»‘ bÃ i: {len(rows)}",
            "Danh sÃ¡ch mÃ£ bÃ i:",
            codes_text,
        ]
        return jsonify({"ok": True, "rows": rows, "codes_text": codes_text, "compact_text": compact_text, "log": "\n".join(log_lines)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400



@app.post("/api/prepare-quiz")
def api_prepare_quiz():
    payload = request.get_json(force=True)
    try:
        questions, rows = prepare_quiz_items(payload.get("text", ""))
        prepare_id = uuid.uuid4().hex
        prepared_quizzes[prepare_id] = {"questions": questions, "rows": rows, "created_at": time.time(), "zip_dir": None}
        ok_count = sum(1 for row in rows if row.get("can_upload"))
        bad_count = len(rows) - ok_count
        log_lines = [f"Chuáº©n bá»‹ dá»¯ liá»‡u quiz: {ok_count}/{len(rows)} cÃ¢u há»£p lá»‡."]
        for row in rows:
            if row.get("can_upload"):
                log_lines.append(f"âœ“ CÃ¢u {row['index']}: {row['title']} ({row['type']}) há»£p lá»‡.")
            else:
                log_lines.append(f"âœ— CÃ¢u {row['index']}: {row.get('error')}")
        return jsonify({"ok": bad_count == 0, "can_upload": bad_count == 0 and ok_count > 0, "prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

@app.post("/api/prepare-quiz-zip")
def api_prepare_quiz_zip():
    try:
        uploaded = request.files.get("zip_file")
        if not uploaded:
            raise ValueError("KhÃ´ng tÃ¬m tháº¥y file zip_file")

        prepare_id = uuid.uuid4().hex
        root = RUNTIME / prepare_id
        source_dir = root / "source"
        source_dir.mkdir(parents=True, exist_ok=True)
        zip_path = root / "uploaded.zip"
        uploaded.save(zip_path)

        extract_zip(zip_path, source_dir)

        # find markdown file
        md_files = list(source_dir.rglob("*.md")) + list(source_dir.rglob("*.txt"))
        if not md_files:
            raise ValueError("KhÃ´ng tÃ¬m tháº¥y file .md hoáº·c .txt trong thÆ° má»¥c ZIP.")

        md_content = md_files[0].read_text(encoding="utf-8")
        questions, rows = prepare_quiz_items(md_content)

        prepared_quizzes[prepare_id] = {"questions": questions, "rows": rows, "created_at": time.time(), "zip_dir": source_dir}
        ok_count = sum(1 for row in rows if row.get("can_upload"))
        bad_count = len(rows) - ok_count
        log_lines = [f"ÄÃ£ Ä‘á»c file ZIP, chuáº©n bá»‹ {ok_count}/{len(rows)} cÃ¢u há»£p lá»‡."]
        return jsonify({"ok": bad_count == 0, "can_upload": bad_count == 0 and ok_count > 0, "prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

@app.post("/api/upload-quiz")
def api_upload_quiz():
    payload = request.get_json(force=True)
    prepare_id = payload.get("prepare_id")
    target = payload.get("target")
    account = payload.get("account")
    if prepare_id not in prepared_quizzes:
        return jsonify({"ok": False, "error": "Prepare ID khÃ´ng há»£p lá»‡ hoáº·c Ä‘Ã£ háº¿t háº¡n"}), 400

    quiz_data = prepared_quizzes[prepare_id]
    questions = quiz_data["questions"]
    rows = quiz_data["rows"]
    zip_dir = quiz_data.get("zip_dir")

    try:
        session = login_quiz_target(target, account)
        info = quiz_target_info(target)
        base_url = info["base_url"]

        log_lines = []
        for i, q in enumerate(questions):
            row = rows[i]
            if not row.get("can_upload"):
                continue
            try:
                # Thay tháº¿ áº£nh náº¿u cÃ³ zip_dir
                content = q.get("content", "")
                if zip_dir and ("[áº¢nh:" in content or "![áº¢nh]" in content or "<img" in content):
                    def repl_img(match):
                        filename = match.group(1).strip()
                        img_path = find_image_in_dir(filename, zip_dir)
                        if img_path:
                            link = upload_quiz_image(session, base_url, img_path)
                            if link:
                                return f"![áº¢nh]({link})"
                        return match.group(0)
                    content = re.sub(r'\[áº¢nh:\s*(.*?)\]', repl_img, content)
                    content = re.sub(r'!\[áº¢nh\]\((.*?)\)', lambda m: repl_img(re.match(r'(.*)', m.group(1))), content)
                    q["content"] = content

                url = create_quiz_question(session, base_url, q, shuffle_choices=payload.get("shuffle_choices", True), is_public=payload.get("is_public", False))
                row["status"] = "ThÃ nh cÃ´ng"
                row["link"] = url
                log_lines.append(f"CÃ¢u {i+1}: OK")
            except Exception as e:
                row["status"] = f"Lá»—i: {e}"
                log_lines.append(f"CÃ¢u {i+1}: {e}")

        has_errors = any("Lá»—i" in r["status"] for r in rows if r.get("can_upload"))
        return jsonify({"ok": not has_errors, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/prepare-course-clone")
def api_prepare_course_clone():
    payload = request.get_json(force=True)
    account = payload.get("account", {})
    try:
        source_slug = extract_hncode_course_slug(payload.get("source_url", ""))
        dest_slug = extract_hncode_course_slug(payload.get("dest_url", ""))
        if source_slug == dest_slug:
            raise RuntimeError("Course nguá»“n vÃ  course Ä‘Ã­ch Ä‘ang trÃ¹ng nhau.")
        include_lessons = bool(payload.get("include_lessons", True))
        include_contests = bool(payload.get("include_contests", True))
        if not include_lessons and not include_contests:
            raise RuntimeError("HÃ£y chá»n Clone lesson hoáº·c Clone contest.")
        session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
        dest_course_id = hncode_course_admin_id(session, dest_slug)
        source_lessons = hncode_course_lessons(session, source_slug) if include_lessons else []
        source_contests = hncode_course_contests(session, source_slug) if include_contests else []
        dest_lessons = hncode_course_lessons(session, dest_slug)
        dest_contests = hncode_course_contests(session, dest_slug)
        dest_lesson_titles = {row["title"].strip().casefold() for row in dest_lessons}
        dest_contest_keys = {row["key"] for row in dest_contests}
        rows: list[dict] = []
        log_lines = [
            "Chuáº©n bá»‹ Clone Course HNCode",
            f"Nguá»“n: {source_slug}",
            f"ÄÃ­ch: {dest_slug}",
            f"Lesson nguá»“n: {len(source_lessons)}",
            f"Contest nguá»“n: {len(source_contests)}",
        ]
        for item in source_lessons:
            exists = item["title"].strip().casefold() in dest_lesson_titles
            row = {
                **item,
                "selected": not exists,
                "can_clone": not exists,
                "status": "ÄÃ£ cÃ³ lesson cÃ¹ng tÃªn á»Ÿ Ä‘Ã­ch" if exists else "âœ“ Sáºµn sÃ ng",
                "new_key": "",
            }
            rows.append(row)
            log_lines.append(f"Lesson {item['order']}. {item['title']}: {row['status']}")
        suffix = payload.get("contest_suffix", "")
        for item in source_contests:
            new_key = default_course_clone_contest_key(item["key"], dest_slug, suffix)
            in_dest = new_key in dest_contest_keys
            global_exists = False
            if not in_dest:
                try:
                    global_exists = bool(admin_contest_change_url(session, TARGETS["hncode"]["base_url"], new_key))
                except Exception:
                    global_exists = False
            if in_dest:
                status_text = "ÄÃ£ cÃ³ contest Ä‘Ã­ch trong course"
            elif global_exists:
                status_text = "MÃ£ contest Ä‘Ã­ch Ä‘Ã£ tá»“n táº¡i trÃªn HNCode"
            else:
                status_text = "âœ“ Sáºµn sÃ ng"
            row = {
                **item,
                "selected": status_text.startswith("âœ“"),
                "can_clone": status_text.startswith("âœ“"),
                "status": status_text,
                "new_key": new_key,
            }
            rows.append(row)
            log_lines.append(f"Contest {item['key']} â†’ {new_key}: {status_text}")
        prepare_id = uuid.uuid4().hex
        prepared_course_clones[prepare_id] = {
            "created_at": time.time(),
            "source_slug": source_slug,
            "dest_slug": dest_slug,
            "dest_course_id": dest_course_id,
            "rows": rows,
        }
        return jsonify(
            {
                "ok": True,
                "prepare_id": prepare_id,
                "rows": rows,
                "can_clone": any(row.get("selected") for row in rows),
                "log": "\n".join(log_lines),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/confirm-course-clone")
def api_confirm_course_clone():
    payload = request.get_json(force=True)
    account = payload.get("account", {})
    prepare_id = payload.get("prepare_id", "")
    state = prepared_course_clones.get(prepare_id)
    if not state:
        return jsonify({"ok": False, "error": "Dá»¯ liá»‡u chuáº©n bá»‹ Clone Course Ä‘Ã£ háº¿t háº¡n. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i."}), 400
    rows_by_id = {(row["kind"], row["key"]): row for row in state["rows"]}
    requested_rows = payload.get("rows", [])
    result_rows = []
    ok = True
    log_lines = [
        "Clone Course HNCode",
        f"Nguá»“n: {state['source_slug']}",
        f"ÄÃ­ch: {state['dest_slug']}",
    ]
    try:
        session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
        for requested in requested_rows:
            key = requested.get("key", "")
            kind = requested.get("kind", "")
            base = dict(rows_by_id.get((kind, key), requested))
            base["selected"] = bool(requested.get("selected"))
            if kind == "contest":
                base["new_key"] = (requested.get("new_key") or base.get("new_key") or "").strip()
            if not base["selected"]:
                base["status"] = "Bá» qua"
                result_rows.append(base)
                log_lines.append(f"- {kind} {key}: bá» qua.")
                continue
            try:
                if kind == "lesson":
                    link = clone_hncode_lesson_native(session, state["source_slug"], key, base.get("title") or f"Lesson {key}", state["dest_slug"], state["dest_course_id"])
                    base["status"] = "âœ“ ÄÃ£ clone"
                    base["link"] = link
                    log_lines.append(f"âœ“ Lesson {key}: Ä‘Ã£ clone.")
                elif kind == "contest":
                    new_key = base.get("new_key", "")
                    if not re.fullmatch(r"[a-z0-9_-]+", new_key):
                        raise RuntimeError("MÃ£ contest Ä‘Ã­ch chá»‰ nÃªn gá»“m chá»¯ thÆ°á»ng, sá»‘, dáº¥u gáº¡ch dÆ°á»›i hoáº·c gáº¡ch ngang.")
                    link = clone_hncode_contest_native(session, key, new_key, state["dest_slug"], state["dest_course_id"])
                    base["status"] = "âœ“ ÄÃ£ clone"
                    base["link"] = link
                    log_lines.append(f"âœ“ Contest {key} â†’ {new_key}: Ä‘Ã£ clone.")
                else:
                    raise RuntimeError(f"Loáº¡i dÃ²ng khÃ´ng há»£p lá»‡: {kind}")
            except Exception as item_exc:
                ok = False
                base["status"] = "âœ— Lá»—i"
                base["error"] = str(item_exc)
                log_lines.append(f"âœ— {kind} {key}: {item_exc}")
            result_rows.append(base)
        if not result_rows:
            ok = False
            log_lines.append("KhÃ´ng cÃ³ dÃ²ng nÃ o Ä‘Æ°á»£c gá»­i lÃªn Ä‘á»ƒ clone.")
        return jsonify({"ok": ok, "rows": result_rows, "log": "\n".join(log_lines), "course_link": hncode_course_page_url(state["dest_slug"])})
    except Exception as exc:
        return jsonify({"ok": False, "rows": result_rows, "error": str(exc)}), 400


@app.post("/api/prepare-contest-to-lesson")
def api_prepare_contest_to_lesson():
    payload = request.get_json(force=True)
    contest_url_value = payload.get("contest_url", "")
    source = contest_lesson_source_from_url(payload.get("source", "hncode"), contest_url_value)
    source_account = payload.get("source_account", {})
    account = payload.get("account", {})
    try:
        contest_key = extract_hncode_contest_key(contest_url_value)
        course_slug, lesson_id = extract_hncode_lesson_ref(payload.get("lesson_url", ""))
        dst_session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
        if source == "hnoj":
            src_session = login_hncode(TARGETS["hnoj"]["base_url"], source_account.get("username", ""), source_account.get("password", ""))
            contest_rows = hnoj_contest_problem_rows(src_session, contest_key)
            source_label = "HNOJ"
        else:
            contest_rows = hncode_contest_problem_rows(dst_session, contest_key)
            source_label = "HNCode"
        lesson_page = dst_session.get(hncode_lesson_edit_url(course_slug, lesson_id), timeout=30)
        if not lesson_page.ok:
            raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c lesson Ä‘Ã­ch: HTTP {lesson_page.status_code}")
        existing_ids = {row["problem"] for row in lesson_problem_rows_from_page(lesson_page.text, lesson_id)}
        rows = []
        log_lines = [
            f"Chuáº©n bá»‹ sao chÃ©p bÃ i {source_label} Contest â†’ Lesson HNCode",
            f"Contest: {contest_key}",
            f"Lesson: {hncode_lesson_url(course_slug, lesson_id)}",
        ]
        for item in contest_rows:
            source_code = item["code"]
            dest_code = normalize_problem_code_for_target(source_code, "hncode")
            problem_id = admin_problem_id(dst_session, TARGETS["hncode"]["base_url"], dest_code)
            if not problem_id:
                status_text = "Thiáº¿u trÃªn HNCode, sáº½ chuyá»ƒn khi xÃ¡c nháº­n" if source == "hnoj" else "âœ— KhÃ´ng tÃ¬m tháº¥y bÃ i trong admin HNCode"
                selected = source == "hnoj"
            elif problem_id in existing_ids:
                status_text = "ÄÃ£ cÃ³ trong lesson"
                selected = False
            else:
                status_text = "âœ“ Sáºµn sÃ ng"
                selected = True
            row = {
                "index": item["order"],
                "source_code": source_code,
                "code": dest_code,
                "title": item["title"],
                "score": item["points"],
                "problem_id": problem_id or "",
                "selected": selected,
                "status": status_text,
            }
            rows.append(row)
            log_lines.append(f"{item['order']}. {source_code} â†’ {dest_code} - {item['title']} - {status_text}")
        prepare_id = uuid.uuid4().hex
        root = RUNTIME / ("contest_lesson_copy_" + prepare_id)
        root.mkdir(parents=True, exist_ok=True)
        prepared_lesson_copies[prepare_id] = {
            "created_at": time.time(),
            "source": source,
            "contest_key": contest_key,
            "course_slug": course_slug,
            "lesson_id": lesson_id,
            "rows": rows,
            "root": root,
        }
        return jsonify(
            {
                "ok": True,
                "prepare_id": prepare_id,
                "rows": rows,
                "can_copy": any(row.get("selected") for row in rows),
                "lesson_link": hncode_lesson_url(course_slug, lesson_id),
                "log": "\n".join(log_lines),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/confirm-contest-to-lesson")
def api_confirm_contest_to_lesson():
    payload = request.get_json(force=True)
    account = payload.get("account", {})
    source_account = payload.get("source_account", {})
    prepare_id = payload.get("prepare_id", "")
    state = prepared_lesson_copies.get(prepare_id)
    if not state:
        return jsonify({"ok": False, "error": "Dá»¯ liá»‡u chuáº©n bá»‹ Ä‘Ã£ háº¿t háº¡n. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i."}), 400
    try:
        rows_by_code = {row["code"]: row for row in state["rows"]}
        requested_rows = payload.get("rows", [])
        selected_refs = []
        result_rows = []
        source = state.get("source", "hncode")
        source_label = "HNOJ" if source == "hnoj" else "HNCode"
        log_lines = [
            f"Sao chÃ©p bÃ i tá»« Contest {source_label} â†’ Lesson HNCode",
            f"Contest: {state['contest_key']}",
            f"Lesson: {hncode_lesson_url(state['course_slug'], state['lesson_id'])}",
        ]
        dst_session = None
        src_session = None
        for requested in requested_rows:
            code = requested.get("code", "")
            base = dict(rows_by_code.get(code, requested))
            base["selected"] = bool(requested.get("selected"))
            base["score"] = str(requested.get("score") or base.get("score") or "100")
            if not base["selected"]:
                base["status"] = "Bá» qua"
            elif "ÄÃ£ cÃ³" in str(rows_by_code.get(code, {}).get("status", "")):
                base["status"] = "ÄÃ£ cÃ³ trong lesson"
            else:
                if not dst_session:
                    dst_session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
                if not base.get("problem_id") and source == "hnoj":
                    if not src_session:
                        src_session = login_hncode(TARGETS["hnoj"]["base_url"], source_account.get("username", ""), source_account.get("password", ""))
                    source_code = base.get("source_code") or code
                    log_lines.append(f"Äang chuyá»ƒn {source_code} sang HNCode...")
                    try:
                        info, zip_path, cases, _attachments = fetch_source_problem(src_session, TARGETS["hnoj"]["base_url"], source_code, state["root"])
                        upload_transfer_to_dmoj(
                            dst_session,
                            "hncode",
                            code,
                            info,
                            zip_path,
                            cases,
                            {"upload_statement": True, "upload_tests": True},
                            list(TARGETS["hncode"]["languages"].values()),
                            log_lines,
                        )
                    except ProblemAlreadyExists:
                        log_lines.append(f"{code}: bÃ i Ä‘Ã£ cÃ³ trÃªn HNCode, dÃ¹ng láº¡i bÃ i hiá»‡n cÃ³.")
                    base["problem_id"] = admin_problem_id(dst_session, TARGETS["hncode"]["base_url"], code) or ""
                if not base.get("problem_id"):
                    base["status"] = "âœ— KhÃ´ng tÃ¬m tháº¥y bÃ i trong admin HNCode"
                else:
                    selected_refs.append(base)
                    base["status"] = "Äang thÃªm..."
            result_rows.append(base)
        link = hncode_lesson_url(state["course_slug"], state["lesson_id"])
        if selected_refs:
            if not dst_session:
                dst_session = login_hncode(TARGETS["hncode"]["base_url"], account.get("username", ""), account.get("password", ""))
            added_ids: set[str] = set()
            failed_by_id: dict[str, str] = {}
            for ref in selected_refs:
                problem_id = str(ref.get("problem_id") or ref.get("id") or "")
                try:
                    link = copy_hncode_contest_to_lesson(dst_session, state["course_slug"], state["lesson_id"], [ref])
                    added_ids.add(problem_id)
                except Exception as item_exc:
                    failed_by_id[problem_id] = str(item_exc)
            for row in result_rows:
                if str(row.get("problem_id")) in added_ids and row.get("selected"):
                    row["status"] = "âœ“ ÄÃ£ thÃªm"
                    row["link"] = link
                    log_lines.append(f"âœ“ {row['code']}: Ä‘Ã£ thÃªm vÃ o lesson.")
                elif str(row.get("problem_id")) in failed_by_id and row.get("selected"):
                    row["status"] = "âœ— Lá»—i"
                    row["error"] = failed_by_id[str(row.get("problem_id"))]
                    log_lines.append(f"âœ— {row.get('code')}: {row['error']}")
                elif row["status"] == "Bá» qua":
                    log_lines.append(f"- {row.get('code')}: bá» qua.")
                elif row["status"] == "ÄÃ£ cÃ³ trong lesson":
                    log_lines.append(f"- {row.get('code')}: Ä‘Ã£ cÃ³ trong lesson.")
        else:
            log_lines.append("KhÃ´ng cÃ³ bÃ i má»›i Ä‘Æ°á»£c chá»n Ä‘á»ƒ thÃªm.")
        ok = all(not row.get("selected") or row.get("status", "").startswith("âœ“") or "ÄÃ£ cÃ³" in row.get("status", "") for row in result_rows)
        return jsonify({"ok": ok, "rows": result_rows, "link": link, "log": "\n".join(log_lines)})
    except Exception as exc:
        rows = payload.get("rows", [])
        for row in rows:
            row["status"] = "âœ— Lá»—i"
            row["error"] = str(exc)
        return jsonify({"ok": False, "rows": rows, "error": str(exc)}), 400


def decode_text_smart(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def extract_hncode_contest_key_any(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        raise RuntimeError("ChÆ°a nháº­p URL hoáº·c mÃ£ contest.")
    match = re.search(r"/contest/([^/?#\s]+)", value)
    return match.group(1) if match else value.strip().strip("/")


def hncode_student_session(username: str, password: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    login_url = urljoin(TARGETS["hncode"]["base_url"], "/accounts/login/?next=/")
    page = session.get(login_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c trang Ä‘Äƒng nháº­p HNCode: HTTP {page.status_code}")
    result = session.post(
        login_url,
        data={"username": username, "password": password, "csrfmiddlewaretoken": csrf_token(page.text), "next": "/"},
        headers={"Referer": login_url},
        allow_redirects=True,
        timeout=30,
    )
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if not result.ok or errors:
        raise RuntimeError("Form Ä‘Äƒng nháº­p HNCode bÃ¡o lá»—i: " + "; ".join(errors or [f"HTTP {result.status_code}"]))
    if "sessionid" not in session.cookies.get_dict() or "/accounts/login" in result.url:
        raise RuntimeError("HNCode login did not create a session")
    return session


def parse_hncode_contest_problems(session: requests.Session, contest_key: str) -> list[dict]:
    page = session.get(urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}/problems"), timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c danh sÃ¡ch bÃ i contest {contest_key}: HTTP {page.status_code}")
    rows = extract_contest_problem_rows_from_html(page.text, contest_key, "100")
    if not rows:
        ranking = session.get(urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}/ranking/"), timeout=30)
        if ranking.ok:
            rows = extract_contest_problem_rows_from_html(ranking.text, contest_key, "100")
    if not rows:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y bÃ i nÃ o trong contest {contest_key}.")
    for row in rows:
        try:
            row["points"] = float(str(row.get("points") or "100").replace(",", "."))
        except ValueError:
            row["points"] = 100.0
    change_url = admin_contest_change_url(session, TARGETS["hncode"]["base_url"], contest_key)
    if change_url:
        admin_page = session.get(change_url, timeout=30)
        if admin_page.ok:
            total = int(input_value(admin_page.text, "contest_problems-TOTAL_FORMS", "0") or "0")
            point_rows = []
            for idx in range(total):
                points = input_value(admin_page.text, f"contest_problems-{idx}-points", "")
                order = input_value(admin_page.text, f"contest_problems-{idx}-order", str(idx + 1)) or str(idx + 1)
                if points:
                    try:
                        point_rows.append((int(float(order)), float(str(points).replace(",", "."))))
                    except ValueError:
                        pass
            point_rows.sort(key=lambda item: item[0])
            if len(point_rows) >= len(rows):
                for row, (_order, points) in zip(rows, point_rows):
                    row["points"] = points
    return rows


def read_hncode_grading_accounts(csv_path: Path) -> list[dict]:
    reader = csv.DictReader(decode_text_smart(csv_path.read_bytes()).splitlines())
    missing = {"username", "password", "name"} - set(reader.fieldnames or [])
    if missing:
        raise RuntimeError("File tÃ i khoáº£n thiáº¿u cá»™t: " + ", ".join(sorted(missing)))
    accounts = []
    for index, row in enumerate(reader, 1):
        username = (row.get("username") or "").strip()
        password = (row.get("password") or "").strip()
        name = (row.get("name") or "").strip()
        if username and password and name:
            accounts.append({"index": index, "username": username, "password": password, "name": name})
    if not accounts:
        raise RuntimeError("KhÃ´ng Ä‘á»c Ä‘Æ°á»£c tÃ i khoáº£n há»£p lá»‡ nÃ o trong file CSV.")
    return accounts


def normalize_grading_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def grading_source_root(extract_root: Path) -> Path:
    dirs = [item for item in extract_root.iterdir() if item.is_dir()]
    return dirs[0] if len(dirs) == 1 and any(path.is_file() for path in dirs[0].rglob("*")) else extract_root


def map_grading_problem_code(stem: str, contest_problems: list[dict]) -> str:
    raw = re.sub(r"[^A-Za-z0-9_]+", "", stem).lower()
    codes = [problem["code"] for problem in contest_problems]
    if raw in codes:
        return raw
    for code in codes:
        if code.startswith(raw + "_") or code.split("_", 1)[0] == raw:
            return code
    for code in codes:
        if raw in code.replace("_", "") or code.replace("_", "") in raw:
            return code
    return raw


def collect_hncode_grading_files(source_root: Path, accounts: list[dict], contest_problems: list[dict]) -> tuple[list[dict], list[str]]:
    account_by_key = {normalize_grading_key(account["name"]): account for account in accounts}
    problem_by_code = {problem["code"]: problem for problem in contest_problems}
    allowed_suffixes = {".cpp", ".cc", ".cxx", ".c", ".py", ".pas"}
    rows: list[dict] = []
    warnings: list[str] = []
    for student_dir in sorted((item for item in source_root.iterdir() if item.is_dir()), key=lambda path: path.name.lower()):
        account = account_by_key.get(normalize_grading_key(student_dir.name))
        if not account:
            warnings.append(f"KhÃ´ng tÃ¬m tháº¥y tÃ i khoáº£n CSV cho thÆ° má»¥c {student_dir.name}.")
            continue
        files = sorted((path for path in student_dir.rglob("*") if path.is_file() and path.suffix.lower() in allowed_suffixes), key=lambda path: path.name.lower())
        if not files:
            warnings.append(f"ThÆ° má»¥c {student_dir.name} khÃ´ng cÃ³ file code.")
            continue
        for path in files:
            code = map_grading_problem_code(path.stem, contest_problems)
            problem = problem_by_code.get(code)
            rows.append({
                "original_key": f"{account['username']}::{path.relative_to(source_root).as_posix()}",
                "selected": bool(problem),
                "student": account["name"],
                "username": account["username"],
                "problem": code,
                "problem_title": problem["title"] if problem else "",
                "contest_points": problem["points"] if problem else 0,
                "language": path.suffix.lower().lstrip("."),
                "file": path.name,
                "relative_path": path.relative_to(source_root).as_posix(),
                "local_path": str(path),
                "status": "ÄÃ£ chuáº©n bá»‹" if problem else "KhÃ´ng khá»›p bÃ i trong contest",
                "submission_url": "",
                "percent": "",
                "score": "",
                "message": "",
            })
    if not rows:
        raise RuntimeError("KhÃ´ng tÃ¬m tháº¥y file bÃ i lÃ m nÃ o khá»›p tÃ i khoáº£n trong zip.")
    return rows, warnings


def join_hncode_contest_if_needed(session: requests.Session, contest_key: str, contest_password: str) -> str:
    join_url = urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}/join")
    page = session.get(join_url, timeout=30, allow_redirects=True)
    if f"/contest/{contest_key}/problems" in page.url:
        return "ÄÃ£ tham gia"
    parser = FormDataParser()
    parser.feed(page.text)
    form = next((item for item in parser.forms if any(name == "access_code" for name, _value in item)), None)
    if not form:
        return "KhÃ´ng cáº§n nháº­p máº­t kháº©u"
    result = session.post(
        join_url,
        data=[(name, contest_password if name == "access_code" else value) for name, value in form],
        headers={"Referer": join_url},
        allow_redirects=True,
        timeout=30,
    )
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Join contest bÃ¡o lá»—i: " + "; ".join(errors))
    return "ÄÃ£ nháº­p máº­t kháº©u contest"


def preferred_languages_for_source(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".cpp", ".cc", ".cxx"}:
        return ["C++17", "GNU C++17", "C++20", "GNU C++20", "C++"]
    if suffix == ".c":
        return ["C", "C11", "GNU C"]
    if suffix == ".py":
        return ["PyPy 3", "Pypy 3", "Python 3", "Python3", "Python"]
    if suffix == ".pas":
        return ["Pascal", "FPC"]
    return ["C++17", "C++"]


def submit_hncode_grading_file(session: requests.Session, problem_code: str, source_path: Path) -> str:
    submit_url = urljoin(TARGETS["hncode"]["base_url"], f"/problem/{problem_code}/submit")
    page = session.get(submit_url, timeout=30, allow_redirects=True)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c trang ná»™p bÃ i {problem_code}: HTTP {page.status_code}")
    language_id = language_id_from_submit_page(page.text, preferred_languages_for_source(source_path))
    if not language_id:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y ngÃ´n ngá»¯ phÃ¹ há»£p cho file {source_path.name}")
    result = session.post(
        submit_url,
        data={"csrfmiddlewaretoken": csrf_token(page.text), "source": read_text_smart(source_path), "language": language_id, "judge": ""},
        headers={"Referer": submit_url},
        allow_redirects=True,
        timeout=30,
    )
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if not result.ok or errors:
        raise RuntimeError("Submit form bÃ¡o lá»—i: " + "; ".join(errors or [f"HTTP {result.status_code}"]))
    if "/submission/" not in result.url:
        raise RuntimeError(f"Submit chÆ°a táº¡o submission; URL sau POST: {result.url}")
    return result.url


def parse_hncode_submission_result(page: str) -> dict:
    plain = html.unescape(re.sub(r"<[^>]+>", " ", page))
    plain = re.sub(r"\s+", " ", plain)
    total_match = re.search(r"Tá»•ng cá»™ng:\s*([0-9]+(?:[.,][0-9]+)?)\s*/\s*100", plain, re.I)
    score_match = re.search(r"Äiá»ƒm:\s*([0-9]+(?:[.,][0-9]+)?)\s*/\s*([0-9]+(?:[.,][0-9]+)?)", plain, re.I)
    verdict = ""
    for candidate in ["Accepted", "Wrong Answer", "Time Limit Exceeded", "Runtime Error", "Compilation Error", "Memory Limit Exceeded", "Output Limit Exceeded"]:
        if candidate in plain:
            verdict = candidate
            break
    if total_match:
        return {"done": True, "percent": float(total_match.group(1).replace(",", ".")), "verdict": verdict or "Done"}
    if score_match:
        got = float(score_match.group(1).replace(",", "."))
        total = float(score_match.group(2).replace(",", "."))
        return {"done": True, "percent": 100.0 * got / total if total else 0.0, "verdict": verdict or "Done"}
    if verdict and verdict in {"Compilation Error", "Runtime Error", "Wrong Answer", "Time Limit Exceeded", "Memory Limit Exceeded", "Output Limit Exceeded"}:
        return {"done": True, "percent": 0.0, "verdict": verdict}
    pending = any(word.lower() in plain.lower() for word in ["Queued", "Äang cháº¥m", "Processing", "grading", "Chá» cháº¥m"])
    return {"done": not pending, "percent": None, "verdict": verdict or "Äang cháº¥m"}


def poll_hncode_submission(session: requests.Session, submission_url: str) -> dict:
    while True:
        page = session.get(submission_url, timeout=30)
        if not page.ok:
            raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c submission: HTTP {page.status_code}")
        result = parse_hncode_submission_result(page.text)
        if result.get("done") and result.get("percent") is not None:
            return result
        time.sleep(2)


def html_cell_text(fragment: str) -> str:
    text = html.unescape(re.sub(r"<[^>]+>", " ", fragment))
    return re.sub(r"\s+", " ", text).strip()


def number_from_rank_text(value: str) -> float | str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return round(float(normalized), 2)
    except ValueError:
        return text


def parse_hncode_ranking_table(page: str) -> tuple[list[dict], list[str]]:
    table_match = re.search(r'<table\b[^>]*id=["\']users-table["\'][^>]*>([\s\S]*?)</table>', page, re.I)
    if not table_match:
        return [], []
    table = table_match.group(1)
    thead = re.search(r"<thead\b[^>]*>([\s\S]*?)</thead>", table, re.I)
    problem_codes = []
    if thead:
        for th in re.findall(r"<th\b[^>]*problem-score-col[^>]*>([\s\S]*?)</th>", thead.group(1), re.I):
            code_match = re.search(r'class=["\']problem-code["\'][^>]*>([\s\S]*?)</div>', th, re.I)
            if code_match:
                problem_codes.append(html_cell_text(code_match.group(1)))
    tbody = re.search(r"<tbody\b[^>]*>([\s\S]*?)</tbody>", table, re.I)
    if not tbody:
        return [], problem_codes
    ranking_rows = []
    for tr in re.findall(r"<tr\b[^>]*>([\s\S]*?)</tr>", tbody.group(1), re.I):
        cells = re.findall(r"<td\b([^>]*)>([\s\S]*?)</td>", tr, re.I)
        if len(cells) < 3:
            continue
        rank = html_cell_text(cells[0][1])
        user_cell = cells[1][1]
        username_match = re.search(r'<a\b[^>]*href=["\']/user/[^"\']+["\'][^>]*>([\s\S]*?)</a>', user_cell, re.I)
        fullname_match = re.search(r'class=["\'][^"\']*\bfullname\b[^"\']*["\'][^>]*>([\s\S]*?)</div>', user_cell, re.I)
        participation_match = re.search(r"<sub\b[^>]*>\s*\[([0-9]+)\]\s*</sub>", user_cell, re.I)
        total_cell = cells[2][1]
        total_score = html_cell_text(re.sub(r'<div\b[^>]*class=["\']solving-time["\'][\s\S]*?</div>', "", total_cell, flags=re.I))
        total_time_match = re.search(r'class=["\']solving-time["\'][^>]*>([\s\S]*?)</div>', total_cell, re.I)
        item = {
            "rank": number_from_rank_text(rank),
            "username": html_cell_text(username_match.group(1)) if username_match else html_cell_text(user_cell),
            "fullname": html_cell_text(fullname_match.group(1)) if fullname_match else "",
            "participation": participation_match.group(1) if participation_match else "",
            "total": number_from_rank_text(total_score),
            "time": html_cell_text(total_time_match.group(1)) if total_time_match else "",
            "scores": {},
            "times": {},
        }
        problem_cells = [cell for attrs, cell in cells[3:] if "problem-score-col" in attrs]
        for code, cell in zip(problem_codes, problem_cells):
            score_match = re.search(r"<span\b[^>]*>([\s\S]*?)</span>", cell, re.I)
            time_match = re.search(r'class=["\']solving-time["\'][^>]*>([\s\S]*?)</div>', cell, re.I)
            score_text = html_cell_text(score_match.group(1)) if score_match else html_cell_text(cell)
            item["scores"][code] = number_from_rank_text(score_text)
            item["times"][code] = html_cell_text(time_match.group(1)) if time_match else ""
        ranking_rows.append(item)
    return ranking_rows, problem_codes


def fetch_hncode_contest_ranking(session: requests.Session, contest_key: str) -> tuple[list[dict], list[str]]:
    all_rows: list[dict] = []
    problem_codes: list[str] = []
    seen_keys: set[tuple[str, str, str]] = set()
    for page_num in range(1, 51):
        page = session.get(
            urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}/ranking/"),
            params={"friend": "0", "virtual": "1", "page": str(page_num)},
            timeout=30,
        )
        if not page.ok:
            raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c báº£ng rank contest: HTTP {page.status_code}")
        rows, codes = parse_hncode_ranking_table(page.text)
        if codes and not problem_codes:
            problem_codes = codes
        new_rows = []
        for row in rows:
            key = (str(row.get("username", "")), str(row.get("participation", "")), str(row.get("rank", "")))
            if key not in seen_keys:
                seen_keys.add(key)
                new_rows.append(row)
        if not new_rows:
            break
        all_rows.extend(new_rows)
        if len(rows) < 100:
            break
    return all_rows, problem_codes


def write_hncode_grading_excel(rows: list[dict], contest_problems: list[dict], accounts: list[dict], output_path: Path, ranking_rows: list[dict] | None = None, ranking_problem_codes: list[str] | None = None) -> None:
    from openpyxl import Workbook
    from copy import copy

    wb = Workbook()
    ws = wb.active
    ws.title = "Bang diem"
    problem_codes = list(ranking_problem_codes or []) or [problem["code"] for problem in contest_problems]
    if ranking_rows:
        ws.append(["Rank", "Username", "Há» tÃªn", "LÆ°á»£t áº£o", "Tá»•ng Ä‘iá»ƒm", "Thá»i gian", *problem_codes])
        for item in ranking_rows:
            ws.append([
                item.get("rank", ""),
                item.get("username", ""),
                item.get("fullname", ""),
                item.get("participation", ""),
                item.get("total", ""),
                item.get("time", ""),
                *[item.get("scores", {}).get(code, "") for code in problem_codes],
            ])
    else:
        ws.append(["STT", "Há»c sinh", "Username", *problem_codes, "Tá»•ng Ä‘iá»ƒm", "Sá»‘ bÃ i Ä‘Ã£ ná»™p"])
        by_student_problem: dict[tuple[str, str], dict] = {}
        for row in rows:
            key = (row.get("username", ""), row.get("problem", ""))
            current = by_student_problem.get(key)
            if current is None or float(row.get("score") or 0) > float(current.get("score") or 0):
                by_student_problem[key] = row
        for account in accounts:
            total = 0.0
            count = 0
            values = [account["index"], account["name"], account["username"]]
            for code in problem_codes:
                row = by_student_problem.get((account["username"], code))
                score = float(row.get("score") or 0) if row else 0.0
                total += score
                if row and row.get("submission_url"):
                    count += 1
                values.append(round(score, 2))
            values.extend([round(total, 2), count])
            ws.append(values)
    for cell in ws[1]:
        cell.font = copy(cell.font)
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "D2"
    autosize_worksheet(ws)
    ws = wb.create_sheet("Chi tiet nop bai")
    ws.append(["Há»c sinh", "Username", "MÃ£ bÃ i", "TÃªn bÃ i", "Äiá»ƒm bÃ i", "%", "Äiá»ƒm quy Ä‘á»•i", "File", "Tráº¡ng thÃ¡i", "Submission", "ThÃ´ng bÃ¡o"])
    for row in rows:
        ws.append([row.get("student"), row.get("username"), row.get("problem"), row.get("problem_title"), row.get("contest_points"), row.get("percent"), row.get("score"), row.get("relative_path"), row.get("status"), "Má»Ÿ submission" if row.get("submission_url") else "", row.get("message")])
        if row.get("submission_url"):
            ws.cell(ws.max_row, 10).hyperlink = row["submission_url"]
            ws.cell(ws.max_row, 10).style = "Hyperlink"
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)
    ws = wb.create_sheet("Danh sach bai")
    ws.append(["Thá»© tá»±", "MÃ£ bÃ i", "TÃªn bÃ i", "Äiá»ƒm contest"])
    for problem in contest_problems:
        ws.append([problem["order"], problem["code"], problem["title"], problem["points"]])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    autosize_worksheet(ws)
    wb.save(output_path)


@app.post("/api/prepare-hncode-grading")
def api_prepare_hncode_grading():
    progress_id = request.form.get("progress_id")
    try:
        contest_key = extract_hncode_contest_key_any(request.form.get("contest_url", ""))
        zip_file = request.files.get("zip_file")
        csv_file = request.files.get("csv_file")
        if not zip_file or not zip_file.filename:
            return jsonify({"error": "ChÆ°a chá»n file zip bÃ i lÃ m."}), 400
        if not csv_file or not csv_file.filename:
            return jsonify({"error": "ChÆ°a chá»n file CSV tÃ i khoáº£n."}), 400
        prepare_id = uuid.uuid4().hex
        root = RUNTIME / ("hncode_grading_" + prepare_id)
        source_zip = root / "bai_lam.zip"
        account_csv = root / "tai_khoan.csv"
        extract_root = root / "extract"
        root.mkdir(parents=True, exist_ok=True)
        zip_file.save(source_zip)
        csv_file.save(account_csv)
        progress_update(progress_id, phase="prepare-hncode-grading", done=0, total=3, rows=[], message="Äang Ä‘á»c contest HNCode")
        admin_username = request.form.get("admin_username", "")
        admin_password = request.form.get("admin_password", "")
        admin_session = login_hncode(TARGETS["hncode"]["base_url"], admin_username, admin_password)
        contest_problems = parse_hncode_contest_problems(admin_session, contest_key)
        accounts = read_hncode_grading_accounts(account_csv)
        safe_extract_zip(source_zip, extract_root)
        source_root = grading_source_root(extract_root)
        rows, warnings = collect_hncode_grading_files(source_root, accounts, contest_problems)
        prepared_hncode_grading[prepare_id] = {"root": root, "source_root": source_root, "contest_key": contest_key, "contest_problems": contest_problems, "accounts": accounts, "rows": rows, "output": "", "admin_username": admin_username}
        log_lines = [f"Contest: {contest_key}", f"ÄÃ£ Ä‘á»c {len(contest_problems)} bÃ i: " + ", ".join(problem["code"] for problem in contest_problems), f"ÄÃ£ Ä‘á»c {len(accounts)} tÃ i khoáº£n.", f"ÄÃ£ tÃ¬m tháº¥y {len(rows)} file bÃ i lÃ m."]
        log_lines.extend(f"- {warning}" for warning in warnings)
        progress_update(progress_id, phase="prepare-hncode-grading", done=3, total=3, rows=rows, message="ÄÃ£ chuáº©n bá»‹ dá»¯ liá»‡u cháº¥m")
        progress_finish(progress_id, True, "ÄÃ£ chuáº©n bá»‹ dá»¯ liá»‡u cháº¥m")
        return jsonify({"prepare_id": prepare_id, "rows": rows, "problems": contest_problems, "accounts": accounts, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


@app.post("/api/confirm-hncode-grading")
def api_confirm_hncode_grading():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id")
    state = prepared_hncode_grading.get(payload.get("prepare_id", ""))
    if not state:
        progress_finish(progress_id, False, "Dá»¯ liá»‡u chuáº©n bá»‹ cháº¥m Ä‘Ã£ háº¿t háº¡n")
        return jsonify({"error": "Dá»¯ liá»‡u chuáº©n bá»‹ cháº¥m Ä‘Ã£ háº¿t háº¡n. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i."}), 400
    try:
        requested = {row.get("original_key"): row for row in payload.get("rows", [])}
        rows = []
        for base in state["rows"]:
            row = dict(base)
            if row["original_key"] in requested:
                row["selected"] = bool(requested[row["original_key"]].get("selected"))
            rows.append(row)
        selected_rows = [row for row in rows if row.get("selected")]
        if not selected_rows:
            raise RuntimeError("ChÆ°a chá»n bÃ i nÃ o Ä‘á»ƒ ná»™p cháº¥m.")
        contest_password = payload.get("contest_password", "")
        account_by_username = {account["username"]: account for account in state["accounts"]}
        sessions: dict[str, requests.Session] = {}
        done = 0
        log_lines = [f"Cháº¥m bÃ i HNCode contest {state['contest_key']}: {len(selected_rows)} file Ä‘Æ°á»£c chá»n."]
        progress_update(progress_id, phase="confirm-hncode-grading", done=0, total=len(selected_rows), rows=rows, message="Báº¯t Ä‘áº§u ná»™p bÃ i")
        for row in rows:
            if not row.get("selected"):
                row["status"] = "Bá» qua"
                continue
            try:
                account = account_by_username[row["username"]]
                session = sessions.get(row["username"])
                if session is None:
                    session = hncode_student_session(account["username"], account["password"])
                    log_lines.append(f"{account['name']} ({account['username']}): {join_hncode_contest_if_needed(session, state['contest_key'], contest_password)}.")
                    sessions[row["username"]] = session
                progress_update(progress_id, phase="confirm-hncode-grading", done=done, total=len(selected_rows), rows=rows, message=f"{row['student']} - {row['problem']}: Ä‘ang ná»™p")
                row["status"] = "Äang ná»™p"
                row["submission_url"] = submit_hncode_grading_file(session, row["problem"], Path(row["local_path"]))
                result = poll_hncode_submission(session, row["submission_url"])
                percent = result.get("percent")
                row["percent"] = "" if percent is None else round(float(percent), 2)
                row["score"] = "" if percent is None else round(float(row.get("contest_points") or 0) * float(percent) / 100.0, 2)
                row["status"] = "âœ“ ÄÃ£ cháº¥m" if percent is not None else "âœ“ ÄÃ£ ná»™p"
                row["message"] = result.get("verdict") or ""
                log_lines.append(f"âœ“ {row['student']} - {row['problem']}: {row['message']}, {row['percent']}%, Ä‘iá»ƒm {row['score']}.")
            except Exception as exc:
                row["status"] = "âœ— Lá»—i"
                row["message"] = str(exc)
                log_lines.append(f"âœ— {row.get('student')} - {row.get('problem')}: {exc}")
            done += 1
            progress_update(progress_id, phase="confirm-hncode-grading", done=done, total=len(selected_rows), rows=rows, message=f"{row.get('student')} - {row.get('problem')}: {row.get('status')}")
        output_path = Path(state["root"]) / "bang_diem_hncode.xlsx"
        ranking_rows: list[dict] = []
        ranking_problem_codes: list[str] = []
        try:
            admin_account = payload.get("admin_account", {})
            rank_session = login_hncode(TARGETS["hncode"]["base_url"], admin_account.get("username", ""), admin_account.get("password", ""))
            ranking_rows, ranking_problem_codes = fetch_hncode_contest_ranking(rank_session, state["contest_key"])
            log_lines.append(f"ÄÃ£ Ä‘á»c láº¡i báº£ng rank contest: {len(ranking_rows)} dÃ²ng.")
        except Exception as exc:
            log_lines.append(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c báº£ng rank, Excel dÃ¹ng dá»¯ liá»‡u submission vá»«a ná»™p: {exc}")
        write_hncode_grading_excel(rows, state["contest_problems"], state["accounts"], output_path, ranking_rows, ranking_problem_codes)
        state["rows"] = rows
        state["output"] = str(output_path)
        ok = all((not row.get("selected")) or str(row.get("status", "")).startswith("âœ“") for row in rows)
        progress_finish(progress_id, ok, "ÄÃ£ hoÃ n táº¥t cháº¥m bÃ i")
        return jsonify({"ok": ok, "rows": rows, "log": "\n".join(log_lines), "download_url": f"/api/download-hncode-grading/{payload.get('prepare_id', '')}"})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


@app.get("/api/download-hncode-grading/<prepare_id>")
def api_download_hncode_grading(prepare_id: str):
    state = prepared_hncode_grading.get(prepare_id)
    if not state or not state.get("output"):
        return jsonify({"error": "ChÆ°a cÃ³ file báº£ng Ä‘iá»ƒm Ä‘á»ƒ táº£i."}), 404
    path = Path(state["output"])
    if not path.exists():
        return jsonify({"error": "File báº£ng Ä‘iá»ƒm khÃ´ng cÃ²n tá»“n táº¡i."}), 404
    return send_file(path, as_attachment=True, download_name=path.name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/progress/<progress_id>")
def api_progress(progress_id: str):
    if not valid_progress_id(progress_id):
        return jsonify({"error": "progress_id khÃ´ng há»£p lá»‡"}), 400
    return jsonify(job_service.read_job(PROGRESS_DIR, progress_id))


@app.post("/api/misc/last-submissions")
def api_misc_last_submissions():
    uploaded = request.files.get("zip_file")
    if not uploaded or not uploaded.filename:
        return jsonify({"error": "ChÆ°a chá»n file zip data."}), 400
    if Path(uploaded.filename).suffix.lower() != ".zip":
        return jsonify({"error": "File data pháº£i lÃ  .zip."}), 400
    job_root = RUNTIME / "misc" / uuid.uuid4().hex
    input_zip = job_root / "input.zip"
    extract_root = job_root / "extract"
    output_dir = job_root / "Last_Submissions"
    try:
        job_root.mkdir(parents=True, exist_ok=True)
        uploaded.save(input_zip)
        safe_extract_zip(input_zip, extract_root)
        data_root = find_scratch_data_root(extract_root)
        summary = collect_last_scratch_submissions(data_root, output_dir)
        if summary["total"] == 0:
            return jsonify({"error": "KhÃ´ng tÃ¬m tháº¥y thÆ° má»¥c thÃ­ sinh nÃ o trong file zip."}), 400
        output_zip = job_root / "last_submissions.zip"
        with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for item in sorted(output_dir.iterdir(), key=lambda path: path.name.lower()):
                if item.is_file():
                    zf.write(item, item.name)
        summary_payload = {
            "total": summary["total"],
            "found": summary["found"],
            "missing": summary["missing"],
            "filename": output_zip.name,
        }
        response = send_file(output_zip, as_attachment=True, download_name=output_zip.name, mimetype="application/zip")
        response.headers["X-Last-Submissions-Summary"] = quote(json.dumps(summary_payload, ensure_ascii=False))
        return response
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/misc/ai-code-warning")
def api_misc_ai_code_warning():
    job_root = RUNTIME / "misc" / uuid.uuid4().hex
    try:
        job_root.mkdir(parents=True, exist_ok=True)
        uploaded = request.files.get("zip_file")
        source_zips: list[Path] = []
        if uploaded and uploaded.filename:
            if Path(uploaded.filename).suffix.lower() != ".zip":
                return jsonify({"error": "File data pháº£i lÃ  .zip."}), 400
            input_zip = job_root / safe_output_part(Path(uploaded.filename).name)
            uploaded.save(input_zip)
            source_zips = [input_zip]
        else:
            folder_path = (request.form.get("folder_path") or "").strip()
            if not folder_path:
                return jsonify({"error": "HÃ£y chá»n file zip hoáº·c nháº­p folder chá»©a cÃ¡c zip contest."}), 400
            folder = Path(folder_path)
            if not folder.exists() or not folder.is_dir():
                return jsonify({"error": f"Folder khÃ´ng tá»“n táº¡i: {folder_path}"}), 400
            source_zips = sorted(folder.glob("*.zip"))
            if not source_zips:
                return jsonify({"error": f"KhÃ´ng tÃ¬m tháº¥y file .zip nÃ o trong folder: {folder_path}"}), 400
        output_path = job_root / "ai_code_warning_report.xlsx"
        summary = build_ai_warning_report(source_zips, output_path)
        response = send_file(
            output_path,
            as_attachment=True,
            download_name=output_path.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["X-AI-Warning-Summary"] = quote(json.dumps(summary, ensure_ascii=False))
        return response
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/tinhoctre-browser/start")
def api_tinhoctre_browser_start():
    try:
        browser = find_edge_executable()
        port = int(os.getenv("TINHOCTRE_CHROME_DEBUG_PORT", "9223"))
        url = "https://tinhoctre.vn/admin/judge/problem/add/"
        subprocess.Popen(
            [
                str(browser),
                f"--remote-debugging-port={port}",
                "--remote-debugging-address=127.0.0.1",
                "--remote-allow-origins=*",
                "--profile-directory=Default",
                "--new-window",
                url,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return jsonify(
            {
                "ok": True,
                "message": "ÄÃ£ má»Ÿ Edge báº±ng profile máº·c Ä‘á»‹nh. HÃ£y Ä‘Äƒng nháº­p admin vÃ  Ä‘áº£m báº£o tháº¥y form táº¡o bÃ i, rá»“i báº¥m Láº¥y cookie tá»« Edge. Náº¿u khÃ´ng láº¥y Ä‘Æ°á»£c cookie, hÃ£y Ä‘Ã³ng háº¿t Edge rá»“i báº¥m nÃºt nÃ y láº¡i.",
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/tinhoctre-browser/cookie")
def api_tinhoctre_browser_cookie():
    try:
        cookie = cookie_from_tinhoctre_debug_browser()
        save_tinhoctre_cookie(cookie)
        s = session_from_cookie(cookie)
        check = s.get("https://tinhoctre.vn/admin/judge/problem/add/", timeout=30)
        if not check.ok or not is_problem_add_form(check.text):
            raise RuntimeError(tinhoctre_admin_cookie_error(check.url))
        return jsonify({"ok": True, "cookie": cookie, "message": "ÄÃ£ láº¥y Cookie TinHocTre tá»« Edge vÃ  kiá»ƒm tra má»Ÿ Ä‘Æ°á»£c form admin táº¡o bÃ i."})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/tinhoctre-browser/quick-cookie")
def api_tinhoctre_browser_quick_cookie():
    try:
        stop_edge_processes()
        time.sleep(1)
        chrome = find_edge_executable()
        port = int(os.getenv("TINHOCTRE_CHROME_DEBUG_PORT", "9223"))
        url = "https://tinhoctre.vn/admin/judge/problem/add/"
        subprocess.Popen(
            [
                str(chrome),
                f"--remote-debugging-port={port}",
                "--remote-debugging-address=127.0.0.1",
                "--remote-allow-origins=*",
                "--profile-directory=Default",
                "--new-window",
                url,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        time.sleep(4)
        cookie = cookie_from_tinhoctre_debug_browser()
        save_tinhoctre_cookie(cookie)
        s = session_from_cookie(cookie)
        check = s.get(url, timeout=30)
        if not check.ok or not is_problem_add_form(check.text):
            raise RuntimeError(tinhoctre_admin_cookie_error(check.url))
        return jsonify({"ok": True, "cookie": cookie, "message": "ÄÃ£ tá»± Ä‘Ã³ng/má»Ÿ Edge, láº¥y Cookie TinHocTre vÃ  kiá»ƒm tra má»Ÿ Ä‘Æ°á»£c form admin táº¡o bÃ i."})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/prepare-upload")
def api_prepare_upload():
    progress_id = None
    try:
        payload = upload_payload()
        progress_id = payload.get("progress_id")
        prepare_id = uuid.uuid4().hex
        root = RUNTIME / prepare_id
        source_dir = root / "source"
        build_root = root / "generated"
        root.mkdir(parents=True, exist_ok=True)
        build_root.mkdir(parents=True, exist_ok=True)
        source_path = receive_upload_source_file(root, payload)
        if source_path.suffix.lower() == ".md":
            source_dir.mkdir(parents=True, exist_ok=True)
            bundles = split_combined_markdown_bundles(source_path, source_dir)
            tests: dict[str, GeneratedTests | None] = {bundle.code: None for bundle in bundles}
            source_name = source_path.name
            log_lines = [f"ÄÃ£ Ä‘á»c {len(bundles)} bÃ i tá»« file Markdown {source_name}."]
        else:
            extract_zip(source_path, source_dir)
            bundles = discover_bundles(source_dir)
            tests = {}
            source_name = source_path.name
            log_lines = [f"ÄÃ£ Ä‘á»c {len(bundles)} bÃ i tá»« {source_name}."]
        rows = []
        progress_update(progress_id, phase="prepare-upload", done=0, total=len(bundles), rows=rows, message="Báº¯t Ä‘áº§u chuáº©n bá»‹ dá»¯ liá»‡u")
        solutions_md: dict[str, Path | None] = {}
        metadata: dict[str, dict] = {}
        for index, bundle in enumerate(bundles, 1):
            generated = tests.get(bundle.code)
            source = "Markdown tá»•ng há»£p"
            if bundle.generator or bundle.test_zip:
                generated = generate_tests(bundle, build_root)
                tests[bundle.code] = generated
                source = "gentest" if bundle.generator else "zip cÃ³ sáºµn"
            meta = metadata_from_statement(bundle.statement, payload)
            metadata[bundle.code] = meta
            solution_md = find_named_file(source_dir, ["sol"], bundle.index, bundle.code, ".md") if source_path.suffix.lower() != ".md" else None
            solutions_md[bundle.code] = solution_md
            rows.append(
                {
                    "original_code": bundle.code,
                    "code": bundle.code,
                    "name": bundle.name,
                    "points": meta["points"],
                    "tags": meta["tags"],
                    "time_limit": payload.get("time_limit") or "1.0",
                    "memory_limit": payload.get("memory_limit") or "1048576",
                    "partial": meta["partial"],
                    "test_file": generated.zip_path.name if generated else "KhÃ´ng cÃ³ test",
                    "test_count": len(generated.input_files) if generated else 0,
                    "upload_tests_default": bool(generated),
                    "upload_solution_default": bool(solution_md),
                }
            )
            test_text = f"{len(generated.input_files)} test" if generated else "khÃ´ng cÃ³ test"
            solution_text = ", cÃ³ lá»i giáº£i Markdown" if solution_md else ""
            log_lines.append(f"- {bundle.code}: {bundle.name}, Ä‘iá»ƒm {meta['points']}, tags {meta['tags'] or 'trá»‘ng'}, {test_text}, nguá»“n {source}{solution_text}.")
            progress_update(progress_id, phase="prepare-upload", done=index, total=len(bundles), rows=rows, message=f"{bundle.code}: Ä‘Ã£ chuáº©n bá»‹ {test_text}")
        prepared_uploads[prepare_id] = {"root": root, "bundles": {b.code: b for b in bundles}, "tests": tests, "solutions": solutions_md, "metadata": metadata}
        progress_finish(progress_id, True, f"ÄÃ£ chuáº©n bá»‹ {len(bundles)}/{len(bundles)} bÃ i")
        return jsonify({"prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


def upload_payload() -> dict:
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        raw = request.form.get("payload", "{}")
        return json.loads(raw)
    return request.get_json(force=True)


def receive_upload_source_file(root: Path, payload: dict) -> Path:
    uploaded = request.files.get("zip_file")
    if uploaded:
        original = Path(uploaded.filename or "uploaded_package.zip")
        suffix = original.suffix.lower() if original.suffix.lower() in {".zip", ".md"} else ".zip"
        upload_path = root / f"uploaded_package{suffix}"
        uploaded.save(upload_path)
        return upload_path
    source_path = Path(payload["zip_path"])
    if not source_path.exists():
        raise FileNotFoundError(f"KhÃ´ng tÃ¬m tháº¥y file: {source_path}")
    if source_path.suffix.lower() not in {".zip", ".md"}:
        raise RuntimeError("Chá»‰ há»— trá»£ file .zip hoáº·c file Markdown .md.")
    return source_path


def split_combined_markdown_bundles(markdown_path: Path, source_dir: Path) -> list[ProblemBundle]:
    return bundle_service.split_combined_markdown_bundles(markdown_path, source_dir)

def statement_header_parts(statement_path: Path) -> list[str]:
    return bundle_service.statement_header_parts(statement_path)

def metadata_from_statement(statement_path: Path, defaults: dict) -> dict:
    return bundle_service.metadata_from_statement(statement_path, defaults)

@app.post("/api/prepare-single-upload")
def api_prepare_single_upload():
    progress_id = None
    try:
        payload = upload_payload()
        progress_id = payload.get("progress_id")
        target = payload.get("target") or "hncode"
        raw_code = (payload.get("code") or "").strip()
        statement_text = (payload.get("statement_text") or "").strip()
        inferred_name, inferred_code = infer_statement_title(statement_text)
        code = (raw_code or inferred_code).strip().lower()
        name = (payload.get("name") or inferred_name or code).strip()
        if not code:
            raise RuntimeError("HÃ£y nháº­p mÃ£ bÃ i hoáº·c dÃ¹ng dÃ²ng Ä‘áº§u Ä‘á» bÃ i dáº¡ng: TÃªn bÃ i | ma_bai.")
        if not name:
            raise RuntimeError("HÃ£y nháº­p tÃªn bÃ i toÃ¡n.")
        prepare_note = ""
        if target in {"hncode", "tinhoctre"} and not re.fullmatch(r"[a-z0-9_]+", code):
            normalized = normalize_problem_code_for_target(code, target)
            label = TARGETS.get(target, {}).get("label", "HNCode/TinHocTre")
            prepare_note = (
                f"MÃ£ {code} cÃ³ kÃ½ tá»± ngoÃ i chuáº©n táº¡o má»›i cá»§a {label}. Khi xÃ¡c nháº­n, náº¿u bÃ i nÃ y Ä‘Ã£ tá»“n táº¡i thÃ¬ tool dÃ¹ng Ä‘Ãºng mÃ£ nÃ y; "
                f"náº¿u táº¡o má»›i thÃ¬ Ä‘á»•i thÃ nh {normalized}."
            )

        prepare_id = uuid.uuid4().hex
        root = RUNTIME / prepare_id
        source_dir = root / "single"
        build_root = root / "generated"
        source_dir.mkdir(parents=True, exist_ok=True)
        build_root.mkdir(parents=True, exist_ok=True)

        statement_path = source_dir / f"{code}.md"
        if statement_text:
            statement_path.write_text(statement_text.strip() + "\n", encoding="utf-8")
        else:
            statement_path.write_text(f"{name} | {code}\n", encoding="utf-8")

        generator_path: Path | None = None
        generator_text = (payload.get("generator_text") or "").strip()
        generator_filename = Path(payload.get("generator_filename") or "").name
        if generator_text:
            suffix = Path(generator_filename).suffix.lower() if generator_filename else ".py"
            if suffix not in {".py", ".cpp"}:
                suffix = ".py"
            generator_path = source_dir / f"gentest_{code}{suffix}"
            generator_path.write_text(repair_python_main_guard(generator_text) + "\n", encoding="utf-8")

        test_zip_path: Path | None = None
        uploaded_test_zip = request.files.get("test_zip")
        if uploaded_test_zip and uploaded_test_zip.filename:
            test_zip_path = source_dir / f"{code}.zip"
            uploaded_test_zip.save(test_zip_path)

        bundle = ProblemBundle(1, code, name, statement_path, generator_path if generator_path and generator_path.suffix.lower() == ".py" else None, test_zip_path, None, None)
        tests: GeneratedTests | None = None
        test_source = "KhÃ´ng cÃ³ test"
        log_lines = [f"ÄÃ£ chuáº©n bá»‹ bÃ i {code}: {name}."]
        if prepare_note:
            log_lines.append(f"- {prepare_note}")
        if test_zip_path:
            input_files, output_files = zip_case_files(test_zip_path)
            tests = GeneratedTests(test_zip_path, input_files, output_files)
            test_source = test_zip_path.name
            log_lines.append(f"- DÃ¹ng zip test cÃ³ sáºµn: {test_zip_path.name}, {len(input_files)} test.")
        elif bundle.generator:
            tests = generate_tests(bundle, build_root)
            test_source = tests.zip_path.name
            log_lines.append(f"- ÄÃ£ cháº¡y gentest Python vÃ  sinh {len(tests.input_files)} test: {tests.zip_path.name}.")
        elif generator_path and generator_path.suffix.lower() == ".cpp":
            tests = generate_tests_from_cpp_generator(generator_path, build_root, code)
            test_source = tests.zip_path.name
            log_lines.append(f"- ÄÃ£ compile/cháº¡y C++ generator vÃ  sinh {len(tests.input_files)} test: {tests.zip_path.name}.")

        solution_path: Path | None = None
        solution_text = (payload.get("solution_text") or "").strip()
        if solution_text:
            solution_path = source_dir / f"solution_{code}.md"
            solution_path.write_text(solution_text + "\n", encoding="utf-8")
            log_lines.append("- CÃ³ lá»i giáº£i/hÆ°á»›ng dáº«n Markdown.")

        rows = [
            {
                "original_code": code,
                "code": code,
                "name": name,
                "points": payload.get("points") or "100",
                "tags": payload.get("tags") or "",
                "time_limit": payload.get("time_limit") or "1.0",
                "memory_limit": payload.get("memory_limit") or "1024M",
                "partial": bool(payload.get("partial", True)),
                "test_file": test_source,
                "test_count": len(tests.input_files) if tests else 0,
                "upload_statement_default": bool(statement_text),
                "upload_tests_default": bool(tests),
                "upload_solution_default": bool(solution_path),
                "status": "ÄÃ£ chuáº©n bá»‹" if bool(statement_text) or bool(tests) or bool(solution_path) else "ChÆ°a cÃ³ pháº§n nÃ o Ä‘á»ƒ up",
                "note": prepare_note,
            }
        ]
        prepared_single_uploads[prepare_id] = {
            "root": root,
            "bundles": {code: bundle},
            "tests": {code: tests},
            "solutions": {code: solution_path},
        }
        progress_finish(progress_id, True, "ÄÃ£ chuáº©n bá»‹ 1/1 bÃ i")
        return jsonify({"prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


def infer_statement_title(statement: str) -> tuple[str, str]:
    for line in statement.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) >= 2 and parts[0] and re.fullmatch(r"[A-Za-z0-9_-]+", parts[1]):
            return parts[0], parts[1]
        break
    return "", ""


def repair_python_main_guard(text: str) -> str:
    return bundle_service.repair_python_main_guard(text)

def generate_tests_from_cpp_generator(generator_path: Path, build_root: Path, code: str) -> GeneratedTests:
    return bundle_service.generate_tests_from_cpp_generator(generator_path, build_root, code)

def find_generated_zip_for_single(build_dir: Path, bundle: ProblemBundle) -> Path | None:
    return bundle_service.find_generated_zip_for_single(build_dir, bundle)

def zip_generated_case_files(build_dir: Path, code: str) -> Path:
    return bundle_service.zip_generated_case_files(build_dir, code)

@app.post("/api/confirm-single-upload")
def api_confirm_single_upload():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id") or payload.get("settings", {}).get("progress_id")
    try:
        prepare_id = payload.get("prepare_id")
        if not prepare_id or prepare_id not in prepared_single_uploads:
            return jsonify({"ok": False, "error": "Dá»¯ liá»‡u Up 1 bÃ i Ä‘Ã£ háº¿t háº¡n. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i."}), 400
        settings = dict(payload.get("settings") or {})
        rows = payload.get("rows") or []
        target = settings.get("target") or "hncode"
        state = prepared_single_uploads[prepare_id]
        result_rows, log_lines = upload_rows(target, settings, rows, state, progress_id)
        append_single_solution_uploads(target, settings, result_rows, state, log_lines)
        ok = all((not row.get("selected")) or row["status"].startswith("âœ“") for row in result_rows)
        progress_finish(progress_id, ok, "ÄÃ£ hoÃ n táº¥t Up 1 bÃ i")
        return jsonify({"ok": ok, "rows": result_rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"ok": False, "error": str(exc)}), 400


def append_single_solution_uploads(target: str, settings: dict, rows: list[dict], state: dict, log_lines: list[str]) -> None:
    if not any(row.get("selected") and row.get("upload_solution") for row in rows):
        return
    target_info = TARGETS[target]
    session = login_upload_target(target, target_info, settings)
    for row in rows:
        if not row.get("selected") or not row.get("upload_solution") or not row.get("status", "").startswith("âœ“"):
            continue
        code = row.get("code") or row.get("original_code")
        solution_path = state.get("solutions", {}).get(row.get("original_code")) or state.get("solutions", {}).get(code)
        if not solution_path:
            continue
        try:
            update_problem_solution_markdown(session, target_info["base_url"], code, read_text_smart(solution_path))
            row["status"] += " vÃ  lá»i giáº£i"
            log_lines.append(f"{code}: Ä‘Ã£ up lá»i giáº£i/hÆ°á»›ng dáº«n Markdown.")
        except Exception as exc:
            row["status"] = "âœ— Lá»—i"
            row["error"] = str(exc)
            log_lines.append(f"âœ— {code}: khÃ´ng up lá»i giáº£i Ä‘Æ°á»£c: {exc}")


def update_problem_solution_markdown(session, base_url: str, code: str, content: str) -> str:
    solution_url = urljoin(base_url, f"/problem/{code}/edit/solutions")
    page = session.get(solution_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c trang lá»i giáº£i {code}: HTTP {page.status_code}")
    parser = FormDataParser()
    parser.feed(page.text)
    form = next((form for form in parser.forms if any(name == "content" for name, _value in form)), None)
    if not form:
        raise RuntimeError("KhÃ´ng tÃ¬m tháº¥y form lá»i giáº£i cÃ³ trÆ°á»ng content.")
    data = set_single_form_fields(form, {"content": content})
    result = session.post(solution_url, data=data, headers={"Referer": solution_url}, allow_redirects=True, timeout=30)
    if not result.ok:
        raise RuntimeError(f"Up lá»i giáº£i lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form lá»i giáº£i bÃ¡o lá»—i:\n" + "\n".join(errors))
    return result.url


@app.post("/api/confirm-upload")
def api_confirm_upload():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id") or payload.get("settings", {}).get("progress_id")
    try:
        prepare_id = payload.get("prepare_id")
        if not prepare_id or prepare_id not in prepared_uploads:
            return jsonify(
                {
                    "ok": False,
                    "error": "Dá»¯ liá»‡u chuáº©n bá»‹ Ä‘Ã£ háº¿t háº¡n hoáº·c server vá»«a khá»Ÿi Ä‘á»™ng láº¡i. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i rá»“i má»›i XÃ¡c nháº­n Up bÃ i.",
                }
            ), 400
        state = prepared_uploads[prepare_id]
        target = payload["settings"]["target"]
        result_rows, log_lines = upload_rows(target, payload["settings"], payload["rows"], state, progress_id)
        append_single_solution_uploads(target, payload["settings"], result_rows, state, log_lines)
        ok = all((not row.get("selected")) or row["status"].startswith("âœ“") for row in result_rows)
        progress_finish(progress_id, ok, "ÄÃ£ hoÃ n táº¥t up bÃ i")
        return jsonify({"ok": ok, "rows": result_rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"ok": False, "error": str(exc)}), 400


def upload_rows(target: str, settings: dict, rows: list[dict], state: dict, progress_id: str | None = None) -> tuple[list[dict], list[str]]:
    target_info = TARGETS[target]
    log_lines = [f"ÄÃ­ch: {target_info['label']}", "Táº¡o bÃ i qua admin form: /admin/judge/problem/add/"]
    selected_language_ids = language_ids_for_target(target, settings.get("languages", []))
    if not selected_language_ids:
        log_lines.append("NgÃ´n ngá»¯ cho phÃ©p: form/admin hiá»‡n táº¡i khÃ´ng cÃ³ ID tÆ°Æ¡ng á»©ng, backend bá» qua an toÃ n.")
    if settings.get("creator"):
        log_lines.append("Creators Ä‘Æ°á»£c hiá»ƒn thá»‹ trÃªn giao diá»‡n; backend chá»‰ set náº¿u form admin há»— trá»£ trá»±c tiáº¿p.")

    session = login_upload_target(target, target_info, settings)
    result_rows = []
    total = len([row for row in rows if row.get("selected")])
    done = 0
    progress_update(progress_id, phase="confirm-upload", done=done, total=total, rows=result_rows, message="Báº¯t Ä‘áº§u up bÃ i")
    for row in rows:
        row = dict(row)
        if not row.get("selected"):
            row["status"] = "Bá» qua"
            result_rows.append(row)
            continue
        try:
            raw_code = row["code"] or row["original_code"]
            dest_code, code_note = resolve_problem_code_for_upload(session, target, target_info["base_url"], raw_code)
            if dest_code != raw_code:
                row["code"] = dest_code
                log_lines.append(code_note or f"{raw_code}: mÃ£ Ä‘Ã­ch {TARGETS[target]['label']} Ä‘Æ°á»£c Ä‘á»•i thÃ nh {dest_code}")
            bundle = replace(state["bundles"][row["original_code"]], code=dest_code, name=row["name"])
            tests = state["tests"].get(row["original_code"])
            action_status = upload_one_problem(session, target, target_info, bundle, tests, row, settings, selected_language_ids, log_lines)
            row["status"] = action_status or "âœ“ ThÃ nh cÃ´ng"
            row["link"] = problem_url(target_info["base_url"], bundle.code)
        except ProblemAlreadyExists as exc:
            row["status"] = "âœ— BÃ i Ä‘Ã£ tá»“n táº¡i"
            log_lines.append(f"âœ— {row.get('code')}: {exc}. Bá» qua bÃ i nÃ y vÃ  tiáº¿p tá»¥c cÃ¡c bÃ i khÃ¡c.")
        except Exception as exc:
            row["status"] = "âœ— Lá»—i"
            log_lines.append(f"âœ— {row.get('code')}: {exc}")
        result_rows.append(row)
        done += 1
        progress_update(progress_id, phase="confirm-upload", done=done, total=total, rows=result_rows, message=f"{row.get('code')}: {row.get('status')}")
    return result_rows, log_lines


def login_upload_target(target: str, target_info: dict, settings: dict):
    try:
        return login_hncode(target_info["base_url"], settings.get("username", ""), settings.get("password", ""))
    except Exception as exc:
        label = target_info.get("label", target)
        message = str(exc).replace("HNCode", label)
        raise RuntimeError(message)


def is_problem_add_form(page: str) -> bool:
    return bool(
        re.search(r"<input\b[^>]*name=[\"']code[\"']", page, re.S)
        and re.search(r"<textarea\b[^>]*name=[\"']description[\"']", page, re.S)
    )


def tinhoctre_admin_cookie_error(final_url: str = "") -> str:
    suffix = f" URL hiá»‡n táº¡i: {final_url}" if final_url else ""
    return (
        "Cookie TinHocTre chÆ°a vÃ o Ä‘Æ°á»£c form admin táº¡o bÃ i. "
        "CÃ³ thá»ƒ báº¡n copy cookie khi chÆ°a Ä‘Äƒng nháº­p admin, cookie Ä‘Ã£ háº¿t háº¡n, hoáº·c tÃ i khoáº£n khÃ´ng cÃ³ quyá»n staff/admin. "
        "HÃ£y má»Ÿ https://tinhoctre.vn/admin/judge/problem/add/ trÃªn cÃ¹ng trÃ¬nh duyá»‡t, Ä‘áº£m báº£o tháº¥y form táº¡o bÃ i, "
        "rá»“i copy láº¡i Request Header Cookie vÃ  dÃ¡n vÃ o tab TÃ i khoáº£n."
        + suffix
    )


def collect_problem_edit_form_data(page: str) -> list[tuple[str, str]]:
    parser = FormDataParser()
    parser.feed(page)
    for form in parser.forms:
        names = {name for name, _value in form}
        if "code" in names and "description" in names:
            return form
    return []


def set_single_form_fields(data: list[tuple[str, str]], updates: dict[str, str]) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for name, value in data:
        if name in updates:
            if name not in seen:
                out.append((name, updates[name]))
                seen.add(name)
            continue
        out.append((name, value))
    for name, value in updates.items():
        if name not in seen:
            out.append((name, value))
    return out


def set_form_fields(
    data: list[tuple[str, str]],
    updates: dict[str, str],
    *,
    multi_updates: dict[str, list[str]] | None = None,
    checkbox_updates: dict[str, bool] | None = None,
) -> list[tuple[str, str]]:
    multi_updates = multi_updates or {}
    checkbox_updates = checkbox_updates or {}
    skip_names = set(updates) | set(multi_updates) | set(checkbox_updates)
    out: list[tuple[str, str]] = []
    for name, value in data:
        if name not in skip_names:
            out.append((name, value))
    out.extend(updates.items())
    for name, values in multi_updates.items():
        out.extend((name, value) for value in values if value)
    for name, checked in checkbox_updates.items():
        if checked:
            out.append((name, "on"))
    return out


def normalized_lookup_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9_]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def select_options_with_text(page: str, name: str) -> list[dict[str, str | bool]]:
    match = re.search(r"<select\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", page, re.S)
    if not match:
        return []
    options: list[dict[str, str | bool]] = []
    for option in re.finditer(r"<option\b([^>]*)>(.*?)</option>", match.group(1), re.S):
        attrs = option.group(1)
        value = re.search(r"value=[\"']([^\"']*)", attrs)
        label = html.unescape(re.sub(r"<.*?>", " ", option.group(2))).strip()
        label = re.sub(r"\s+", " ", label)
        options.append(
            {
                "value": html.unescape(value.group(1)) if value else "",
                "text": label,
                "selected": "selected" in attrs,
            }
        )
    return options


def resolve_hncode_type_ids(page: str, tags_text: object, fallback_ids: list[str], default_type_id: str) -> list[str]:
    options = select_options_with_text(page, "types")
    valid_values = {str(option["value"]) for option in options if option.get("value")}
    ids: list[str] = []

    def add(value: str) -> None:
        if value and value in valid_values and value not in ids:
            ids.append(value)

    lookup: dict[str, str] = {}
    for option in options:
        value = str(option.get("value") or "")
        text = str(option.get("text") or "")
        if not value:
            continue
        labels = [text]
        if " - " in text:
            labels.extend(part.strip() for part in text.split(" - ", 1))
        if "-" in text:
            labels.extend(part.strip() for part in text.split("-", 1))
        for label in labels:
            key = normalized_lookup_text(label)
            if key:
                lookup.setdefault(key, value)

    for raw_tag in re.split(r"[,;|]+", str(tags_text or "")):
        tag = raw_tag.strip()
        if not tag:
            continue
        if tag in valid_values:
            add(tag)
            continue
        alias_id = HNCODE_TYPE_ALIASES.get(tag.lower()) or HNCODE_TYPE_ALIASES.get(normalized_lookup_text(tag))
        if alias_id:
            add(alias_id)
            continue
        add(lookup.get(normalized_lookup_text(tag), ""))

    if not ids:
        for value in fallback_ids or []:
            add(str(value))
    if not ids and default_type_id:
        add(str(default_type_id))
    return ids or ([str(default_type_id)] if default_type_id else [])


def update_existing_problem_statement(
    session,
    target: str,
    base_url: str,
    bundle: ProblemBundle,
    settings: dict,
) -> str:
    edit_url = urljoin(base_url, f"/problem/{bundle.code}/edit")
    page = session.get(edit_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a bÃ i {bundle.code}: HTTP {page.status_code}")
    data = collect_problem_edit_form_data(page.text)
    if not data:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y form sá»­a Ä‘á» bÃ i cho {bundle.code}. TÃ i khoáº£n cÃ³ thá»ƒ chÆ°a cÃ³ quyá»n sá»­a bÃ i nÃ y.")
    description = statement_for_target(
        target,
        read_text_smart(bundle.statement),
        skip_title_line=bool(settings.get("skip_statement_title", True)),
    )
    data = set_single_form_fields(
        data,
        {
            "code": bundle.code,
            "name": bundle.name,
            "description": description,
        },
    )
    result = session.post(edit_url, data=data, headers={"Referer": edit_url}, allow_redirects=True, timeout=30)
    if not result.ok:
        raise RuntimeError(f"Ghi Ä‘Ã¨ Ä‘á» bÃ i lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form ghi Ä‘Ã¨ Ä‘á» bÃ i bÃ¡o lá»—i:\n" + "\n".join(errors))
    if "/accounts/login" in result.url or "/admin/login" in result.url:
        raise RuntimeError(f"Ghi Ä‘Ã¨ Ä‘á» bÃ i bá»‹ chuyá»ƒn vá» trang Ä‘Äƒng nháº­p: {result.url}")
    verify = session.get(edit_url, timeout=30)
    if not verify.ok:
        raise RuntimeError(f"KhÃ´ng kiá»ƒm tra láº¡i Ä‘Æ°á»£c Ä‘á» bÃ i sau khi ghi Ä‘Ã¨: HTTP {verify.status_code}")
    saved_description = textarea_value(verify.text, "description")
    saved_name = input_value_from_page(verify.text, "name", "")
    if saved_description.strip() != description.strip():
        debug_dir = RUNTIME / "debug_overwrite_statement"
        debug_dir.mkdir(parents=True, exist_ok=True)
        (debug_dir / f"{bundle.code}_expected.md").write_text(description, encoding="utf-8", errors="replace")
        (debug_dir / f"{bundle.code}_saved.md").write_text(saved_description, encoding="utf-8", errors="replace")
        (debug_dir / f"{bundle.code}_post.html").write_text(result.text, encoding="utf-8", errors="replace")
        (debug_dir / f"{bundle.code}_verify.html").write_text(verify.text, encoding="utf-8", errors="replace")
        raise RuntimeError(
            f"HNCode nháº­n POST nhÆ°ng Ä‘á» bÃ i {bundle.code} chÆ°a khá»›p ná»™i dung má»›i. "
            f"ÄÃ£ lÆ°u debug táº¡i {debug_dir}."
        )
    if saved_name and saved_name != bundle.name:
        raise RuntimeError(f"HNCode nháº­n POST nhÆ°ng tÃªn bÃ i {bundle.code} chÆ°a Ä‘á»•i: {saved_name!r}")
    return result.url


def update_hncode_problem_metadata(
    session,
    base_url: str,
    code: str,
    *,
    target_label: str = "HNCode",
    name: str,
    points: str,
    partial: bool,
    time_limit: str,
    memory_limit: str,
    type_ids: list[str],
    group_id: str,
    default_type_id: str = "",
    tags_text: object = "",
) -> str:
    edit_url = urljoin(base_url, f"/problem/{code}/edit")
    page = session.get(edit_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form metadata HNCode {code}: HTTP {page.status_code}")
    data = collect_problem_edit_form_data(page.text)
    if not data:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y form metadata HNCode cho {code}.")
    resolved_type_ids = resolve_hncode_type_ids(page.text, tags_text, type_ids or [default_type_id], default_type_id)
    data = set_form_fields(
        data,
        {
            "code": code,
            "name": name,
            "points": str(points or "100"),
            "time_limit": str(time_limit or "1.0"),
            "memory_limit": memory_limit_to_kb(memory_limit or "1048576"),
            "memory_unit": "KB",
            "group": group_id,
        },
        multi_updates={"types": resolved_type_ids},
        checkbox_updates={"partial": bool(partial)},
    )
    result = session.post(edit_url, data=data, headers={"Referer": edit_url}, allow_redirects=True, timeout=30)
    if not result.ok:
        raise RuntimeError(f"Cáº­p nháº­t metadata HNCode {code} lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError(f"Form metadata HNCode {code} bÃ¡o lá»—i:\n" + "\n".join(errors))
    if "/accounts/login" in result.url or "/admin/login" in result.url:
        raise RuntimeError(f"Cáº­p nháº­t metadata HNCode {code} bá»‹ chuyá»ƒn vá» trang Ä‘Äƒng nháº­p: {result.url}")
    verify = session.get(edit_url, timeout=30)
    if not verify.ok:
        raise RuntimeError(f"KhÃ´ng kiá»ƒm tra láº¡i metadata HNCode {code}: HTTP {verify.status_code}")
    saved_points = input_value_from_page(verify.text, "points", "")
    saved_type_ids = selected_values(verify.text, "types")
    if not same_numeric_value(saved_points, str(points or "100")):
        debug_dir = RUNTIME / "debug_hncode_metadata"
        debug_dir.mkdir(parents=True, exist_ok=True)
        (debug_dir / f"{code}_post.html").write_text(result.text, encoding="utf-8", errors="replace")
        (debug_dir / f"{code}_verify.html").write_text(verify.text, encoding="utf-8", errors="replace")
        raise RuntimeError(
            f"HNCode nháº­n POST nhÆ°ng Points cá»§a {code} váº«n lÃ  {saved_points!r}, "
            f"khÃ´ng pháº£i {points!r}. ÄÃ£ lÆ°u debug táº¡i {debug_dir}."
        )
    missing_type_ids = [value for value in resolved_type_ids if value not in saved_type_ids]
    if missing_type_ids:
        debug_dir = RUNTIME / "debug_hncode_metadata"
        debug_dir.mkdir(parents=True, exist_ok=True)
        (debug_dir / f"{code}_post.html").write_text(result.text, encoding="utf-8", errors="replace")
        (debug_dir / f"{code}_verify.html").write_text(verify.text, encoding="utf-8", errors="replace")
        raise RuntimeError(
            f"HNCode nháº­n POST nhÆ°ng Problem Types cá»§a {code} váº«n lÃ  {saved_type_ids}, "
            f"chÆ°a cÃ³ {missing_type_ids}. ÄÃ£ lÆ°u debug táº¡i {debug_dir}."
        )
    saved_description = textarea_value(verify.text, "description")
    ensure_hncode_vi_translation(session, base_url, code, name, saved_description)
    return result.url


def find_hncode_admin_problem_change_url(session, base_url: str, code: str) -> str:
    search_url = urljoin(base_url, f"/admin/judge/problem/?q={quote(code)}")
    page = session.get(search_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c trang admin tÃ¬m bÃ i {code}: HTTP {page.status_code}")
    match = re.search(r"/admin/judge/problem/(\d+)/change/", page.text)
    if not match:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y admin change URL cho bÃ i {code}.")
    return urljoin(base_url, f"/admin/judge/problem/{match.group(1)}/change/")


def ensure_hncode_vi_translation(session, base_url: str, code: str, name: str, description: str) -> None:
    change_url = find_hncode_admin_problem_change_url(session, base_url, code)
    page = session.get(change_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c admin form bÃ i {code}: HTTP {page.status_code}")
    data = collect_problem_edit_form_data(page.text)
    if not data:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c admin form Ä‘á»ƒ cáº­p nháº­t báº£n dá»‹ch tiáº¿ng Viá»‡t cho {code}.")
    object_id = re.search(r"/admin/judge/problem/(\d+)/change/", change_url)
    problem_id = object_id.group(1) if object_id else ""
    data = [(key, value) for key, value in data if "__prefix__" not in key]
    values = dict(data)
    total = int(values.get("translations-TOTAL_FORMS") or "0")
    target_index: int | None = None
    for index in range(total):
        if values.get(f"translations-{index}-language") == "vi":
            target_index = index
            break
    if target_index is None:
        target_index = total
        total += 1
    updates = {
        "translations-TOTAL_FORMS": str(total),
        f"translations-{target_index}-language": "vi",
        f"translations-{target_index}-name": name,
        f"translations-{target_index}-description": description,
        f"translations-{target_index}-id": values.get(f"translations-{target_index}-id", ""),
        f"translations-{target_index}-problem": values.get(f"translations-{target_index}-problem", problem_id),
    }
    data = set_single_form_fields(data, updates)
    data.append(("_save", "LÆ°u"))
    result = session.post(change_url, data=data, headers={"Referer": change_url}, allow_redirects=True, timeout=30)
    if not result.ok:
        raise RuntimeError(f"Cáº­p nháº­t báº£n dá»‹ch tiáº¿ng Viá»‡t cho {code} lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError(f"Form báº£n dá»‹ch tiáº¿ng Viá»‡t HNCode {code} bÃ¡o lá»—i:\n" + "\n".join(errors))


def same_numeric_value(left: str, right: str) -> bool:
    try:
        return abs(float(str(left).strip()) - float(str(right).strip())) < 1e-9
    except (TypeError, ValueError):
        return str(left).strip() == str(right).strip()


def upload_one_problem(
    session,
    target: str,
    target_info: dict,
    bundle: ProblemBundle,
    tests: GeneratedTests | None,
    row: dict,
    settings: dict,
    language_ids: list[str],
    log_lines: list[str],
) -> str:
    base_url = target_info["base_url"]

    def refresh_hncode_metadata() -> None:
        if target not in {"hncode", "tinhoctre"}:
            return
        tags_text = row.get("tags") or settings.get("tags")
        type_ids = type_ids_from_tags(tags_text, target) or [target_info["type_id"]]
        update_hncode_problem_metadata(
            session,
            base_url,
            bundle.code,
            target_label=target_info["label"],
            name=row.get("name") or bundle.name,
            points=str(row.get("points") or settings.get("points") or "100"),
            partial=bool(row.get("partial", settings.get("partial", True))),
            time_limit=row.get("time_limit") or settings.get("time_limit") or "1.0",
            memory_limit=row.get("memory_limit") or settings.get("memory_limit") or "1048576",
            type_ids=type_ids,
            group_id=target_info["group_id"],
            default_type_id=target_info["type_id"],
            tags_text=tags_text,
        )
        log_lines.append(f"{bundle.code}: Ä‘Ã£ cáº­p nháº­t láº¡i Ä‘iá»ƒm vÃ  dáº¡ng bÃ i táº­p {target_info['label']}.")

    exists = problem_exists_for_target(session, target, base_url, bundle.code)
    if exists:
        overwrite_row = bool(row.get("overwrite") or settings.get("overwrite_existing"))
        overwrite_statement = bool(settings.get("overwrite_statement") or overwrite_row) and bool(row.get("upload_statement"))
        overwrite_tests = bool(settings.get("overwrite_tests") or overwrite_row) and bool(row.get("upload_tests"))
        if not (overwrite_statement or overwrite_tests):
            raise ProblemAlreadyExists(f"MÃ£ bÃ i {bundle.code} Ä‘Ã£ tá»“n táº¡i táº¡i {problem_url(base_url, bundle.code)}")
        log_lines.append(f"{bundle.code}: bÃ i Ä‘Ã£ tá»“n táº¡i, chuyá»ƒn sang cháº¿ Ä‘á»™ ghi Ä‘Ã¨ pháº§n Ä‘Æ°á»£c chá»n.")
        actions: list[str] = []
        if row.get("upload_statement"):
            if overwrite_statement:
                change_url = update_existing_problem_statement(session, target, base_url, bundle, settings)
                log_lines.append(f"{bundle.code}: Ä‘Ã£ ghi Ä‘Ã¨ Ä‘á» bÃ i ({change_url}).")
                actions.append("Ä‘á» bÃ i")
            else:
                log_lines.append(f"{bundle.code}: khÃ´ng ghi Ä‘Ã¨ Ä‘á» bÃ i vÃ¬ chÆ°a tÃ­ch Ghi Ä‘Ã¨ Ä‘á» bÃ i.")
        if row.get("upload_tests"):
            if overwrite_tests:
                if tests is None:
                    raise RuntimeError("BÃ i nÃ y khÃ´ng cÃ³ bá»™ test trong dá»¯ liá»‡u chuáº©n bá»‹. HÃ£y bá» tÃ­ch Up test hoáº·c dÃ¹ng file zip/gentest.")
                upload_tests_for_target(session, target, base_url, bundle.code, tests)
                log_lines.append(f"{bundle.code}: Ä‘Ã£ ghi Ä‘Ã¨ {len(tests.input_files)} test.")
                actions.append("test")
            else:
                log_lines.append(f"{bundle.code}: khÃ´ng ghi Ä‘Ã¨ test vÃ¬ chÆ°a tÃ­ch Ghi Ä‘Ã¨ test.")
        submit_if_requested(session, base_url, bundle, settings, log_lines)
        if overwrite_statement or overwrite_tests or overwrite_row:
            refresh_hncode_metadata()
        return "âœ“ Ghi Ä‘Ã¨ " + " vÃ  ".join(actions) if actions else "âœ“ KhÃ´ng cÃ³ pháº§n ghi Ä‘Ã¨"
    actions: list[str] = []
    if row.get("upload_statement"):
        type_ids = type_ids_from_tags(row.get("tags") or settings.get("tags"), target) or [target_info["type_id"]]
        type_id = type_ids[0] if type_ids else target_info["type_id"]
        info = ProblemInfo(
            code=bundle.code,
            name=bundle.name,
            description=statement_for_target(
                target,
                read_text_smart(bundle.statement),
                skip_title_line=bool(settings.get("skip_statement_title", True)),
            ),
            points=str(row.get("points") or settings.get("points") or "100"),
            partial=bool(row.get("partial", settings.get("partial", True))),
            time_limit=row.get("time_limit") or settings.get("time_limit") or "1.0",
            memory_limit=memory_limit_to_kb(row.get("memory_limit") or settings.get("memory_limit") or "1048576"),
            memory_unit="KB",
        )
        change_url = create_hncode_problem(
            session,
            base_url,
            info,
            dest_code=bundle.code,
            type_id=",".join(type_ids),
            group_id=target_info["group_id"],
            default_type_id=target_info["type_id"],
            default_group_id=target_info["group_id"],
            public=False,
            allow_all_languages=False,
            allowed_language_ids=language_ids,
        )
        log_lines.append(f"{bundle.code}: Ä‘Ã£ táº¡o Ä‘á» qua admin form ({change_url}).")
        actions.append("táº¡o Ä‘á»")
    else:
        log_lines.append(f"{bundle.code}: khÃ´ng upload Ä‘á».")

    if row.get("upload_tests"):
        if tests is None:
            raise RuntimeError("BÃ i nÃ y khÃ´ng cÃ³ bá»™ test trong dá»¯ liá»‡u chuáº©n bá»‹. HÃ£y bá» tÃ­ch Up test hoáº·c dÃ¹ng file zip/gentest.")
        upload_tests_for_target(session, target, base_url, bundle.code, tests)
        log_lines.append(f"{bundle.code}: Ä‘Ã£ upload {len(tests.input_files)} test.")
        actions.append("upload test")
    else:
        log_lines.append(f"{bundle.code}: khÃ´ng upload test.")

    submit_if_requested(session, base_url, bundle, settings, log_lines)
    if target in {"hncode", "tinhoctre"} and row.get("upload_statement"):
        refresh_hncode_metadata()
    return "âœ“ " + " vÃ  ".join(actions) if actions else "âœ“ ThÃ nh cÃ´ng"


def problem_exists_for_target(session, target: str, base_url: str, code: str) -> bool:
    return upload_service.problem_exists_for_target(session, target, base_url, code)

def resolve_problem_code_for_upload(session, target: str, base_url: str, raw_code: str) -> tuple[str, str]:
    return upload_service.resolve_problem_code_for_upload(session, target, base_url, raw_code)

def statement_for_target(target: str, statement: str, *, skip_title_line: bool = False) -> str:
    return upload_service.statement_for_target(target, statement, skip_title_line=skip_title_line)

def problem_info_for_target(info: ProblemInfo, target: str) -> ProblemInfo:
    return replace(info, description=statement_for_target(target, info.description))


def create_tinhoctre_admin_problem(
    session,
    base_url: str,
    info: ProblemInfo,
    *,
    dest_code: str,
    type_id: str,
    group_id: str,
    allowed_language_ids: list[str],
) -> str:
    add_url = urljoin(base_url, "/admin/judge/problem/add/")
    page = session.get(add_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"TinHocTre add page failed: HTTP {page.status_code}")
    if not is_problem_add_form(page.text):
        raise RuntimeError(tinhoctre_admin_cookie_error(page.url))
    token = csrf_token(page.text)
    language_ids = [value for value in allowed_language_ids if value]
    data: list[tuple[str, str]] = [
        ("csrfmiddlewaretoken", token),
        ("code", dest_code),
        ("name", info.name),
        ("submission_source_visibility_mode", selected_option_value(page.text, "submission_source_visibility_mode", "F")),
        ("testcase_visibility_mode", selected_option_value(page.text, "testcase_visibility_mode", "O")),
        ("testcase_result_visibility_mode", selected_option_value(page.text, "testcase_result_visibility_mode", "A")),
        ("description", info.description),
        ("pdf_url", ""),
        ("source", ""),
        ("license", selected_option_value(page.text, "license", "")),
        ("og_image", ""),
        ("summary", ""),
        ("types", type_id),
        ("group", group_id),
        ("points", info.points),
        ("time_limit", info.time_limit),
        ("memory_limit", info.memory_limit),
        ("change_message", ""),
        ("language_limits-TOTAL_FORMS", input_value_from_page(page.text, "language_limits-TOTAL_FORMS", "3")),
        ("language_limits-INITIAL_FORMS", input_value_from_page(page.text, "language_limits-INITIAL_FORMS", "0")),
        ("language_limits-MIN_NUM_FORMS", input_value_from_page(page.text, "language_limits-MIN_NUM_FORMS", "0")),
        ("language_limits-MAX_NUM_FORMS", input_value_from_page(page.text, "language_limits-MAX_NUM_FORMS", "1000")),
        ("problemclarification_set-TOTAL_FORMS", input_value_from_page(page.text, "problemclarification_set-TOTAL_FORMS", "0")),
        ("problemclarification_set-INITIAL_FORMS", input_value_from_page(page.text, "problemclarification_set-INITIAL_FORMS", "0")),
        ("problemclarification_set-MIN_NUM_FORMS", input_value_from_page(page.text, "problemclarification_set-MIN_NUM_FORMS", "0")),
        ("problemclarification_set-MAX_NUM_FORMS", input_value_from_page(page.text, "problemclarification_set-MAX_NUM_FORMS", "1000")),
        ("solution-TOTAL_FORMS", input_value_from_page(page.text, "solution-TOTAL_FORMS", "0")),
        ("solution-INITIAL_FORMS", input_value_from_page(page.text, "solution-INITIAL_FORMS", "0")),
        ("solution-MIN_NUM_FORMS", input_value_from_page(page.text, "solution-MIN_NUM_FORMS", "0")),
        ("solution-MAX_NUM_FORMS", input_value_from_page(page.text, "solution-MAX_NUM_FORMS", "1")),
        ("translations-TOTAL_FORMS", input_value_from_page(page.text, "translations-TOTAL_FORMS", "0")),
        ("translations-INITIAL_FORMS", input_value_from_page(page.text, "translations-INITIAL_FORMS", "0")),
        ("translations-MIN_NUM_FORMS", input_value_from_page(page.text, "translations-MIN_NUM_FORMS", "0")),
        ("translations-MAX_NUM_FORMS", input_value_from_page(page.text, "translations-MAX_NUM_FORMS", "1000")),
        ("_continue", "Save and continue editing"),
    ]
    if input_checked(page.text, "allow_judging"):
        data.append(("allow_judging", "on"))
    if info.partial:
        data.append(("partial", "on"))
    for value in language_ids:
        data.append(("allowed_languages", value))

    total_language_limits = int(input_value_from_page(page.text, "language_limits-TOTAL_FORMS", "3") or "0")
    for index in range(total_language_limits):
        data.extend(
            [
                (f"language_limits-{index}-id", input_value_from_page(page.text, f"language_limits-{index}-id", "")),
                (f"language_limits-{index}-problem", input_value_from_page(page.text, f"language_limits-{index}-problem", "")),
                (f"language_limits-{index}-language", selected_option_value(page.text, f"language_limits-{index}-language", "")),
                (f"language_limits-{index}-time_limit", input_value_from_page(page.text, f"language_limits-{index}-time_limit", "")),
                (f"language_limits-{index}-memory_limit", input_value_from_page(page.text, f"language_limits-{index}-memory_limit", "")),
            ]
        )

    result = session.post(add_url, data=data, headers={"Referer": add_url}, allow_redirects=True, timeout=30)
    if "/admin/login/" in result.url:
        raise RuntimeError(tinhoctre_admin_cookie_error(result.url))
    if not result.ok:
        errors = form_errors(result.text)
        detail = ("\n" + "\n".join(errors)) if errors else ""
        raise RuntimeError(f"TinHocTre create problem failed: HTTP {result.status_code}{detail}")
    errors = form_errors(result.text)
    if errors:
        raise RuntimeError("TinHocTre create problem form errors:\n" + "\n".join(errors))
    if "/change/" not in result.url and dest_code not in result.text:
        raise RuntimeError(f"TinHocTre did not appear to save {dest_code}; final URL: {result.url}")
    return result.url


def input_value_from_page(page: str, name: str, default: str = "") -> str:
    match = re.search(r"<input\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>", page, re.S)
    if not match:
        return default
    value = re.search(r"value=[\"']([^\"']*)", match.group(0))
    return html.unescape(value.group(1)) if value else default


def input_checked(page: str, name: str) -> bool:
    match = re.search(r"<input\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>", page, re.S)
    return bool(match and re.search(r"\bchecked\b", match.group(0)))


def selected_option_value(page: str, name: str, default: str = "") -> str:
    match = re.search(r"<select\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", page, re.S)
    if not match:
        return default
    options = list(re.finditer(r"<option\b([^>]*)>(.*?)</option>", match.group(1), re.S))
    for option in options:
        attrs = option.group(1)
        if "selected" in attrs:
            value = re.search(r"value=[\"']([^\"']*)", attrs)
            return html.unescape(value.group(1)) if value else default
    if options:
        value = re.search(r"value=[\"']([^\"']*)", options[0].group(1))
        return html.unescape(value.group(1)) if value else default
    return default


def upload_tests_for_target(session, target: str, base_url: str, code: str, tests: GeneratedTests) -> None:
    upload_service.upload_tests_for_target(session, target, base_url, code, tests, upload_hncode_tests, upload_tinhoctre_tests)

def submit_if_requested(session, base_url: str, bundle: ProblemBundle, settings: dict, log_lines: list[str]) -> None:
    fallback = None if ("hncode.edu.vn" in base_url or "tinhoctre.vn" in base_url) else submit_solution
    upload_service.submit_if_requested(session, base_url, bundle, settings, log_lines, compact_form_red_errors, fallback_submit_solution=fallback)

def submit_solution_file(session, base_url: str, code: str, source_path: Path, preferred_languages: list[str]) -> str:
    return upload_service.submit_solution_file(session, base_url, code, source_path, preferred_languages, compact_form_red_errors)

def language_id_from_submit_page(page: str, preferred_languages: list[str]) -> str:
    return upload_service.language_id_from_submit_page(page, preferred_languages)

def normalize_language_label(label: str) -> str:
    return upload_service.normalize_language_label(label)

def language_ids_for_target(target: str, names: list[str]) -> list[str]:
    return upload_service.language_ids_for_target(TARGETS[target], names)

def memory_limit_to_kb(value: object) -> str:
    return upload_service.memory_limit_to_kb(value)

def type_id_from_tags(value: object) -> str:
    match = re.search(r"\b\d+\b", str(value or ""))
    return match.group(0) if match else ""


def type_ids_from_tags(value: object, target: str) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    numeric_ids = re.findall(r"\b\d+\b", text)
    if numeric_ids:
        return list(dict.fromkeys(numeric_ids))
    if target != "hncode":
        return []
    ids: list[str] = []
    for raw_tag in re.split(r"[,;|]+", text):
        tag = re.sub(r"\s+", " ", raw_tag.strip().lower())
        if not tag:
            continue
        type_id = HNCODE_TYPE_ALIASES.get(tag)
        if type_id and type_id not in ids:
            ids.append(type_id)
    return ids


def problem_url(base_url: str, code: str) -> str:
    return upload_service.problem_url(base_url, code)

def normalize_problem_code_for_target(code: str, target: str) -> str:
    return upload_service.normalize_problem_code_for_target(code, target)

def validate_problem_code_for_target(code: str, target: str) -> None:
    upload_service.validate_problem_code_for_target(code, target)

def test_data_url(base_url: str, code: str) -> str:
    return upload_service.test_data_url(base_url, code)

def session_from_cookie(cookie_header: str):
    s = tinhoctre_session()
    parsed = SimpleCookie()
    parsed.load(cookie_header)
    for key, morsel in parsed.items():
        s.cookies.set(key, morsel.value, domain=".tinhoctre.vn")
        s.cookies.set(key, morsel.value, domain="tinhoctre.vn")
    return s


def tinhoctre_cookie_file() -> Path:
    return RUNTIME / "tinhoctre_cookie.txt"


def save_tinhoctre_cookie(cookie_header: str) -> None:
    RUNTIME.mkdir(parents=True, exist_ok=True)
    tinhoctre_cookie_file().write_text(cookie_header, encoding="utf-8")


def load_tinhoctre_cookie() -> str:
    path = tinhoctre_cookie_file()
    if path.exists():
        return path.read_text(encoding="utf-8").strip()
    return ""


def find_edge_executable() -> Path:
    candidates = [
        shutil.which("msedge"),
        shutil.which("msedge.exe"),
        os.environ.get("EDGE_PATH"),
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        shutil.which("chrome"),
        shutil.which("chrome.exe"),
        os.environ.get("CHROME_PATH"),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return Path(candidate)
    raise RuntimeError("KhÃ´ng tÃ¬m tháº¥y Edge/Chrome trÃªn mÃ¡y local. HÃ£y cÃ i Edge hoáº·c Ä‘áº·t biáº¿n mÃ´i trÆ°á»ng EDGE_PATH.")


def stop_edge_processes() -> None:
    if os.name == "nt":
        subprocess.run(["taskkill", "/IM", "msedge.exe", "/F"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return
    subprocess.run(["pkill", "-f", "msedge"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def cdp_json(path: str, port: int) -> dict:
    with urlopen(f"http://127.0.0.1:{port}{path}", timeout=5) as response:
        return json.loads(response.read().decode("utf-8"))


def cookie_from_tinhoctre_debug_browser() -> str:
    try:
        import websocket
    except Exception as exc:
        raise RuntimeError("Thiáº¿u thÆ° viá»‡n websocket-client Ä‘á»ƒ Ä‘á»c cookie Edge. HÃ£y cÃ i: pip install websocket-client") from exc

    port = int(os.getenv("TINHOCTRE_CHROME_DEBUG_PORT", "9223"))
    deadline = time.time() + 20
    last_error: Exception | None = None
    version = None
    while time.time() < deadline:
        try:
            version = cdp_json("/json/version", port)
            break
        except Exception as exc:
            last_error = exc
            time.sleep(0.5)
    if not version:
        raise RuntimeError(f"KhÃ´ng káº¿t ná»‘i Ä‘Æ°á»£c Edge Ä‘Äƒng nháº­p TinHocTre á»Ÿ cá»•ng {port}. HÃ£y Ä‘Ã³ng háº¿t Edge, rá»“i báº¥m Má»Ÿ Edge Ä‘Äƒng nháº­p TinHocTre trÆ°á»›c.") from last_error

    ws_url = version.get("webSocketDebuggerUrl")
    if not ws_url:
        raise RuntimeError("Edge DevTools khÃ´ng tráº£ webSocketDebuggerUrl.")
    ws = websocket.create_connection(ws_url, timeout=10)
    counter = 0

    def cdp(method: str, params: dict | None = None):
        nonlocal counter
        counter += 1
        ws.send(json.dumps({"id": counter, "method": method, "params": params or {}}))
        while True:
            message = json.loads(ws.recv())
            if message.get("id") == counter:
                if "error" in message:
                    raise RuntimeError(message["error"].get("message", str(message["error"])))
                return message.get("result", {})

    try:
        try:
            result = cdp("Network.getAllCookies")
            cookies = result.get("cookies", [])
        except Exception:
            result = cdp("Storage.getCookies")
            cookies = result.get("cookies", [])
    finally:
        ws.close()

    useful = []
    for cookie in cookies:
        domain = cookie.get("domain", "")
        name = cookie.get("name", "")
        value = cookie.get("value", "")
        if "tinhoctre.vn" in domain and name and value:
            useful.append((name, value))
    if not useful:
        raise RuntimeError("KhÃ´ng tháº¥y cookie tinhoctre.vn trong Edge. HÃ£y Ä‘Äƒng nháº­p TinHocTre admin trong cá»­a sá»• Edge vá»«a má»Ÿ rá»“i thá»­ láº¡i.")

    priority = {"cf_clearance": 0, "aws-waf-token": 1, "csrftoken": 2, "sessionid": 3}
    useful.sort(key=lambda item: (priority.get(item[0], 50), item[0]))
    return "; ".join(f"{name}={value}" for name, value in useful)


def login_tinhoctre_source(account: dict, first_code: str):
    base_url = TARGETS["tinhoctre"]["base_url"]
    try:
        return login_hncode(base_url, account.get("username", ""), account.get("password", ""))
    except Exception as exc:
        raise RuntimeError(str(exc).replace("HNCode", "TinHocTre")) from exc


def login_problem_source(target: str, account: dict, first_code: str):
    base_url = TARGETS[target]["base_url"]
    username = account.get("username", "")
    password = account.get("password", "")
    if target == "tinhoctre":
        return login_tinhoctre_source(account, first_code)
    return login_hncode(base_url, username, password)


def contest_url(base_url: str, key: str) -> str:
    return urljoin(base_url, f"/contest/{key}")


def admin_contest_change_url(session, base_url: str, key: str) -> str | None:
    page = session.get(urljoin(base_url, "/admin/judge/contest/"), params={"q": key})
    if not page.ok:
        return None
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", page.text, re.S):
        plain = html.unescape(re.sub(r"<.*?>", " ", row))
        if re.search(rf"\b{re.escape(key)}\b", plain):
            link = re.search(r'href="(/admin/judge/contest/\d+/change/[^"]*)"', row)
            if link:
                return urljoin(base_url, html.unescape(link.group(1)))
    return None


def admin_problem_id(session, base_url: str, code: str) -> str | None:
    return hncode_service.find_problem_admin_id(session, base_url, code)


def public_contest_problem_codes(session, base_url: str, key: str) -> list[str]:
    return [row["code"] for row in public_contest_problem_rows(session, base_url, key)]


def public_contest_problem_rows(session, base_url: str, key: str) -> list[dict]:
    rows = hncode_service.list_contest_problems(session, base_url, key, default_points="100")
    if rows:
        return rows
    page = session.get(contest_url(base_url, key), timeout=30)
    if not page.ok:
        return []
    codes: list[str] = []
    for code in re.findall(r"(?:/problem/|/contest/" + re.escape(key) + r"/problems/)([A-Za-z0-9_-]+)", page.text):
        if code not in codes:
            codes.append(code)
    return [{"code": code, "title": code, "points": "100", "order": index} for index, code in enumerate(codes, 1)]


def problem_has_test_zip(session, base_url: str, code: str) -> bool:
    page = session.get(test_data_url(base_url, code))
    return page.ok and bool(re.search(r'href=[\"\'][^\"\']+\.zip[\"\']', page.text))


def upload_existing_problem_tests(session, dest: str, code: str, zip_path: Path, cases) -> None:
    base_url = TARGETS[dest]["base_url"]
    if dest == "hnoj":
        tests = GeneratedTests(zip_path, [case.input_file for case in cases], [case.output_file for case in cases])
        upload_tinhoctre_tests(session, base_url, code, tests)
    elif dest == "tinhoctre":
        tests = GeneratedTests(zip_path, [case.input_file for case in cases], [case.output_file for case in cases])
        upload_tinhoctre_tests(session, base_url, code, tests)
    else:
        upload_hncode_tests(session, base_url, code, zip_path, cases)


def selected_values(page: str, name: str) -> list[str]:
    match = re.search(r"<select\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", page, re.S)
    if not match:
        return []
    values = []
    for option in re.finditer(r"<option\b([^>]*)>", match.group(1), re.S):
        attrs = option.group(1)
        if "selected" not in attrs:
            continue
        value = re.search(r"value=[\"']([^\"']*)", attrs)
        if value:
            values.append(html.unescape(value.group(1)))
    return values


def select_option_values(page: str, name: str) -> list[str]:
    match = re.search(r"<select\b[^>]*name=[\"']" + re.escape(name) + r"[\"'][^>]*>(.*?)</select>", page, re.S)
    if not match:
        return []
    values = []
    for option in re.finditer(r"<option\b([^>]*)>", match.group(1), re.S):
        value = re.search(r"value=[\"']([^\"']*)", option.group(1))
        if value:
            values.append(html.unescape(value.group(1)))
    return values


def valid_select_value(page: str, name: str, wanted: str, default: str = "") -> str:
    values = select_option_values(page, name)
    if wanted and wanted in values:
        return wanted
    selected = selected_option(page, name, "")
    if selected:
        return selected
    if default and default in values:
        return default
    return values[0] if values else wanted


class FormDataParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.forms: list[list[tuple[str, str]]] = []
        self.current: list[tuple[str, str]] | None = None
        self.select: dict | None = None
        self.textarea: dict | None = None

    @staticmethod
    def attrs_dict(attrs) -> dict[str, str]:
        return {str(k): "" if v is None else str(v) for k, v in attrs}

    def handle_starttag(self, tag: str, attrs) -> None:
        attrs_map = self.attrs_dict(attrs)
        if tag == "form":
            self.current = []
            return
        if self.current is None:
            return
        if tag == "input":
            name = attrs_map.get("name", "")
            if not name or attrs_map.get("disabled") is not None:
                return
            input_type = attrs_map.get("type", "text").lower()
            if input_type in {"file", "submit", "button", "image", "reset"}:
                return
            if input_type in {"checkbox", "radio"}:
                if "checked" in attrs_map:
                    self.current.append((name, attrs_map.get("value") or "on"))
                return
            self.current.append((name, attrs_map.get("value", "")))
            return
        if tag == "select":
            name = attrs_map.get("name", "")
            self.select = {
                "name": name,
                "multiple": "multiple" in attrs_map,
                "disabled": "disabled" in attrs_map,
                "options": [],
            }
            return
        if tag == "option" and self.select is not None:
            self.select["options"].append(
                {
                    "value": attrs_map.get("value", ""),
                    "selected": "selected" in attrs_map,
                }
            )
            return
        if tag == "textarea":
            name = attrs_map.get("name", "")
            self.textarea = {"name": name, "disabled": "disabled" in attrs_map, "parts": []}

    def handle_data(self, data: str) -> None:
        if self.textarea is not None:
            self.textarea["parts"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "form" and self.current is not None:
            self.forms.append(self.current)
            self.current = None
            return
        if self.current is None:
            return
        if tag == "select" and self.select is not None:
            name = self.select.get("name", "")
            if name and not self.select.get("disabled"):
                options = self.select.get("options") or []
                selected = [option for option in options if option.get("selected")]
                if not selected and not self.select.get("multiple") and options:
                    selected = [options[0]]
                for option in selected:
                    self.current.append((name, str(option.get("value", ""))))
            self.select = None
            return
        if tag == "textarea" and self.textarea is not None:
            name = self.textarea.get("name", "")
            if name and not self.textarea.get("disabled"):
                self.current.append((str(name), "".join(self.textarea.get("parts") or [])))
            self.textarea = None


def collect_contest_form_data(page: str) -> list[tuple[str, str]]:
    parser = FormDataParser()
    parser.feed(page)
    for form in parser.forms:
        if any(name == "contest_problems-TOTAL_FORMS" for name, _value in form):
            return form
    return []


def strip_html_text(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(re.sub(r"<.*?>", " ", value, flags=re.S))).strip()


def compact_form_red_errors(page: str) -> list[str]:
    errors = []
    for block in re.findall(r'<div\b[^>]*class=["\'][^"\']*\bred\b[^"\']*["\'][^>]*>(.*?)</div>', page, re.S | re.I):
        text = strip_html_text(block)
        if text and text not in errors:
            errors.append(text)
    return errors


def extract_hncode_course_slug(value: str) -> str:
    value = (value or "").strip()
    if not value:
        raise RuntimeError("ChÆ°a nháº­p URL hoáº·c mÃ£ course HNCode.")
    match = re.search(r"/course/([^/?#\s]+)", value)
    if match:
        return html.unescape(match.group(1)).strip("/")
    if re.fullmatch(r"[A-Za-z0-9_-]+", value):
        return value
    raise RuntimeError("KhÃ´ng Ä‘á»c Ä‘Æ°á»£c mÃ£ course. HÃ£y nháº­p URL dáº¡ng https://hncode.edu.vn/course/<ma_course>.")


def hncode_course_page_url(course_slug: str, path: str = "") -> str:
    return urljoin(TARGETS["hncode"]["base_url"], f"/course/{course_slug}{path}")


def hncode_course_admin_id(session: requests.Session, course_slug: str) -> str:
    page = session.get(hncode_course_page_url(course_slug, "/edit_lessons"), timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c course {course_slug}: HTTP {page.status_code}")
    match = re.search(r"/admin/judge/course/(\d+)/change/", page.text)
    if not match:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c ID admin cá»§a course {course_slug}.")
    return match.group(1)


def hncode_course_lessons(session: requests.Session, course_slug: str) -> list[dict]:
    page = session.get(hncode_course_page_url(course_slug, "/edit_lessons"), timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c danh sÃ¡ch lesson course {course_slug}: HTTP {page.status_code}")
    rows: list[dict] = []
    seen: set[str] = set()
    for lesson_id, block in re.findall(r'<li\b[^>]*class=["\'][^"\']*\bsortable-item\b[^"\']*["\'][^>]*data-id=["\']?(\d+)["\']?[^>]*>(.*?)</li>', page.text, re.S | re.I):
        if lesson_id in seen:
            continue
        seen.add(lesson_id)
        title_match = re.search(r'<a\b[^>]*href=["\']/course/[^"\']+/lesson/' + re.escape(lesson_id) + r'["\'][^>]*>(.*?)</a>', block, re.S | re.I)
        order_match = re.search(r'<span\b[^>]*class=["\'][^"\']*\bitem-order\b[^"\']*["\'][^>]*>(.*?)</span>', block, re.S | re.I)
        points_match = re.search(r'<span\b[^>]*class=["\'][^"\']*\bitem-points\b[^"\']*["\'][^>]*>(.*?)</span>', block, re.S | re.I)
        rows.append(
            {
                "kind": "lesson",
                "key": lesson_id,
                "title": strip_html_text(title_match.group(1)) if title_match else f"Lesson {lesson_id}",
                "order": strip_html_text(order_match.group(1)).rstrip(".") if order_match else str(len(rows) + 1),
                "points": strip_html_text(points_match.group(1)) if points_match else "",
            }
        )
    return rows


def hncode_course_contests(session: requests.Session, course_slug: str) -> list[dict]:
    page = session.get(hncode_course_page_url(course_slug, "/contests"), timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c danh sÃ¡ch contest course {course_slug}: HTTP {page.status_code}")
    rows: list[dict] = []
    seen: set[str] = set()
    for block in re.findall(r'<li\b[^>]*class=["\'][^"\']*\bsortable-item\b[^"\']*["\'][^>]*>(.*?)</li>', page.text, re.S | re.I):
        link_match = re.search(r'<a\b[^>]*href=["\']/contest/([A-Za-z0-9_-]+)["\'][^>]*>(.*?)</a>', block, re.S | re.I)
        if not link_match:
            continue
        key = html.unescape(link_match.group(1)).strip()
        if not key or key in seen:
            continue
        seen.add(key)
        order_match = re.search(r'<span\b[^>]*class=["\'][^"\']*\bitem-order\b[^"\']*["\'][^>]*>(.*?)</span>', block, re.S | re.I)
        points_match = re.search(r'<input\b[^>]*class=["\'][^"\']*\binline-points-edit\b[^"\']*["\'][^>]*value=["\']?([^"\'> ]*)', block, re.S | re.I)
        rows.append(
            {
                "kind": "contest",
                "key": key,
                "title": strip_html_text(link_match.group(2)) or key,
                "order": strip_html_text(order_match.group(1)).rstrip(".") if order_match else str(len(rows) + 1),
                "points": html.unescape(points_match.group(1)) if points_match else "",
            }
        )
    return rows


def find_hncode_course_lesson_url(session: requests.Session, course_slug: str, title: str) -> str | None:
    wanted = (title or "").strip().casefold()
    if not wanted:
        return None
    matches = [row for row in hncode_course_lessons(session, course_slug) if row.get("title", "").strip().casefold() == wanted]
    if not matches:
        return None
    chosen = matches[-1]
    return hncode_course_page_url(course_slug, f"/lesson/{chosen['key']}")


def find_hncode_course_contest_url(session: requests.Session, course_slug: str, contest_key: str) -> str | None:
    contest_key = (contest_key or "").strip()
    if not contest_key:
        return None
    for row in hncode_course_contests(session, course_slug):
        if row.get("key") == contest_key:
            return urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}")
    return None


def default_course_clone_contest_key(source_key: str, dest_slug: str, suffix: str = "") -> str:
    suffix = (suffix or "").strip()
    if not suffix:
        suffix = "_" + dest_slug
    if not suffix.startswith("_") and not suffix.startswith("-"):
        suffix = "_" + suffix
    raw = f"{source_key}{suffix}".lower()
    raw = re.sub(r"[^a-z0-9_-]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_-")
    return raw or source_key


def collect_form_with_field(page: str, field_name: str) -> list[tuple[str, str]]:
    parser = FormDataParser()
    parser.feed(page)
    for form in parser.forms:
        if any(name == field_name for name, _value in form):
            return form
    return []


def replace_form_fields(data: list[tuple[str, str]], updates: dict[str, str], remove_names: set[str] | None = None) -> list[tuple[str, str]]:
    remove_names = set(remove_names or set()) | set(updates)
    out = [(name, value) for name, value in data if name not in remove_names]
    out.extend((name, value) for name, value in updates.items())
    return out


def lesson_quiz_rows_from_page(page: str, lesson_id: str) -> list[dict]:
    prefix = f"quizzes_{lesson_id}"
    total = int(input_value_from_page(page, f"{prefix}-TOTAL_FORMS", "0") or "0")
    rows: list[dict] = []
    for index in range(total):
        quiz_id = selected_option_value(page, f"{prefix}-{index}-quiz", "") or input_value_from_page(page, f"{prefix}-{index}-quiz", "")
        if not quiz_id:
            continue
        rows.append(
            {
                "id": input_value_from_page(page, f"{prefix}-{index}-id", ""),
                "lesson": input_value_from_page(page, f"{prefix}-{index}-lesson", ""),
                "quiz": quiz_id,
                "points": input_value_from_page(page, f"{prefix}-{index}-points", "0"),
                "max_attempts": input_value_from_page(page, f"{prefix}-{index}-max_attempts", "0"),
                "order": input_value_from_page(page, f"{prefix}-{index}-order", str(index)),
                "is_visible": input_checked(page, f"{prefix}-{index}-is_visible"),
                "delete": input_checked(page, f"{prefix}-{index}-DELETE"),
            }
        )
    return rows


def remove_lesson_item_fields(data: list[tuple[str, str]], lesson_id: str) -> list[tuple[str, str]]:
    prefixes = (f"problems_{lesson_id}-", f"quizzes_{lesson_id}-")
    return [(name, value) for name, value in data if not any(name.startswith(prefix) for prefix in prefixes)]


def append_lesson_quiz_formset(data: list[tuple[str, str]], lesson_id: str, rows: list[dict], initial_forms: int) -> list[tuple[str, str]]:
    prefix = f"quizzes_{lesson_id}"
    out = list(data)
    out.extend(
        [
            (f"{prefix}-TOTAL_FORMS", str(len(rows))),
            (f"{prefix}-INITIAL_FORMS", str(initial_forms)),
            (f"{prefix}-MIN_NUM_FORMS", "0"),
            (f"{prefix}-MAX_NUM_FORMS", "1000"),
        ]
    )
    for index, row in enumerate(rows):
        out.extend(
            [
                (f"{prefix}-{index}-order", str(row.get("order", index))),
                (f"{prefix}-{index}-lesson", str(row.get("lesson", ""))),
                (f"{prefix}-{index}-id", str(row.get("id", ""))),
                (f"{prefix}-{index}-quiz", str(row.get("quiz", ""))),
                (f"{prefix}-{index}-points", str(row.get("points", "0") or "0")),
                (f"{prefix}-{index}-max_attempts", str(row.get("max_attempts", "0") or "0")),
            ]
        )
        if row.get("is_visible"):
            out.append((f"{prefix}-{index}-is_visible", "on"))
        if row.get("delete"):
            out.append((f"{prefix}-{index}-DELETE", "on"))
    return out


def copy_hncode_lesson_items(session: requests.Session, dest_course_slug: str, dest_lesson_id: str, source_page: str, source_lesson_id: str) -> None:
    edit_url = hncode_course_page_url(dest_course_slug, f"/edit_lessons_new/{dest_lesson_id}")
    page = session.get(edit_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a lesson Ä‘Ã­ch {dest_lesson_id}: HTTP {page.status_code}")
    form_data = collect_lesson_form_data(page.text, dest_lesson_id)
    if not form_data:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y form danh sÃ¡ch bÃ i/quiz trong lesson Ä‘Ã­ch {dest_lesson_id}.")
    source_problems = lesson_problem_rows_from_page(source_page, source_lesson_id)
    source_quizzes = lesson_quiz_rows_from_page(source_page, source_lesson_id)
    if not source_problems and not source_quizzes:
        return
    base_data = remove_lesson_item_fields(form_data, dest_lesson_id)
    problem_rows = lesson_problem_rows_from_page(page.text, dest_lesson_id)
    quiz_rows = lesson_quiz_rows_from_page(page.text, dest_lesson_id)
    problem_initial = int(input_value_from_page(page.text, f"problems_{dest_lesson_id}-INITIAL_FORMS", str(len(problem_rows))) or str(len(problem_rows)))
    quiz_initial = int(input_value_from_page(page.text, f"quizzes_{dest_lesson_id}-INITIAL_FORMS", str(len(quiz_rows))) or str(len(quiz_rows)))
    existing_problem_ids = {str(row.get("problem")) for row in problem_rows}
    existing_quiz_ids = {str(row.get("quiz")) for row in quiz_rows}
    for row in source_problems:
        problem_id = str(row.get("problem") or "")
        if not problem_id or problem_id in existing_problem_ids:
            continue
        problem_rows.append(
            {
                "id": "",
                "lesson": str(dest_lesson_id),
                "problem": problem_id,
                "score": str(row.get("score") or "100"),
                "order": str(row.get("order") or len(problem_rows)),
                "delete": False,
            }
        )
        existing_problem_ids.add(problem_id)
    for row in source_quizzes:
        quiz_id = str(row.get("quiz") or "")
        if not quiz_id or quiz_id in existing_quiz_ids:
            continue
        quiz_rows.append(
            {
                "id": "",
                "lesson": str(dest_lesson_id),
                "quiz": quiz_id,
                "points": str(row.get("points") or "0"),
                "max_attempts": str(row.get("max_attempts") or "0"),
                "order": str(row.get("order") or len(quiz_rows)),
                "is_visible": bool(row.get("is_visible")),
                "delete": False,
            }
        )
        existing_quiz_ids.add(quiz_id)
    data = append_lesson_problem_formset(base_data, dest_lesson_id, problem_rows, problem_initial)
    data = append_lesson_quiz_formset(data, dest_lesson_id, quiz_rows, quiz_initial)
    result = session.post(edit_url, data=data, headers={"Referer": edit_url}, allow_redirects=True, timeout=60)
    if not result.ok:
        raise RuntimeError(f"LÆ°u danh sÃ¡ch bÃ i/quiz lesson {dest_lesson_id} lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form sá»­a lesson bÃ¡o lá»—i:\n" + "\n".join(errors))


def clone_hncode_lesson_native(session: requests.Session, source_course: str, lesson_id: str, title: str, dest_course_slug: str, dest_course_id: str) -> str:
    source_edit_url = hncode_course_page_url(source_course, f"/edit_lessons_new/{lesson_id}")
    source_page = session.get(source_edit_url, timeout=30)
    if not source_page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a lesson nguá»“n {lesson_id}: HTTP {source_page.status_code}")
    title = input_value_from_page(source_page.text, "title", title) or title
    points = input_value_from_page(source_page.text, "points", "100") or "100"
    content = textarea_value(source_page.text, "content")
    order = input_value_from_page(source_page.text, "order", "")
    create_url = hncode_course_page_url(dest_course_slug, "/lesson/create")
    page = session.get(create_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form táº¡o lesson á»Ÿ course Ä‘Ã­ch {dest_course_slug}: HTTP {page.status_code}")
    form_data = collect_form_with_field(page.text, "title")
    if not form_data:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y form táº¡o lesson á»Ÿ course Ä‘Ã­ch {dest_course_slug}.")
    data = replace_form_fields(
        form_data,
        {"title": title, "points": points, "content": content, "order": order},
        remove_names={"is_visible"},
    )
    if input_checked(source_page.text, "is_visible"):
        data.append(("is_visible", "on"))
    result = session.post(create_url, data=data, headers={"Referer": create_url}, allow_redirects=True, timeout=60)
    if not result.ok:
        raise RuntimeError(f"Táº¡o lesson {lesson_id} á»Ÿ course Ä‘Ã­ch lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form táº¡o lesson bÃ¡o lá»—i:\n" + "\n".join(errors))
    link = find_hncode_course_lesson_url(session, dest_course_slug, title)
    if not link:
        raise RuntimeError(f"Táº¡o lesson {lesson_id} xong nhung chua thay lesson moi trong course dich {dest_course_slug}.")
    match = re.search(r"/lesson/(\d+)", link)
    if not match:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c ID lesson má»›i tá»« link {link}.")
    copy_hncode_lesson_items(session, dest_course_slug, match.group(1), source_page.text, lesson_id)
    return link


def hncode_contest_edit_problem_rows(page: str) -> list[dict]:
    total = int(input_value_from_page(page, "rows-TOTAL_FORMS", "0") or "0")
    rows: list[dict] = []
    for index in range(total):
        problem_id = selected_option_value(page, f"rows-{index}-problem", "") or input_value_from_page(page, f"rows-{index}-problem", "")
        if not problem_id:
            continue
        rows.append(
            {
                "problem": problem_id,
                "points": input_value_from_page(page, f"rows-{index}-points", "100") or "100",
                "order": input_value_from_page(page, f"rows-{index}-order", str(index + 1)) or str(index + 1),
            }
        )
    return rows


def clone_hncode_contest_native(session: requests.Session, contest_key: str, new_key: str, dest_course_slug: str, dest_course_id: str) -> str:
    source_url = urljoin(TARGETS["hncode"]["base_url"], f"/contest/{contest_key}/edit")
    source_page = session.get(source_url, timeout=30)
    if not source_page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a contest nguá»“n {contest_key}: HTTP {source_page.status_code}")
    name = input_value_from_page(source_page.text, "name", contest_key) or contest_key
    start_time = input_value_from_page(source_page.text, "start_time", "")
    end_time = input_value_from_page(source_page.text, "end_time", "")
    problem_rows = hncode_contest_edit_problem_rows(source_page.text)
    if not problem_rows:
        raise RuntimeError(f"Contest nguá»“n {contest_key} khÃ´ng cÃ³ bÃ i Ä‘á»ƒ clone.")
    add_url = hncode_course_page_url(dest_course_slug, "/add_contest")
    page = session.get(add_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form thÃªm contest vÃ o course Ä‘Ã­ch {dest_course_slug}: HTTP {page.status_code}")
    form_data = collect_form_with_field(page.text, "key")
    if not form_data:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y form thÃªm contest vÃ o course Ä‘Ã­ch {dest_course_slug}.")
    data = replace_form_fields(
        form_data,
        {
            "points": "1000",
            "key": new_key,
            "name": name,
            "start_time": start_time,
            "end_time": end_time,
        },
        remove_names={"problems"},
    )
    for row in sorted(problem_rows, key=lambda item: int(str(item.get("order") or "0")) if str(item.get("order") or "0").isdigit() else 0):
        data.append(("problems", str(row["problem"])))
    result = session.post(add_url, data=data, headers={"Referer": add_url}, allow_redirects=True, timeout=90)
    if not result.ok:
        raise RuntimeError(f"Táº¡o contest {new_key} trong course Ä‘Ã­ch lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form thÃªm contest vÃ o course bÃ¡o lá»—i:\n" + "\n".join(errors))
    link = find_hncode_course_contest_url(session, dest_course_slug, new_key)
    if not link:
        raise RuntimeError(f"Táº¡o contest {new_key} xong nhung chua thay trong course dich {dest_course_slug}.")
    return link


def extract_hncode_contest_key(value: str) -> str:
    value = (value or "").strip()
    if not value:
        raise RuntimeError("ChÆ°a nháº­p URL hoáº·c mÃ£ contest HNCode.")
    match = re.search(r"/contest/([^/?#\s]+)", value)
    if match:
        return html.unescape(match.group(1)).strip("/")
    if re.fullmatch(r"[A-Za-z0-9_-]+", value):
        return value
    raise RuntimeError("KhÃ´ng Ä‘á»c Ä‘Æ°á»£c mÃ£ contest. HÃ£y nháº­p URL dáº¡ng https://oj.hncode.edu.vn/contest/<ma_contest>.")


def contest_lesson_source_from_url(source: str, contest_url_value: str) -> str:
    text = (contest_url_value or "").strip().lower()
    if "hnoj.edu.vn" in text:
        return "hnoj"
    if "hncode.edu.vn" in text or "oj.hncode.edu.vn" in text:
        return "hncode"
    return source if source in {"hncode", "hnoj"} else "hncode"


def extract_hncode_lesson_ref(value: str) -> tuple[str, str]:
    value = (value or "").strip()
    match = re.search(r"/course/([^/?#\s]+)/lesson/(\d+)", value)
    if not match:
        match = re.search(r"/course/([^/?#\s]+)/edit_lessons_new/(\d+)", value)
    if not match:
        raise RuntimeError("KhÃ´ng Ä‘á»c Ä‘Æ°á»£c lesson. HÃ£y nháº­p URL dáº¡ng https://oj.hncode.edu.vn/course/<course>/lesson/<id>.")
    return html.unescape(match.group(1)), match.group(2)


def hncode_lesson_url(course_slug: str, lesson_id: str) -> str:
    return urljoin(TARGETS["hncode"]["base_url"], f"/course/{course_slug}/lesson/{lesson_id}")


def hncode_lesson_edit_url(course_slug: str, lesson_id: str) -> str:
    return urljoin(TARGETS["hncode"]["base_url"], f"/course/{course_slug}/edit_lessons_new/{lesson_id}")


def contest_lesson_score(value: str, default: str = "100") -> str:
    return hncode_service.contest_lesson_score(value, default)


def extract_contest_problem_rows_from_html(page: str, contest_key: str = "", default_points: str = "100") -> list[dict]:
    return hncode_service.extract_contest_problem_rows_from_html(page, contest_key, default_points)


def hncode_contest_problem_rows(session, contest_key: str) -> list[dict]:
    rows = hncode_service.list_contest_problems(session, TARGETS["hncode"]["base_url"], contest_key, default_points="1")
    if not rows:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y bÃ i nÃ o trong contest {contest_key}.")
    return rows


def source_problem_title(session: requests.Session, base_url: str, code: str) -> str:
    page = session.get(urljoin(base_url, f"/problem/{code}/edit"), timeout=30)
    if page.ok:
        title = input_value(page.text, "name", "")
        if title:
            return title
    public = session.get(urljoin(base_url, f"/problem/{code}"), timeout=30)
    if public.ok:
        match = re.search(r"<h1\b[^>]*>(.*?)</h1>", public.text, re.S | re.I)
        if match:
            title = strip_html_text(match.group(1))
            if title:
                return title
    return code


def extract_problem_link_rows_from_html(page: str, default_points: str = "") -> list[dict]:
    return hncode_service.extract_problem_link_rows_from_html(page, default_points)


def admin_problem_code_name_by_id(session: requests.Session, base_url: str, problem_id: str) -> tuple[str, str]:
    return hncode_service.find_problem_code_name_by_id(session, base_url, problem_id)


def hncode_lesson_problem_code_rows(session: requests.Session, course_slug: str, lesson_id: str) -> list[dict]:
    rows = hncode_service.list_lesson_problems(session, TARGETS["hncode"]["base_url"], course_slug, lesson_id)
    if not rows:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y bÃ i nÃ o trong lesson {course_slug}/lesson/{lesson_id}.")
    return rows


def hnoj_contest_problem_rows(session: requests.Session, contest_key: str) -> list[dict]:
    rows = public_contest_problem_rows(session, TARGETS["hnoj"]["base_url"], contest_key)
    if rows:
        return rows
    try:
        info = fetch_contest_info(session, TARGETS["hnoj"]["base_url"], contest_key)
    except Exception:
        codes = public_contest_problem_codes(session, TARGETS["hnoj"]["base_url"], contest_key)
        rows = [
            {
                "code": code,
                "title": source_problem_title(session, TARGETS["hnoj"]["base_url"], code),
                "points": "100",
                "order": index,
            }
            for index, code in enumerate(codes, 1)
        ]
        if rows:
            return rows
        raise
    rows: list[dict] = []
    for index, item in enumerate(info.get("problems", []), 1):
        code = item.get("code", "")
        if not code:
            continue
        order_value = item.get("order") or index
        rows.append(
            {
                "code": code,
                "title": source_problem_title(session, TARGETS["hnoj"]["base_url"], code),
                "points": item.get("points") or "100",
                "order": int(order_value) if str(order_value).isdigit() else index,
            }
        )
    if not rows:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y bÃ i nÃ o trong contest HNOJ {contest_key}.")
    return rows


def collect_lesson_form_data(page: str, lesson_id: str) -> list[tuple[str, str]]:
    parser = FormDataParser()
    parser.feed(page)
    marker = f"problems_{lesson_id}-TOTAL_FORMS"
    for form in parser.forms:
        if any(name == marker for name, _value in form):
            return form
    return []


def remove_lesson_problem_fields(data: list[tuple[str, str]], lesson_id: str) -> list[tuple[str, str]]:
    prefix = f"problems_{lesson_id}-"
    return [(name, value) for name, value in data if not name.startswith(prefix)]


def lesson_problem_rows_from_page(page: str, lesson_id: str) -> list[dict]:
    prefix = f"problems_{lesson_id}"
    total = int(input_value_from_page(page, f"{prefix}-TOTAL_FORMS", "0") or "0")
    rows: list[dict] = []
    for index in range(total):
        problem_id = selected_option_value(page, f"{prefix}-{index}-problem", "") or input_value_from_page(page, f"{prefix}-{index}-problem", "")
        if not problem_id:
            continue
        rows.append(
            {
                "id": input_value_from_page(page, f"{prefix}-{index}-id", ""),
                "lesson": input_value_from_page(page, f"{prefix}-{index}-lesson", ""),
                "problem": problem_id,
                "score": input_value_from_page(page, f"{prefix}-{index}-score", "100"),
                "order": input_value_from_page(page, f"{prefix}-{index}-order", str(index)),
                "delete": input_checked(page, f"{prefix}-{index}-DELETE"),
            }
        )
    return rows


def append_lesson_problem_formset(data: list[tuple[str, str]], lesson_id: str, rows: list[dict], initial_forms: int) -> list[tuple[str, str]]:
    prefix = f"problems_{lesson_id}"
    out = list(data)
    out.extend(
        [
            (f"{prefix}-TOTAL_FORMS", str(len(rows))),
            (f"{prefix}-INITIAL_FORMS", str(initial_forms)),
            (f"{prefix}-MIN_NUM_FORMS", "0"),
            (f"{prefix}-MAX_NUM_FORMS", "1000"),
        ]
    )
    for index, row in enumerate(rows):
        out.extend(
            [
                (f"{prefix}-{index}-order", str(row.get("order", index))),
                (f"{prefix}-{index}-lesson", str(row.get("lesson", ""))),
                (f"{prefix}-{index}-id", str(row.get("id", ""))),
                (f"{prefix}-{index}-problem", str(row.get("problem", ""))),
                (f"{prefix}-{index}-score", str(row.get("score", "100") or "100")),
            ]
        )
        if row.get("delete"):
            out.append((f"{prefix}-{index}-DELETE", "on"))
    return out


def copy_hncode_contest_to_lesson(session, course_slug: str, lesson_id: str, problem_refs: list[dict]) -> str:
    edit_url = hncode_lesson_edit_url(course_slug, lesson_id)
    page = session.get(edit_url, timeout=30)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a lesson HNCode: HTTP {page.status_code}")
    form_data = collect_lesson_form_data(page.text, lesson_id)
    if not form_data:
        raise RuntimeError("KhÃ´ng tÃ¬m tháº¥y form danh sÃ¡ch bÃ i trong lesson HNCode.")
    base_data = remove_lesson_item_fields(form_data, lesson_id)
    current_rows = lesson_problem_rows_from_page(page.text, lesson_id)
    quiz_rows = lesson_quiz_rows_from_page(page.text, lesson_id)
    initial_forms = int(input_value_from_page(page.text, f"problems_{lesson_id}-INITIAL_FORMS", str(len(current_rows))) or str(len(current_rows)))
    quiz_initial = int(input_value_from_page(page.text, f"quizzes_{lesson_id}-INITIAL_FORMS", str(len(quiz_rows))) or str(len(quiz_rows)))
    existing_problem_ids = {str(row.get("problem")) for row in current_rows if row.get("problem")}
    next_order = 1
    if current_rows:
        order_values = [int(str(row.get("order") or "0")) for row in current_rows if str(row.get("order") or "0").isdigit()]
        next_order = (max(order_values) + 1) if order_values else len(current_rows) + 1
    added = 0
    for ref in problem_refs:
        problem_id = str(ref.get("problem_id") or ref.get("id") or "")
        if not problem_id or problem_id in existing_problem_ids:
            continue
        current_rows.append(
            {
                "id": "",
                "lesson": str(lesson_id),
                "problem": problem_id,
                "score": str(ref.get("score") or ref.get("points") or "100"),
                "order": str(next_order + added),
                "delete": False,
            }
        )
        existing_problem_ids.add(problem_id)
        added += 1
    if added == 0:
        return hncode_lesson_url(course_slug, lesson_id)
    data = append_lesson_problem_formset(base_data, lesson_id, current_rows, initial_forms)
    data = append_lesson_quiz_formset(data, lesson_id, quiz_rows, quiz_initial)
    result = session.post(edit_url, data=data, headers={"Referer": edit_url}, allow_redirects=True, timeout=60)
    RUNTIME.mkdir(parents=True, exist_ok=True)
    debug_post = RUNTIME / "hncode_lesson_copy_last_post.html"
    debug_post.write_text(result.text, encoding="utf-8", errors="replace")
    if not result.ok:
        raise RuntimeError(f"LÆ°u lesson lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text) + compact_form_red_errors(result.text)
    if errors:
        raise RuntimeError("Form sá»­a lesson bÃ¡o lá»—i:\n" + "\n".join(errors))
    verify = session.get(edit_url, timeout=30)
    if verify.ok:
        debug_verify = RUNTIME / "hncode_lesson_copy_last_verify.html"
        debug_verify.write_text(verify.text, encoding="utf-8", errors="replace")
        saved_ids = {row["problem"] for row in lesson_problem_rows_from_page(verify.text, lesson_id)}
        missing = [str(ref.get("problem_id") or ref.get("id")) for ref in problem_refs if str(ref.get("problem_id") or ref.get("id")) not in saved_ids]
        if missing:
            raise RuntimeError(
                "Lesson chÆ°a lÆ°u Ä‘á»§ bÃ i: "
                + ", ".join(missing)
                + f"\nÄÃ£ lÆ°u HTML debug: {debug_post} vÃ  {debug_verify}"
            )
    return hncode_lesson_url(course_slug, lesson_id)


def remove_contest_problem_fields(data: list[tuple[str, str]]) -> list[tuple[str, str]]:
    prefixes = (
        "contest_problems-",
        "contest_problems-TOTAL_FORMS",
        "contest_problems-INITIAL_FORMS",
        "contest_problems-MIN_NUM_FORMS",
        "contest_problems-MAX_NUM_FORMS",
    )
    return [(name, value) for name, value in data if not any(name.startswith(prefix) for prefix in prefixes)]


def clean_contest_base_field(name: str, value: str) -> str:
    if name in {"format_config", "problem_label_script", "summary"} and not str(value).strip():
        return ""
    return value


def form_has_field(page: str, name: str) -> bool:
    return bool(re.search(r'\bname=["\']' + re.escape(name) + r'["\']', page))


def select2_field_id(page: str, name: str) -> str:
    match = re.search(r'name="' + re.escape(name) + r'"[^>]*data-field_id="([^"]+)"', page)
    return html.unescape(match.group(1)) if match else ""


def profile_id_for_username(session, base_url: str, page: str, username: str) -> str:
    field_id = select2_field_id(page, "authors")
    if not field_id or not username:
        return ""
    result = session.get(
        urljoin(base_url, "/judge-select2/profile/"),
        params={"field_id": field_id, "term": username, "page": 1},
        headers={"Referer": urljoin(base_url, "/admin/judge/contest/add/"), "X-Requested-With": "XMLHttpRequest"},
    )
    if not result.ok:
        return ""
    try:
        data = result.json()
    except json.JSONDecodeError:
        return ""
    fallback = ""
    for item in data.get("results", []):
        text = str(item.get("text", ""))
        value = str(item.get("id", ""))
        if text == username:
            return value
        if not fallback:
            fallback = value
    return fallback


def split_datetime(value: str) -> tuple[str, str]:
    value = (value or "").strip()
    if " " in value:
        date, time = value.split(" ", 1)
        return date.strip(), time.strip()
    return value, ""


def fetch_contest_info(session, base_url: str, key: str) -> dict:
    change_url = admin_contest_change_url(session, base_url, key)
    if not change_url:
        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y contest {key} trong admin.")
    page = session.get(change_url)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c trang sá»­a contest {key}: HTTP {page.status_code}")
    problem_codes = public_contest_problem_codes(session, base_url, key)
    total = int(input_value(page.text, "contest_problems-TOTAL_FORMS", "0") or "0")
    entries = []
    order_rows = []
    for idx in range(total):
        problem_id = input_value(page.text, f"contest_problems-{idx}-problem")
        if not problem_id:
            continue
        order_rows.append(
            {
                "idx": idx,
                "problem_id": problem_id,
                "points": input_value(page.text, f"contest_problems-{idx}-points", "100") or "100",
                "partial": checkbox_checked(page.text, f"contest_problems-{idx}-partial"),
                "is_pretested": checkbox_checked(page.text, f"contest_problems-{idx}-is_pretested"),
                "max_submissions": input_value(page.text, f"contest_problems-{idx}-max_submissions", ""),
                "order": input_value(page.text, f"contest_problems-{idx}-order", str(idx)) or str(idx),
            }
        )
    for row, code in zip(order_rows, problem_codes):
        row["code"] = code
        entries.append(row)
    if not entries and problem_codes:
        entries = [{"code": code, "points": "100", "partial": True, "is_pretested": False, "max_submissions": "", "order": str(i)} for i, code in enumerate(problem_codes)]
    start_time = f"{input_value(page.text, 'start_time_0', '')} {input_value(page.text, 'start_time_1', '')}".strip()
    end_time = f"{input_value(page.text, 'end_time_0', '')} {input_value(page.text, 'end_time_1', '')}".strip()
    return {
        "key": input_value(page.text, "key", key) or key,
        "name": input_value(page.text, "name", key) or key,
        "description": textarea_value(page.text, "description"),
        "start_time": start_time,
        "end_time": end_time,
        "format_name": selected_option(page.text, "format_name", "vnoj") or "vnoj",
        "scoreboard_visibility": selected_option(page.text, "scoreboard_visibility", "H") or "H",
        "points_precision": input_value(page.text, "points_precision", "3") or "3",
        "is_visible": checkbox_checked(page.text, "is_visible"),
        "is_rated": checkbox_checked(page.text, "is_rated"),
        "is_private": checkbox_checked(page.text, "is_private"),
        "problems": entries,
        "change_url": change_url,
    }


def build_contest_post_data(page: str, info: dict, problem_ids: list[dict], dest: str, author_ids: list[str] | None = None) -> list[tuple[str, str]]:
    start_date, start_clock = split_datetime(info.get("start_time", ""))
    end_date, end_clock = split_datetime(info.get("end_time", ""))
    scoreboard_visibility = valid_select_value(page, "scoreboard_visibility", info.get("scoreboard_visibility") or "", "V")
    format_name = valid_select_value(page, "format_name", info.get("format_name") or "", "vnoj")
    data: list[tuple[str, str]] = [
        ("csrfmiddlewaretoken", csrf_token(page)),
        ("key", info["key"]),
        ("name", info["name"]),
        ("description", statement_for_target(dest, info.get("description", ""))),
        ("scoreboard_visibility", scoreboard_visibility),
        ("points_precision", info.get("points_precision") or "3"),
        ("start_time_0", start_date),
        ("start_time_1", start_clock),
        ("end_time_0", end_date),
        ("end_time_1", end_clock),
        ("time_limit", info.get("time_limit", "")),
        ("format_name", format_name),
        ("format_config", info.get("format_config", "")),
        ("frozen_last_minutes", info.get("frozen_last_minutes", "0") or "0"),
        ("rate_limit", info.get("rate_limit", "")),
        ("freeze_after", info.get("freeze_after", "")),
        ("problem_label_script", info.get("problem_label_script", "")),
        ("rating_floor", info.get("rating_floor", "")),
        ("rating_ceiling", info.get("rating_ceiling", "")),
        ("access_code", info.get("access_code", "")),
        ("ranking_access_code", info.get("ranking_access_code", "")),
        ("scoreboard_cache_timeout", info.get("scoreboard_cache_timeout", "0") or "0"),
        ("summary", info.get("summary", "")),
        ("og_image", ""),
        ("logo_override_image", ""),
        ("contest_problems-TOTAL_FORMS", str(len(problem_ids))),
        ("contest_problems-INITIAL_FORMS", "0"),
        ("contest_problems-MIN_NUM_FORMS", "0"),
        ("contest_problems-MAX_NUM_FORMS", "1000"),
        ("contestannouncement_set-TOTAL_FORMS", "0"),
        ("contestannouncement_set-INITIAL_FORMS", "0"),
        ("contestannouncement_set-MIN_NUM_FORMS", "0"),
        ("contestannouncement_set-MAX_NUM_FORMS", "1000"),
        ("official-TOTAL_FORMS", "0"),
        ("official-INITIAL_FORMS", "0"),
        ("official-MIN_NUM_FORMS", "0"),
        ("official-MAX_NUM_FORMS", "1000"),
        ("_continue", "Save and continue editing"),
    ]
    authors = author_ids if author_ids is not None else selected_values(page, "authors")
    data.extend(("authors", value) for value in authors if value)
    if info.get("is_visible", True):
        data.append(("is_visible", "on"))
    for flag in [
        "use_clarifications",
        "push_announcements",
        "hide_problem_tags",
        "hide_problem_authors",
        "show_short_display",
        "show_submission_list",
        "public_scoreboard",
        "run_pretests_only",
        "rate_all",
    ]:
        if info.get(flag):
            data.append((flag, "on"))
    if info.get("is_rated"):
        data.append(("is_rated", "on"))
    if info.get("is_private"):
        data.append(("is_private", "on"))
    has_quiz = form_has_field(page, "contest_problems-0-quiz") or form_has_field(page, "contest_problems-__prefix__-quiz")
    has_result_hidden = form_has_field(page, "contest_problems-0-is_result_hidden") or form_has_field(page, "contest_problems-__prefix__-is_result_hidden")
    has_show_testcases = form_has_field(page, "contest_problems-0-show_testcases") or form_has_field(page, "contest_problems-__prefix__-show_testcases")
    for idx, problem in enumerate(problem_ids):
        data.extend(
            [
                (f"contest_problems-{idx}-id", ""),
                (f"contest_problems-{idx}-contest", ""),
                (f"contest_problems-{idx}-problem", str(problem["id"])),
                (f"contest_problems-{idx}-points", str(problem.get("points") or "100")),
                (f"contest_problems-{idx}-max_submissions", contest_max_submissions_value(problem, dest)),
                (f"contest_problems-{idx}-hidden_subtasks", str(problem.get("hidden_subtasks") or "")),
                (f"contest_problems-{idx}-output_prefix_override", ""),
                (f"contest_problems-{idx}-order", str(problem.get("order", idx))),
            ]
        )
        if has_quiz:
            data.append((f"contest_problems-{idx}-quiz", str(problem.get("quiz") or "")))
        if problem.get("partial", True):
            data.append((f"contest_problems-{idx}-partial", "on"))
        if problem.get("is_pretested"):
            data.append((f"contest_problems-{idx}-is_pretested", "on"))
        if has_result_hidden and problem.get("is_result_hidden"):
            data.append((f"contest_problems-{idx}-is_result_hidden", "on"))
        if has_show_testcases and problem.get("show_testcases"):
            data.append((f"contest_problems-{idx}-show_testcases", "on"))
    return data


def existing_contest_problem_rows(page: str) -> list[dict]:
    total = int(input_value(page, "contest_problems-TOTAL_FORMS", "0") or "0")
    initial = int(input_value(page, "contest_problems-INITIAL_FORMS", "0") or "0")
    rows: list[dict] = []
    for idx in range(initial):
        problem_id = selected_option(page, f"contest_problems-{idx}-problem", "") or input_value(page, f"contest_problems-{idx}-problem", "")
        if not problem_id:
            continue
        rows.append(
            {
                "form_id": input_value(page, f"contest_problems-{idx}-id", ""),
                "contest": input_value(page, f"contest_problems-{idx}-contest", ""),
                "id": problem_id,
                "quiz": selected_option(page, f"contest_problems-{idx}-quiz", "") or input_value(page, f"contest_problems-{idx}-quiz", ""),
                "points": input_value(page, f"contest_problems-{idx}-points", "100") or "100",
                "partial": checkbox_checked(page, f"contest_problems-{idx}-partial"),
                "is_pretested": checkbox_checked(page, f"contest_problems-{idx}-is_pretested"),
                "is_result_hidden": checkbox_checked(page, f"contest_problems-{idx}-is_result_hidden"),
                "show_testcases": checkbox_checked(page, f"contest_problems-{idx}-show_testcases"),
                "max_submissions": input_value(page, f"contest_problems-{idx}-max_submissions", ""),
                "hidden_subtasks": input_value(page, f"contest_problems-{idx}-hidden_subtasks", ""),
                "output_prefix_override": input_value(page, f"contest_problems-{idx}-output_prefix_override", ""),
                "order": input_value(page, f"contest_problems-{idx}-order", str(idx)) or str(idx),
            }
        )
    return rows


def contest_max_submissions_value(problem: dict, dest: str) -> str:
    value = problem.get("max_submissions")
    if value not in (None, ""):
        return str(value)
    return "0" if dest == "hncode" else ""


def append_contest_problem_fields(data: list[tuple[str, str]], page: str, rows: list[dict], initial_forms: int, dest: str) -> None:
    data.extend(
        [
            ("contest_problems-TOTAL_FORMS", str(len(rows))),
            ("contest_problems-INITIAL_FORMS", str(initial_forms)),
            ("contest_problems-MIN_NUM_FORMS", "0"),
            ("contest_problems-MAX_NUM_FORMS", "1000"),
        ]
    )
    has_quiz = form_has_field(page, "contest_problems-0-quiz") or form_has_field(page, "contest_problems-__prefix__-quiz")
    has_result_hidden = form_has_field(page, "contest_problems-0-is_result_hidden") or form_has_field(page, "contest_problems-__prefix__-is_result_hidden")
    has_show_testcases = form_has_field(page, "contest_problems-0-show_testcases") or form_has_field(page, "contest_problems-__prefix__-show_testcases")
    for idx, problem in enumerate(rows):
        data.extend(
            [
                (f"contest_problems-{idx}-id", str(problem.get("form_id") or "")),
                (f"contest_problems-{idx}-contest", str(problem.get("contest") or "")),
                (f"contest_problems-{idx}-problem", str(problem["id"])),
                (f"contest_problems-{idx}-points", str(problem.get("points") or "100")),
                (f"contest_problems-{idx}-max_submissions", contest_max_submissions_value(problem, dest)),
                (f"contest_problems-{idx}-hidden_subtasks", str(problem.get("hidden_subtasks") or "")),
                (f"contest_problems-{idx}-output_prefix_override", str(problem.get("output_prefix_override") or "")),
                (f"contest_problems-{idx}-order", str(problem.get("order", idx))),
            ]
        )
        if has_quiz:
            data.append((f"contest_problems-{idx}-quiz", str(problem.get("quiz") or "")))
        if problem.get("partial", True):
            data.append((f"contest_problems-{idx}-partial", "on"))
        if problem.get("is_pretested"):
            data.append((f"contest_problems-{idx}-is_pretested", "on"))
        if has_result_hidden and problem.get("is_result_hidden"):
            data.append((f"contest_problems-{idx}-is_result_hidden", "on"))
        if has_show_testcases and problem.get("show_testcases"):
            data.append((f"contest_problems-{idx}-show_testcases", "on"))


def append_problems_to_existing_contest(session, base_url: str, dest: str, change_url: str, problem_ids: list[dict]) -> str:
    page = session.get(change_url)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form sá»­a contest: HTTP {page.status_code}")
    base_data = remove_contest_problem_fields(collect_contest_form_data(page.text))
    if not base_data:
        raise RuntimeError("KhÃ´ng Ä‘á»c Ä‘Æ°á»£c form sá»­a contest Ä‘á»ƒ thÃªm bÃ i.")
    rows = existing_contest_problem_rows(page.text)
    initial_forms = int(input_value(page.text, "contest_problems-INITIAL_FORMS", str(len(rows))) or str(len(rows)))
    existing_ids = {str(row["id"]) for row in rows}
    next_order = max([int(str(row.get("order") or 0)) for row in rows if str(row.get("order") or "").isdigit()] or [-1]) + 1
    added = 0
    for problem in problem_ids:
        problem_id = str(problem["id"])
        if problem_id in existing_ids:
            continue
        item = dict(problem)
        item["form_id"] = ""
        item["contest"] = ""
        item["id"] = problem_id
        item["order"] = str(next_order + added)
        item["max_submissions"] = contest_max_submissions_value(item, dest)
        rows.append(item)
        existing_ids.add(problem_id)
        added += 1
    if not added:
        return change_url
    data = [
        (name, clean_contest_base_field(name, value))
        for name, value in base_data
        if name not in {"_save", "_addanother", "_continue"}
    ]
    append_contest_problem_fields(data, page.text, rows, initial_forms, dest)
    data.append(("_continue", "Save and continue editing"))
    result = session.post(change_url, data=data, headers={"Referer": change_url}, allow_redirects=True)
    if not result.ok:
        raise RuntimeError(f"ThÃªm bÃ i vÃ o contest lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text)
    if errors:
        raise RuntimeError("Form thÃªm bÃ i vÃ o contest bÃ¡o lá»—i:\n" + "\n".join(errors))
    if "/change/" not in result.url:
        raise RuntimeError(f"ThÃªm bÃ i vÃ o contest chÆ°a quay láº¡i trang sá»­a: {result.url}")
    return result.url


def create_contest(session, base_url: str, dest: str, info: dict, problem_ids: list[dict], author_username: str = "") -> str:
    change_url = admin_contest_change_url(session, base_url, info["key"])
    if change_url:
        return append_problems_to_existing_contest(session, base_url, dest, change_url, problem_ids)
    add_url = urljoin(base_url, "/admin/judge/contest/add/")
    page = session.get(add_url)
    if not page.ok:
        raise RuntimeError(f"KhÃ´ng má»Ÿ Ä‘Æ°á»£c form táº¡o contest: HTTP {page.status_code}")
    authors = selected_values(page.text, "authors")
    if not authors:
        author_id = profile_id_for_username(session, base_url, page.text, author_username)
        if author_id:
            authors = [author_id]
    result = session.post(add_url, data=build_contest_post_data(page.text, info, problem_ids, dest, authors), headers={"Referer": add_url}, allow_redirects=True)
    if not result.ok:
        raise RuntimeError(f"Táº¡o contest lá»—i HTTP {result.status_code}")
    errors = form_errors(result.text)
    if errors:
        raise RuntimeError("Form táº¡o contest bÃ¡o lá»—i:\n" + "\n".join(errors))
    if "/change/" not in result.url:
        raise RuntimeError(f"Táº¡o contest chÆ°a redirect vÃ o trang sá»­a: {result.url}")
    return result.url


def contest_transfer_root(prepare_id: str) -> Path:
    if not re.fullmatch(r"[0-9a-f]{32}", prepare_id or ""):
        raise RuntimeError("MÃ£ chuáº©n bá»‹ contest khÃ´ng há»£p lá»‡.")
    return RUNTIME / ("contest_transfer_" + prepare_id)


def save_prepared_contest_transfer(prepare_id: str, state: dict) -> None:
    root = Path(state["root"])
    root.mkdir(parents=True, exist_ok=True)
    disk_state = {
        "root": str(root),
        "source": state["source"],
        "dest": state["dest"],
        "items": state["items"],
    }
    (root / "state.json").write_text(json.dumps(disk_state, ensure_ascii=False, indent=2), encoding="utf-8")


def load_prepared_contest_transfer(prepare_id: str) -> dict | None:
    if prepare_id in prepared_contest_transfers:
        return prepared_contest_transfers[prepare_id]
    root = contest_transfer_root(prepare_id)
    state_file = root / "state.json"
    if not state_file.exists():
        return None
    state = json.loads(state_file.read_text(encoding="utf-8"))
    state["root"] = Path(state["root"])
    prepared_contest_transfers[prepare_id] = state
    return state


@app.post("/api/prepare-contest-transfer")
def api_prepare_contest_transfer():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id")
    source = payload["source"]
    dest = payload["dest"]
    codes = [code.strip() for code in payload.get("codes", []) if code.strip()]
    if not codes:
        return jsonify({"error": "ChÆ°a nháº­p mÃ£ contest cáº§n chuyá»ƒn."}), 400
    if source == dest:
        return jsonify({"error": "Nguá»“n vÃ  Ä‘Ã­ch Ä‘ang trÃ¹ng nhau."}), 400
    try:
        prepare_id = uuid.uuid4().hex
        root = RUNTIME / ("contest_transfer_" + prepare_id)
        root.mkdir(parents=True, exist_ok=True)
        source_account = payload["source_account"]
        source_info = CONTEST_TARGETS[source]
        src = login_hncode(source_info["base_url"], source_account["username"], source_account["password"])
        rows = []
        items = {}
        log_lines = [f"Äá»c contest nguá»“n: {source_info['label']} â†’ {TARGETS[dest]['label']}"]
        progress_update(progress_id, phase="prepare-contest-transfer", done=0, total=len(codes), rows=rows, message="Báº¯t Ä‘áº§u Ä‘á»c contest nguá»“n")
        for index, key in enumerate(codes, 1):
            try:
                info = fetch_contest_info(src, source_info["base_url"], key)
                dest_exists = False
                try:
                    dest_account = payload.get("dest_account", {})
                    dst_probe = login_hncode(TARGETS[dest]["base_url"], dest_account.get("username", ""), dest_account.get("password", ""))
                    dest_exists = bool(admin_contest_change_url(dst_probe, TARGETS[dest]["base_url"], info["key"]))
                    for problem in info["problems"]:
                        pid = admin_problem_id(dst_probe, TARGETS[dest]["base_url"], problem["code"])
                        if pid:
                            problem["status"] = "ÄÃ£ cÃ³ á»Ÿ Ä‘Ã­ch, cÃ³ test" if problem_has_test_zip(dst_probe, TARGETS[dest]["base_url"], problem["code"]) else "ÄÃ£ cÃ³ á»Ÿ Ä‘Ã­ch, thiáº¿u test"
                        else:
                            problem["status"] = "Thiáº¿u á»Ÿ Ä‘Ã­ch"
                except Exception:
                    dest_exists = False
                items[key] = info
                rows.append(
                    {
                        "original_key": key,
                        "key": info["key"],
                        "name": info["name"],
                        "start_time": info["start_time"],
                        "end_time": info["end_time"],
                        "problems": info["problems"],
                        "can_transfer": not dest_exists,
                        "status": "ÄÃ£ tá»“n táº¡i á»Ÿ Ä‘Ã­ch" if dest_exists else "ÄÃ£ Ä‘á»c",
                    }
                )
                log_lines.append(f"- {key}: {info['name']}, {len(info['problems'])} bÃ i")
                if dest_exists:
                    log_lines.append(f"  Contest {info['key']} Ä‘Ã£ tá»“n táº¡i á»Ÿ Ä‘Ã­ch, máº·c Ä‘á»‹nh bá» chá»n Ä‘á»ƒ trÃ¡nh táº¡o trÃ¹ng.")
            except Exception as exc:
                rows.append({"original_key": key, "key": key, "name": "", "start_time": "", "end_time": "", "problems": [], "can_transfer": False, "status": "âœ— Lá»—i Ä‘á»c nguá»“n"})
                log_lines.append(f"âœ— {key}: {exc}")
            progress_update(progress_id, phase="prepare-contest-transfer", done=index, total=len(codes), rows=rows, message=f"{key}: {rows[-1]['status']}")
        state = {"root": root, "source": source, "dest": dest, "items": items}
        prepared_contest_transfers[prepare_id] = state
        save_prepared_contest_transfer(prepare_id, state)
        progress_finish(progress_id, True, f"ÄÃ£ Ä‘á»c {len(rows)}/{len(codes)} contest")
        return jsonify({"prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


@app.post("/api/confirm-contest-transfer")
def api_confirm_contest_transfer():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id")
    prepare_id = payload.get("prepare_id")
    state = load_prepared_contest_transfer(prepare_id) if prepare_id else None
    if not state:
        progress_finish(progress_id, False, "Dá»¯ liá»‡u chuáº©n bá»‹ chuyá»ƒn contest Ä‘Ã£ háº¿t háº¡n")
        return jsonify({"error": "Dá»¯ liá»‡u chuáº©n bá»‹ chuyá»ƒn contest Ä‘Ã£ háº¿t háº¡n. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i."}), 400
    source = payload["source"]
    dest = payload["dest"]
    settings = payload.get("settings", {})
    rows = payload.get("rows", [])
    result_rows = []
    log_lines = [f"Chuyá»ƒn contest: {CONTEST_TARGETS[source]['label']} â†’ {TARGETS[dest]['label']}"]
    try:
        source_account = payload["source_account"]
        dest_account = payload["dest_account"]
        src = login_hncode(CONTEST_TARGETS[source]["base_url"], source_account["username"], source_account["password"])
        dst = login_hncode(TARGETS[dest]["base_url"], dest_account["username"], dest_account["password"])
        root = state["root"]
        language_ids = list(TARGETS[dest]["languages"].values())
        total = len([row for row in rows if row.get("selected")])
        done = 0
        progress_update(progress_id, phase="confirm-contest-transfer", done=done, total=total, rows=result_rows, message="Báº¯t Ä‘áº§u chuyá»ƒn contest")
        for row in rows:
            row = dict(row)
            if not row.get("selected"):
                row["status"] = "Bá» qua"
                result_rows.append(row)
                continue
            try:
                info = dict(state["items"].get(row["original_key"]) or {})
                if not info:
                    raise RuntimeError("ChÆ°a Ä‘á»c Ä‘Æ°á»£c dá»¯ liá»‡u contest nguá»“n")
                info["key"] = row.get("key") or info["key"]
                info["name"] = row.get("name") or info["name"]
                selected_codes = {problem.get("code") for problem in row.get("problems", []) if problem.get("selected")}
                if row.get("problems"):
                    info["problems"] = [problem for problem in info["problems"] if problem["code"] in selected_codes]
                if not info["problems"]:
                    raise RuntimeError("ChÆ°a chá»n bÃ i nÃ o trong contest")
                problem_refs = []
                for problem in info["problems"]:
                    code = problem["code"]
                    pid = admin_problem_id(dst, TARGETS[dest]["base_url"], code)
                    if pid and not settings.get("reuse_existing_problems", True):
                        raise RuntimeError(f"BÃ i {code} Ä‘Ã£ cÃ³ á»Ÿ Ä‘Ã­ch vÃ  tÃ¹y chá»n dÃ¹ng láº¡i bÃ i Ä‘Ã£ cÃ³ Ä‘ang táº¯t")
                    if not pid and not settings.get("create_missing_problems", True):
                        raise RuntimeError(f"BÃ i {code} chÆ°a cÃ³ á»Ÿ Ä‘Ã­ch")
                    if not pid:
                        pinfo, zip_path, cases, _zip_url = fetch_source_problem(src, CONTEST_TARGETS[source]["base_url"], code, root)
                        pinfo.time_limit = pinfo.time_limit or settings.get("time_limit") or "1.0"
                        pinfo.memory_limit = pinfo.memory_limit or settings.get("memory_limit") or "1048576"
                        transfer_row = {"upload_statement": True, "upload_tests": True}
                        if dest == "tinhoctre":
                            upload_transfer_to_tinhoctre(dst, dest, code, pinfo, zip_path, cases, transfer_row, root, language_ids, log_lines)
                        else:
                            upload_transfer_to_dmoj(dst, dest, code, pinfo, zip_path, cases, transfer_row, language_ids, log_lines)
                        pid = admin_problem_id(dst, TARGETS[dest]["base_url"], code)
                    elif settings.get("create_missing_problems", True) and not problem_has_test_zip(dst, TARGETS[dest]["base_url"], code):
                        _pinfo, zip_path, cases, _zip_url = fetch_source_problem(src, CONTEST_TARGETS[source]["base_url"], code, root)
                        upload_existing_problem_tests(dst, dest, code, zip_path, cases)
                        log_lines.append(f"{code}: Ä‘Ã£ bá»• sung test cho bÃ i Ä‘Ã£ cÃ³.")
                    if not pid:
                        raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y ID admin cá»§a bÃ i {code} sau khi chuyá»ƒn")
                    problem_ref = dict(problem)
                    problem_ref["id"] = pid
                    problem_refs.append(problem_ref)
                create_contest(dst, TARGETS[dest]["base_url"], dest, info, problem_refs, dest_account.get("username", ""))
                row["status"] = "âœ“ ThÃ nh cÃ´ng"
                row["link"] = contest_url(TARGETS[dest]["base_url"], info["key"])
                log_lines.append(f"âœ“ {info['key']}: Ä‘Ã£ táº¡o/cáº­p nháº­t contest vá»›i {len(problem_refs)} bÃ i theo Ä‘Ãºng thá»© tá»± gá»­i lÃªn.")
            except ContestAlreadyExists as exc:
                row["status"] = "âœ— Contest Ä‘Ã£ tá»“n táº¡i"
                row["link"] = contest_url(TARGETS[dest]["base_url"], row.get("key") or row.get("original_key"))
                log_lines.append(f"âœ— {row.get('key')}: {exc}. Bá» qua contest nÃ y.")
            except ProblemAlreadyExists:
                row["status"] = "âœ— BÃ i Ä‘Ã£ tá»“n táº¡i nhÆ°ng chÆ°a dÃ¹ng láº¡i Ä‘Æ°á»£c"
                log_lines.append(f"âœ— {row.get('key')}: gáº·p bÃ i Ä‘Ã£ tá»“n táº¡i khi chuyá»ƒn problem, hÃ£y báº­t dÃ¹ng láº¡i bÃ i Ä‘Ã£ cÃ³ hoáº·c kiá»ƒm tra mÃ£ bÃ i.")
            except Exception as exc:
                row["status"] = "âœ— Lá»—i"
                log_lines.append(f"âœ— {row.get('key')}: {exc}")
            result_rows.append(row)
            done += 1
            progress_update(progress_id, phase="confirm-contest-transfer", done=done, total=total, rows=result_rows, message=f"{row.get('key')}: {row.get('status')}")
        ok = all((not row.get("selected")) or row.get("status", "").startswith("âœ“") for row in result_rows)
        progress_finish(progress_id, ok, "ÄÃ£ hoÃ n táº¥t chuyá»ƒn contest")
        return jsonify({"ok": ok, "rows": result_rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/create-contest")
def api_create_contest():
    payload = request.get_json(force=True)
    target = payload["target"]
    key = payload.get("key", "").strip()
    name = payload.get("name", "").strip()
    problems = [code.strip() for code in payload.get("problems", []) if code.strip()]
    if not key or not name or not problems:
        return jsonify({"error": "Cáº§n nháº­p mÃ£ contest, tÃªn contest vÃ  danh sÃ¡ch mÃ£ bÃ i."}), 400
    try:
        account = payload["account"]
        dst = login_upload_target(target, TARGETS[target], account)
        refs = []
        for idx, code in enumerate(problems):
            pid = admin_problem_id(dst, TARGETS[target]["base_url"], code)
            if not pid:
                raise RuntimeError(f"KhÃ´ng tÃ¬m tháº¥y bÃ i {code} á»Ÿ {TARGETS[target]['label']}")
            refs.append({"code": code, "id": pid, "points": "100", "partial": True, "is_pretested": False, "max_submissions": "", "order": str(idx)})
        info = {
            "key": key,
            "name": name,
            "description": "",
            "start_time": payload.get("start_time", ""),
            "end_time": payload.get("end_time", ""),
            "format_name": "vnoj",
            "scoreboard_visibility": "H",
            "points_precision": "3",
            "is_visible": True,
            "is_rated": False,
            "is_private": False,
        }
        create_contest(dst, TARGETS[target]["base_url"], target, info, refs, account.get("username", ""))
        link = contest_url(TARGETS[target]["base_url"], key)
        return jsonify({"ok": True, "log": f"âœ“ ÄÃ£ táº¡o/cáº­p nháº­t contest {key}\nLink: {link}", "link": link})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/prepare-transfer")
def api_prepare_transfer():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id")
    source = payload["source"]
    dest = payload["dest"]
    codes = [code.strip() for code in payload.get("codes", []) if code.strip()]
    if not codes:
        return jsonify({"error": "ChÆ°a nháº­p mÃ£ bÃ i cáº§n chuyá»ƒn."}), 400
    if source == dest:
        return jsonify({"error": "Nguá»“n vÃ  Ä‘Ã­ch Ä‘ang trÃ¹ng nhau."}), 400
    try:
        prepare_id = uuid.uuid4().hex
        root = RUNTIME / ("transfer_" + prepare_id)
        root.mkdir(parents=True, exist_ok=True)
        source_account = payload["source_account"]
        src = login_problem_source(source, source_account, codes[0])
        rows = []
        state_items = {}
        log_lines = [f"Äá»c dá»¯ liá»‡u nguá»“n: {TARGETS[source]['label']} â†’ {TARGETS[dest]['label']}"]
        progress_update(progress_id, phase="prepare-transfer", done=0, total=len(codes), rows=rows, message="Báº¯t Ä‘áº§u Ä‘á»c dá»¯ liá»‡u nguá»“n")
        for index, code in enumerate(codes, 1):
            try:
                info, zip_path, cases, zip_url = fetch_source_problem(src, TARGETS[source]["base_url"], code, root)
                state_items[code] = {"info": info, "zip_path": zip_path, "cases": cases, "zip_url": zip_url}
                source_code = info.code or code
                dest_code = normalize_problem_code_for_target(source_code, dest)
                rows.append(
                    {
                        "original_code": code,
                        "code": dest_code,
                        "name": info.name,
                        "time_limit": info.time_limit or payload.get("settings", {}).get("time_limit") or "1.0",
                        "memory_limit": info.memory_limit or payload.get("settings", {}).get("memory_limit") or "1048576",
                        "source_time_limit": info.time_limit or "1.0",
                        "source_memory_limit": info.memory_limit or "1048576",
                        "test_file": zip_path.name,
                        "test_link": test_data_url(TARGETS[source]["base_url"], code),
                        "test_count": len(cases),
                        "status": "ÄÃ£ Ä‘á»c",
                    }
                )
                log_lines.append(f"- {code}: {info.name}, {len(cases)} test, bá»™ test {test_data_url(TARGETS[source]['base_url'], code)}")
            except Exception as exc:
                rows.append(
                    {
                        "original_code": code,
                        "code": code,
                        "name": "",
                        "time_limit": payload.get("settings", {}).get("time_limit") or "1.0",
                        "memory_limit": payload.get("settings", {}).get("memory_limit") or "1048576",
                        "source_time_limit": "1.0",
                        "source_memory_limit": "1048576",
                        "test_file": "Lá»—i khi Ä‘á»c nguá»“n",
                        "test_link": test_data_url(TARGETS[source]["base_url"], code),
                        "test_count": 0,
                        "status": "âœ— Lá»—i Ä‘á»c nguá»“n",
                    }
                )
                log_lines.append(f"âœ— {code}: {exc}")
            progress_update(progress_id, phase="prepare-transfer", done=index, total=len(codes), rows=rows, message=f"{code}: {rows[-1]['status']}")
        prepared_transfers[prepare_id] = {"root": root, "source": source, "dest": dest, "items": state_items}
        progress_finish(progress_id, True, f"ÄÃ£ Ä‘á»c {len(rows)}/{len(codes)} bÃ i")
        return jsonify({"prepare_id": prepare_id, "rows": rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"error": str(exc)}), 400


@app.post("/api/confirm-transfer")
def api_confirm_transfer():
    payload = request.get_json(force=True)
    progress_id = payload.get("progress_id")
    rows = payload["rows"]
    source = payload["source"]
    dest = payload["dest"]
    settings = payload.get("settings", {})
    log_lines = [
        f"Chuyá»ƒn bÃ i: {TARGETS[source]['label']} â†’ {TARGETS[dest]['label']}",
        "Táº¡o bÃ i Ä‘Ã­ch qua admin form: /admin/judge/problem/add/",
    ]
    result_rows = []
    if source == dest:
        for row in rows:
            row["status"] = "âœ— Nguá»“n vÃ  Ä‘Ã­ch trÃ¹ng nhau"
            result_rows.append(row)
        log_lines.append("Nguá»“n vÃ  Ä‘Ã­ch Ä‘ang trÃ¹ng nhau, khÃ´ng thá»±c hiá»‡n chuyá»ƒn.")
        progress_finish(progress_id, False, "Nguá»“n vÃ  Ä‘Ã­ch Ä‘ang trÃ¹ng nhau")
        return jsonify({"ok": False, "rows": result_rows, "log": "\n".join(log_lines)})

    try:
        dest_account = payload["dest_account"]
        prepare_id = payload.get("prepare_id")
        if not prepare_id or prepare_id not in prepared_transfers:
            return jsonify(
                {
                    "ok": False,
                    "error": "Dá»¯ liá»‡u chuáº©n bá»‹ chuyá»ƒn bÃ i Ä‘Ã£ háº¿t háº¡n hoáº·c server vá»«a khá»Ÿi Ä‘á»™ng láº¡i. HÃ£y báº¥m Chuáº©n bá»‹ dá»¯ liá»‡u láº¡i rá»“i má»›i XÃ¡c nháº­n chuyá»ƒn bÃ i.",
                }
            ), 400
        state = prepared_transfers[prepare_id]
        dst = login_hncode(TARGETS[dest]["base_url"], dest_account["username"], dest_account["password"])
        out_dir = state["root"]
        language_ids = language_ids_for_target(dest, settings.get("languages", []))

        total = len([row for row in rows if row.get("selected")])
        done = 0
        progress_update(progress_id, phase="confirm-transfer", done=done, total=total, rows=result_rows, message="Báº¯t Ä‘áº§u chuyá»ƒn bÃ i")
        for row in rows:
            row = dict(row)
            if not row.get("selected"):
                row["status"] = "Bá» qua"
                result_rows.append(row)
                continue
            try:
                item = state["items"].get(row["original_code"])
                if not item:
                    raise RuntimeError("ChÆ°a Ä‘á»c Ä‘Æ°á»£c dá»¯ liá»‡u nguá»“n cho bÃ i nÃ y")
                info = item["info"]
                zip_path = item["zip_path"]
                cases = item["cases"]
                raw_dest_code = row["code"] or row["original_code"]
                dest_code = normalize_problem_code_for_target(raw_dest_code, dest)
                validate_problem_code_for_target(dest_code, dest)
                if dest_code != raw_dest_code:
                    row["code"] = dest_code
                    log_lines.append(f"{raw_dest_code}: mÃ£ Ä‘Ã­ch {TARGETS[dest]['label']} Ä‘Æ°á»£c Ä‘á»•i thÃ nh {dest_code}")
                if row.get("name"):
                    info.name = row["name"]
                info.time_limit = row.get("time_limit") or settings.get("time_limit") or info.time_limit or "1.0"
                info.memory_limit = row.get("memory_limit") or settings.get("memory_limit") or info.memory_limit or "1048576"
                upload_transfer_to_dmoj(dst, dest, dest_code, info, zip_path, cases, row, language_ids, log_lines)
                row["status"] = "âœ“ ThÃ nh cÃ´ng"
                row["link"] = problem_url(TARGETS[dest]["base_url"], dest_code)
            except ProblemAlreadyExists as exc:
                row["status"] = "âœ— BÃ i Ä‘Ã£ tá»“n táº¡i"
                log_lines.append(f"âœ— {row.get('code')}: {exc}. Bá» qua bÃ i nÃ y vÃ  tiáº¿p tá»¥c cÃ¡c bÃ i khÃ¡c.")
            except Exception as exc:
                row["status"] = "âœ— Lá»—i"
                log_lines.append(f"âœ— {row.get('code')}: {exc}")
            result_rows.append(row)
            done += 1
            progress_update(progress_id, phase="confirm-transfer", done=done, total=total, rows=result_rows, message=f"{row.get('code')}: {row.get('status')}")
        ok = all((not row.get("selected")) or row["status"].startswith("âœ“") for row in result_rows)
        progress_finish(progress_id, ok, "ÄÃ£ hoÃ n táº¥t chuyá»ƒn bÃ i")
        return jsonify({"ok": ok, "rows": result_rows, "log": "\n".join(log_lines)})
    except Exception as exc:
        progress_finish(progress_id, False, str(exc))
        return jsonify({"ok": False, "error": str(exc)}), 400


def upload_transfer_to_dmoj(session, dest: str, dest_code: str, info: ProblemInfo, zip_path: Path, cases, row: dict, language_ids: list[str], log_lines: list[str]) -> None:
    base_url = TARGETS[dest]["base_url"]
    exists = destination_problem_exists(session, base_url, dest_code)
    if exists:
        raise ProblemAlreadyExists(f"MÃ£ bÃ i {dest_code} Ä‘Ã£ tá»“n táº¡i táº¡i {problem_url(base_url, dest_code)}")
    if row.get("upload_statement") and not exists:
        dest_info = problem_info_for_target(info, dest)
        create_hncode_problem(
            session,
            base_url,
            dest_info,
            dest_code=dest_code,
            type_id=TARGETS[dest]["type_id"],
            group_id=TARGETS[dest]["group_id"],
            default_type_id=TARGETS[dest]["type_id"],
            default_group_id=TARGETS[dest]["group_id"],
            public=False,
            allow_all_languages=False,
            allowed_language_ids=language_ids,
        )
        log_lines.append(f"{dest_code}: Ä‘Ã£ táº¡o Ä‘á».")
    else:
        log_lines.append(f"{dest_code}: bá» qua táº¡o Ä‘á».")
    if row.get("upload_tests"):
        if dest == "hnoj":
            tests = GeneratedTests(zip_path, [case.input_file for case in cases], [case.output_file for case in cases])
            upload_tinhoctre_tests(session, base_url, dest_code, tests)
        else:
            upload_hncode_tests(session, base_url, dest_code, zip_path, cases)
        log_lines.append(f"{dest_code}: Ä‘Ã£ upload test.")
    else:
        log_lines.append(f"{dest_code}: khÃ´ng upload test.")


def upload_transfer_to_tinhoctre(session, dest: str, dest_code: str, info: ProblemInfo, zip_path: Path, cases, row: dict, out_dir: Path, language_ids: list[str], log_lines: list[str]) -> None:
    base_url = TARGETS[dest]["base_url"]
    statement = out_dir / f"{dest_code}.md"
    dest_info = problem_info_for_target(info, dest)
    statement.write_text(dest_info.description, encoding="utf-8")
    bundle = ProblemBundle(0, dest_code, info.name, statement, None, zip_path, None)
    tests = GeneratedTests(zip_path, [case.input_file for case in cases], [case.output_file for case in cases])
    exists = tinhoctre_problem_exists(session, base_url, dest_code)
    if exists:
        raise ProblemAlreadyExists(f"MÃ£ bÃ i {dest_code} Ä‘Ã£ tá»“n táº¡i táº¡i {problem_url(base_url, dest_code)}")
    if row.get("upload_statement") and not exists:
        create_tinhoctre_admin_problem(
            session,
            base_url,
            dest_info,
            dest_code=dest_code,
            type_id=TARGETS[dest]["type_id"],
            group_id=TARGETS[dest]["group_id"],
            allowed_language_ids=language_ids,
        )
        log_lines.append(f"{dest_code}: Ä‘Ã£ táº¡o Ä‘á».")
    else:
        log_lines.append(f"{dest_code}: bá» qua táº¡o Ä‘á».")
    if row.get("upload_tests"):
        upload_tinhoctre_tests(session, base_url, dest_code, tests)
        log_lines.append(f"{dest_code}: Ä‘Ã£ upload test.")
    else:
        log_lines.append(f"{dest_code}: khÃ´ng upload test.")


if __name__ == "__main__":
    app.run(
        host=os.getenv("TOOL_OJ_HOST", "127.0.0.1"),
        port=int(os.getenv("TOOL_OJ_PORT", "5050")),
        debug=False,
    )
