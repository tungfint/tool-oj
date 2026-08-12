import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from services import grading as grading_service


class GradingServiceTests(unittest.TestCase):
    def test_read_accounts_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "accounts.csv"
            csv_path.write_text(
                "name,username,password\nNguyễn Văn A,hs_a,pass_a\nTrần Thị B,hs_b,pass_b\n",
                encoding="utf-8",
            )

            accounts = grading_service.read_accounts(csv_path)

        self.assertEqual([row["username"] for row in accounts], ["hs_a", "hs_b"])
        self.assertEqual(accounts[0]["name"], "Nguyễn Văn A")
        self.assertEqual(accounts[1]["index"], 2)

    def test_normalize_student_and_file_key(self):
        self.assertEqual(grading_service.normalize_key("Nguyễn Văn A"), "nguyenvana")
        self.assertEqual(grading_service.normalize_key("  Tran-Thi_B  "), "tranthib")
        self.assertEqual(grading_service.normalize_key("Học sinh 01"), "hocsinh01")

    def test_map_problem_code(self):
        problems = [
            {"code": "bai1_tong", "title": "Tổng", "points": 100, "order": 1},
            {"code": "bai2_dayso", "title": "Dãy số", "points": 100, "order": 2},
            {"code": "rank_a", "title": "Rank A", "points": 100, "order": 3},
        ]

        self.assertEqual(grading_service.map_problem_code("bai1", problems), "bai1_tong")
        self.assertEqual(grading_service.map_problem_code("Bai2-DaySo", problems), "bai2_dayso")
        self.assertEqual(grading_service.map_problem_code("ranka", problems), "rank_a")
        self.assertEqual(grading_service.map_problem_code("unknown", problems), "unknown")

    def test_collect_submission_files_builds_prepare_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            student_a = root / "Nguyen Van A"
            student_b = root / "Tran Thi B"
            student_c = root / "Khong Co TK"
            student_a.mkdir()
            student_b.mkdir()
            student_c.mkdir()
            (student_a / "bai1.cpp").write_text("int main(){return 0;}", encoding="utf-8")
            (student_a / "unknown.cpp").write_text("int main(){return 0;}", encoding="utf-8")
            (student_b / "Bai2-DaySo.py").write_text("print(1)", encoding="utf-8")
            (student_c / "bai1.cpp").write_text("int main(){return 0;}", encoding="utf-8")
            accounts = [
                {"index": 1, "username": "hs_a", "password": "p", "name": "Nguyễn Văn A"},
                {"index": 2, "username": "hs_b", "password": "p", "name": "Trần Thị B"},
            ]
            problems = [
                {"code": "bai1_tong", "title": "Tổng", "points": 100, "order": 1},
                {"code": "bai2_dayso", "title": "Dãy số", "points": 80, "order": 2},
            ]

            rows, warnings = grading_service.collect_submission_files(root, accounts, problems)

        by_file = {row["file"]: row for row in rows}
        self.assertEqual(by_file["bai1.cpp"]["problem"], "bai1_tong")
        self.assertTrue(by_file["bai1.cpp"]["selected"])
        self.assertEqual(by_file["Bai2-DaySo.py"]["problem"], "bai2_dayso")
        self.assertEqual(by_file["unknown.cpp"]["status"], "Không khớp bài trong contest")
        self.assertFalse(by_file["unknown.cpp"]["selected"])
        self.assertTrue(any("Không tìm thấy tài khoản" in warning for warning in warnings))

    def test_merge_requested_rows(self):
        saved = [
            {"original_key": "a::one.cpp", "selected": True, "status": "Đã chuẩn bị"},
            {"original_key": "b::two.cpp", "selected": True, "status": "Đã chuẩn bị"},
        ]
        requested = [{"original_key": "a::one.cpp", "selected": False}]

        rows = grading_service.merge_requested_rows(saved, requested)

        self.assertFalse(rows[0]["selected"])
        self.assertTrue(rows[1]["selected"])

    def test_write_excel_sample(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "bang_diem.xlsx"
            accounts = [{"index": 1, "username": "hs_a", "password": "p", "name": "Nguyễn Văn A"}]
            problems = [{"code": "bai1_tong", "title": "Tổng", "points": 100, "order": 1}]
            rows = [
                {
                    "student": "Nguyễn Văn A",
                    "username": "hs_a",
                    "problem": "bai1_tong",
                    "problem_title": "Tổng",
                    "contest_points": 100,
                    "percent": 80,
                    "score": 80,
                    "relative_path": "Nguyen Van A/bai1.cpp",
                    "status": "✓ Đã chấm",
                    "submission_url": "https://hncode.edu.vn/submission/1",
                    "message": "Accepted",
                }
            ]

            grading_service.write_excel(rows, problems, accounts, output)
            workbook = load_workbook(output)

        self.assertEqual(workbook.sheetnames, ["Bang diem", "Chi tiet nop bai", "Danh sach bai"])
        self.assertEqual(workbook["Bang diem"]["B2"].value, "Nguyễn Văn A")
        self.assertEqual(workbook["Bang diem"]["D2"].value, 80)
        self.assertEqual(workbook["Chi tiet nop bai"]["J2"].hyperlink.target, "https://hncode.edu.vn/submission/1")


if __name__ == "__main__":
    unittest.main()
